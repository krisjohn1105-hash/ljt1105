"""Qube 월간 제출 레포트 3종 생성 — QSMA_Qube_Reporting 템플릿을 채운다.

산출물
    (2) Monthly Profit and Loss Report        → 시트 02_MonthlyPL
    (2) ... (Security Level).xlsx             → 종목별 P&L 별도 파일
    (3) Monthly Available Capacity Report     → 시트 03_MonthlyCapacity
    (4) Firmwide AUM Report                   → 시트 04_FirmAUM

02_MonthlyPL 산식 (2026년 7월 제출본과 동일한 표시 방식)
    Part A 는 수수료·거래세를 빼기 전(gross) 손익, Part B 에서 비용을 따로 계상한다.
        R1 Realized   = 실현손익 (스왑 미결제/정산 + 현물 평균원가법 실현)
        R2 Unrealized = 미실현손익 (스왑 Equity MTM 증감 + 현물 평가 미실현)
        R3 Dividend   = 배당
        E1 Commissions / E2 Stamp Duty : GS 거래 리포트에서 추출 (Part A 에 다시
                                         포함되지 않도록 gross 로 환원해 계상)
        E3 = Int Exp/Inc Equity Swaps (스왑계좌 이자)
        E4 = Int Exp/Inc Broker      (GS PB 계좌 이자)
        (D) = A + B  — 비용을 음수로 넣으므로 합산이 맞다.
              qube_monthly_pnl.py 의 Sub-total 과 일치해야 한다.

전월(Prior Month) 열
    수탁사 Citco 가 계산한 값을 쓴다(공식 장부). 대사파일의 'Citco P&L' 시트에서
    P&S=실현 / OTE Change=미실현 / Dividends=배당 / GL 계정=이자로 매핑한다.
    Citco 는 수수료를 instrument P&L 에 포함하므로 E1/E2 는 0 으로 둔다.

(3)(4) 는 GS 리포트 밖의 수치(펀드별 AUM, 전략 캐파)가 필요하므로
report_inputs.json 으로 받는다. --write-inputs 로 템플릿을 만들 수 있다.

사용법
    python qube_monthly_report.py [월폴더] --template ... --citco ... \
        --inputs report_inputs.json --outdir "...\\06. Monthly Report\\202608_report"
"""

import argparse
import datetime as dt
import json
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd

BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

import qube_pnl as base                                       # noqa: E402
import qube_monthly_pnl as mo                                 # noqa: E402
from qube_pnl import CASH_ACCT, SWAP_ACCT, acct8, collect_files, num, read_report, to_date  # noqa: E402

base.REPORT_PATTERNS.setdefault('syn_mtd_act', '*MTD_SynActivity_303171_*.xls')

WARNINGS = []


def warn(message):
    if message not in WARNINGS:
        WARNINGS.append(message)


# --------------------------------------------------------------------------- #
# 수수료 / 거래세
# --------------------------------------------------------------------------- #
def load_trading_costs(root, month_start, month_end, fx_by_date):
    """{'commission','tax'} — 당월 수수료·거래세 (모두 음수).

    현물은 일별 Custody Transaction 의 Commission / Tax / Fees / Other Fees,
    스왑은 MTD SynActivity 의 Commission Amount / Stamp Duty 에서 뽑는다.
    """
    def rate_for(fx_map, when, ccy):
        """통화별 Base 환산율. USD 는 1, 그 외는 해당일(없으면 직전) 포지션 환율."""
        if not ccy or ccy == 'USD':
            return 1.0
        usable = sorted(d for d in fx_map if d <= when and fx_map[d].get(ccy))
        if usable:
            return fx_map[usable[-1]][ccy]
        later = sorted(d for d in fx_map if fx_map[d].get(ccy))
        if later:
            return fx_map[later[0]][ccy]
        warn(f'{ccy} 환율을 찾지 못해 수수료·거래세를 환산하지 못했습니다.')
        return 1.0

    commission = tax = 0.0

    for bdate, path in sorted(collect_files(root, 'custody_trade').items()):
        if not (month_start <= bdate <= month_end):
            continue
        _, idx, rows, _ = read_report(
            path, ['Account Number', 'Product Type', 'Trade Net Amount'])
        if idx is None:
            continue
        fx_col = next((idx[c] for c in ('FX Rate', 'Fx Rate (Settle to Base)')
                       if c in idx), None)
        for row in rows:
            if acct8(row[idx['Account Number']]) != CASH_ACCT:
                continue
            if str(row[idx['Product Type']]).strip().upper() != 'EQ':
                continue
            fx = (num(row[fx_col]) if fx_col is not None else 1.0) or 1.0
            # Commission 은 부호가 붙어 있으나 Tax/Fees/Other Fees 는 양수 표기라
            # 컬럼을 그대로 더하면 부호가 뒤집힌다. 항등식으로 분해한다.
            #   Trade Net Amount = Principal Amount + Commission - (세금·기타비용)
            net_amt = num(row[idx['Trade Net Amount']])
            principal = num(row[idx['Principal Amount']])
            comm = num(row[idx['Commission']])
            commission += comm * fx
            tax += (net_amt - principal - comm) * fx

    # 스왑: MTD 누적 리포트이므로 당월 범위 내 최신본 1개만 사용
    files = {d: p for d, p in collect_files(root, 'syn_mtd_act').items()
             if month_start <= d <= month_end}
    if files:
        _, idx, rows, dm = read_report(files[max(files)],
                                       ['Account Number', 'Trade Date'])
        if idx is not None:
            for row in rows:
                if acct8(row[idx['Account Number']]) != SWAP_ACCT:
                    continue
                trade_date = to_date(row[idx['Trade Date']], dm)
                if trade_date is None or not (month_start <= trade_date <= month_end):
                    continue
                ccy = str(mo.cell(row, idx, 'Contract CCY')).strip().upper()
                fx = rate_for(fx_by_date, trade_date, ccy)
                commission += num(mo.cell(row, idx, 'Commission Amount')) * fx
                tax += num(mo.cell(row, idx, 'Stamp Duty')) * fx
    else:
        warn('MTD SynActivity 리포트가 없어 스왑 수수료·거래세를 반영하지 못했습니다.')

    return {'commission': commission, 'tax': tax}


# --------------------------------------------------------------------------- #
# 실현 / 미실현 분해
# --------------------------------------------------------------------------- #
def split_cash_realized(root, cash_base, cash_end, month_start, month_end, fx_by_date):
    """현물 실현/미실현 분해 (월초 평가액을 원가로 보는 이동평균법).

    실현 + 미실현 = 평가액증감 + 매매순대금 이 되도록 구성한다(배당 제외).
    """
    qty = {s: v['qty'] for s, v in cash_base.items()}
    cost = {s: v['mv'] for s, v in cash_base.items()}     # 월초 평가액 = 기초원가
    realized = 0.0

    trades = base.load_cash_trades(root, fx_by_date)
    for bdate in sorted(trades):
        if not (month_start <= bdate <= month_end):
            continue
        for rec in trades[bdate]['rows']:
            symbol, q, net = rec['symbol'], rec['qty'], rec['net_base']
            if q > 0:                                   # 매수: 원가 가산
                qty[symbol] = qty.get(symbol, 0.0) + q
                cost[symbol] = cost.get(symbol, 0.0) - net     # net<0 이므로 원가 증가
            elif q < 0:                                 # 매도: 평균원가로 원가 차감
                held = qty.get(symbol, 0.0)
                avg = (cost.get(symbol, 0.0) / held) if held else 0.0
                released = avg * min(-q, held) if held else 0.0
                realized += net - released
                qty[symbol] = held + q
                cost[symbol] = cost.get(symbol, 0.0) - released
            else:                                       # 수량 0 (수수료 조정 등)
                realized += net

    end_mv = sum(v['mv'] for v in cash_end.values())
    end_cost = sum(cost.get(s, 0.0) for s in set(cost) | set(cash_end))
    return realized, end_mv - end_cost


def split_swap_components(swap_base, swap_end, settle):
    """스왑 실현/미실현/배당 분해.

    미실현 = Equity MTM(Base) 증감
    실현   = Unsettled P&L 증감 + 당월 결제된 Equity Leg
    배당   = Dividend Accrued 증감 + 당월 결제된 Dividend Leg
    """
    def total(snap, key):
        return sum(r[key] for r in snap.values())

    unrealized = total(swap_end, 'equity_mtm') - total(swap_base, 'equity_mtm')
    realized = (total(swap_end, 'unsettled') - total(swap_base, 'unsettled')
                + sum(v['equity'] for v in settle.values()))
    dividend = (total(swap_end, 'dividend') - total(swap_base, 'dividend')
                + sum(v['dividend'] for v in settle.values()))
    return realized, unrealized, dividend


# --------------------------------------------------------------------------- #
# Citco 전월 실적
# --------------------------------------------------------------------------- #
def load_citco_prior(path):
    """대사파일의 'Citco P&L' 시트 → 전월 열에 넣을 값 (Citco = 공식 장부)."""
    if not path:
        return None
    try:
        df = pd.read_excel(path, 'Citco P&L').dropna(how='all')
    except Exception as exc:
        warn(f"Citco 대사파일을 읽지 못했습니다 ({path}): {exc}")
        return None

    inst = df[df['Type'].isin(['Equity', 'Equity Swap', 'Cross Rate'])]
    gl = df[df['GL Account'].notna()]

    def gl_sum(*names):
        mask = gl['GL Account Name'].astype(str).str.strip().isin(names)
        return float(gl.loc[mask, 'FA P&L'].sum())

    # Cash Balance 유형은 현금계정 발생분(브로커 현금·미수미지급) — 기타로 모은다
    cash_items = float(df.loc[df['Type'] == 'Cash Balance', 'FA P&L'].sum())
    grand = float(df.loc[df['Type'].isna(), 'FA P&L'].sum())     # 합계 행

    prior = {
        'realized': float(inst['P&S'].sum()),
        'unrealized': float(inst['OTE Change'].sum()),
        'dividend': float(inst['Dividends'].sum()),
        # Citco 는 수수료·제비용을 instrument P&L 안에 넣고 별도 컬럼으로도 보여준다.
        # P&S/OTE 합계와 P&L 의 차이가 정확히 이 금액이므로 Part B 로 옮겨 표시한다.
        'commission': float(inst['Actual Comm'].sum()),
        'tax': float(inst['Total Misc Exp'].sum() + inst['SEC Fees'].sum()),
        'swap_interest': gl_sum('Int Exp Equity Swaps', 'Int Inc Swaps'),
        'broker_interest': gl_sum('Int Exp Broker', 'Int Inc Broker'),
        'other': cash_items,
        'grand_total': grand,
    }
    computed = sum(prior[k] for k in ('realized', 'unrealized', 'dividend',
                                      'commission', 'tax',
                                      'swap_interest', 'broker_interest', 'other'))
    if grand and abs(computed - grand) > 1.0:
        warn(f'Citco 전월 분해 합계 {computed:,.2f} 가 대사파일 합계 행 {grand:,.2f} 와 '
             f'{computed - grand:,.2f} 차이납니다 — 매핑 확인 필요.')
    return prior


# --------------------------------------------------------------------------- #
# 8월 실적 산출
# --------------------------------------------------------------------------- #
def build_current(root, prev_root, cost_basis):
    frame, meta = mo.build_monthly(root, prev_root, {}, cost_basis)
    month_start, month_end = meta['month_start'], meta['month_end']

    cash_snaps, fx_by_date = mo.load_cash_snapshots(root)
    swap_snaps = mo.load_swap_snapshots(root)
    prev_cash, _ = mo.load_cash_snapshots(prev_root) if prev_root else ({}, {})
    prev_swap = mo.load_swap_snapshots(prev_root) if prev_root else {}

    cash_base = prev_cash.get(meta['cash_base_date'],
                              cash_snaps.get(meta['cash_base_date'], {}))
    swap_base = prev_swap.get(meta['swap_base_date'],
                              swap_snaps.get(meta['swap_base_date'], {}))
    settle, _, _ = mo.load_swap_settlements(root, month_start, month_end)

    cash_real, cash_unreal = split_cash_realized(
        root, cash_base, cash_snaps[month_end], month_start, month_end, fx_by_date)
    swap_real, swap_unreal, swap_div = split_swap_components(
        swap_base, swap_snaps[month_end], settle)

    div_total, _ = mo.load_dividends_month(root, month_start, month_end)
    costs = load_trading_costs(root, month_start, month_end, fx_by_date)

    swap_interest = float(frame.loc[frame['Description'] == mo.GL_SWAP_INTEREST,
                                    '$ MTD P&L'].sum())
    broker_interest = float(frame.loc[frame['Description'] == mo.GL_BROKER_INTEREST,
                                      '$ MTD P&L'].sum())
    fx_pnl = float(frame.loc[frame['Type'] == mo.TYPE_FX, '$ MTD P&L'].sum())
    cost_adj = float(frame['_cost_adj'].sum())

    # Part A 는 gross(수수료·거래세 제외 전) 로 표시한다 → 비용을 되돌려 더한다.
    # FX 손익과 IPO 취득원가는 실현손익에 포함시킨다(현금 확정분).
    gross_adjust = -(costs['commission'] + costs['tax'])
    current = {
        # 무상입고(IPO) 취득원가는 매도로 확정된 원가이므로 실현손익에 포함
        'realized': cash_real + swap_real + fx_pnl + cost_adj + gross_adjust,
        'unrealized': cash_unreal + swap_unreal,
        'dividend': swap_div + div_total,
        'commission': costs['commission'],
        'tax': costs['tax'],
        'swap_interest': swap_interest,
        'broker_interest': broker_interest,
        'other': 0.0,
        # 자기참조 검증이 되지 않도록 월별 리포트의 실제 Sub-total 과 대조한다
        'net_check': float(frame['$ MTD P&L'].sum()),
        'cost_adj': cost_adj,
    }
    return current, frame, meta


# --------------------------------------------------------------------------- #
# 템플릿 채우기
# --------------------------------------------------------------------------- #
def put(ws, coord, value):
    """병합셀에 안전하게 쓴다 — 병합범위는 좌상단 셀만 쓸 수 있다."""
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            ws.cell(row=rng.min_row, column=rng.min_col).value = value
            return
    ws[coord] = value


def label_rows(ws, col='B', upto=None):
    """{정규화된 라벨: 행번호} — 셀 주소를 하드코딩하지 않기 위해 라벨로 찾는다.

    템플릿 개정판마다 행 위치가 바뀌므로(v7 은 R4 Interest Income 이 있고
    2026-07 제출본은 없다) 반드시 라벨 기준으로 써야 한다.
    """
    from openpyxl.utils import column_index_from_string
    ci = column_index_from_string(col)
    found = {}
    for r in range(1, (upto or ws.max_row) + 1):
        value = ws.cell(row=r, column=ci).value
        if value is None:
            continue
        key = re.sub(r'\s+', ' ', str(value)).strip().lower()
        if key:
            found.setdefault(key, r)
    return found


def find_row(rows, *needles):
    """라벨 부분일치로 행을 찾는다 (앞에 있는 needle 우선)."""
    for needle in needles:
        target = needle.lower()
        for key, row in rows.items():
            if key.startswith(target):
                return row
        for key, row in rows.items():
            if target in key:
                return row
    return None


# 02_MonthlyPL 행 라벨 → (당월값 키, 새 라벨)
REVENUE_MAP = [
    ('realized p&l', 'realized', None),
    ('unrealized p&l', 'unrealized', None),
    ('dividend income', 'dividend', None),
    ('interest income', None, None),          # 이자는 Part B 에 음수로 계상
]
EXPENSE_MAP = [
    ('brokerage commissions', 'commission', None),
    ('stamp duty', 'tax', None),
    ('interest —', 'swap_interest', 'Interest - Equity Swaps (Int Exp / Inc Equity Swaps)'),
    ('interest', 'swap_interest', 'Interest - Equity Swaps (Int Exp / Inc Equity Swaps)'),
    ('financing', 'broker_interest', 'Interest - Broker (Int Exp / Inc Broker)'),
    ('research charge', None, None),
    ('other', 'other', 'Other'),
    ('short sale', None, None),
    ('clearing and settlement', None, None),
    ('custodial fees', None, None),
]


def fill_monthly_pl(ws, month_end, current, prior):
    rows = label_rows(ws)
    month_start = dt.date(month_end.year, month_end.month, 1)
    put(ws, 'C9', month_start)
    put(ws, 'F9', month_end)
    put(ws, 'C10', 'USD')

    used = set()
    for needle, key, relabel in REVENUE_MAP + EXPENSE_MAP:
        r = find_row(rows, needle)
        if r is None or r in used:
            continue
        used.add(r)
        if relabel:
            put(ws, f'B{r}', relabel)
        put(ws, f'C{r}', round(current.get(key, 0.0), 2) if key else 0)
        put(ws, f'D{r}', round(prior.get(key, 0.0), 2) if (key and prior) else 0)

    # 관리보수 0%, 그리고 (D) 행의 전월 열 수식 부호 통일
    fee = find_row(rows, '(c) management fee')
    if fee:
        put(ws, f'C{fee}', 0)
        put(ws, f'D{fee}', 0)
    net = find_row(rows, '(d) monthly net p&l')
    rev = find_row(rows, '(a) total revenue')
    exp = find_row(rows, '(b) total expenses')
    if net and rev and exp and fee:
        # 비용을 음수로 넣으므로 A+B+C 로 합산해야 한다(템플릿 원본은 A-B-C).
        put(ws, f'C{net}', f'=C{rev}+C{exp}+C{fee}')
        put(ws, f'D{net}', f'=D{rev}+D{exp}+D{fee}')
    if prior:
        # Part A / Part B 두 표의 헤더 모두 표시 (전월은 Citco 기준임을 명시)
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(row=r, column=4).value or '').strip().startswith('Prior Month'):
                put(ws, f'D{r}', 'Prior Month (USD) - Citco')
            if str(ws.cell(row=r, column=2).value or '').strip() == 'Item':
                put(ws, f'D{r}', 'Prior Month (USD) - Citco')
    r1 = find_row(rows, 'realized p&l')
    r2 = find_row(rows, 'unrealized p&l')
    if r1:
        put(ws, f'E{r1}', '스왑 미결제/정산 + 현물 평균원가법 실현 + FX + IPO 취득원가 '
                          '(수수료·거래세는 Part B 로 분리한 gross 표시)')
    if r2:
        put(ws, f'E{r2}', '스왑 Equity MTM 증감 + 현물 미실현')
    return {'A': rev, 'B': exp, 'C': fee, 'D': net}


def fill_workbook(template, out_path, month_end, current, prior, inputs, cum_pnl):
    import openpyxl

    shutil.copyfile(template, out_path)
    wb = openpyxl.load_workbook(out_path)

    pl_name = next((n for n in wb.sheetnames if 'MonthlyPL' in n), None)
    if pl_name is None:
        raise SystemExit('템플릿에 02_MonthlyPL 시트가 없습니다.')
    fill_monthly_pl(wb[pl_name], month_end, current, prior)

    cap_name = next((n for n in wb.sheetnames if 'Capacity' in n), None)
    if cap_name:
        cap = wb[cap_name]
        rows = label_rows(cap, col='A')
        put(cap, 'C9', month_end)
        notional = find_row(rows, 'current notional value')
        if notional:
            # 템플릿 수식은 '50M + 당월 손익' 이라 2개월차부터 틀린다 → 누적으로 직접 기입
            put(cap, f'C{notional}', round(50_000_000 + cum_pnl, 2))
            put(cap, f'D{notional}',
                f'Section 5.1 - initial $50M + 설정 후 누적손익 {cum_pnl:,.2f}')
        unused = find_row(rows, 'unused reserved capacity')
        if unused:
            put(cap, f'C{unused}', inputs['unused_reserved_capacity'])
        brows = label_rows(cap)
        for needle, value in (
                ('total programme capacity', inputs['total_programme_capacity']),
                ('total dunamis aum', inputs['dunamis_strategy_aum']),
                ('unused reserved capacity', inputs['unused_reserved_capacity']),
                ('potential allocations', inputs['third_party_allocations'])):
            r = find_row(brows, needle)
            if r:
                put(cap, f'C{r}', value)

    aum_name = next((n for n in wb.sheetnames if n.endswith('FirmAUM')), None)
    if aum_name:
        aum = wb[aum_name]
        rows = label_rows(aum, col='A')
        head = find_row(rows, 'month-end date')
        if head:
            put(aum, f'C{head}', month_end)
        # 브레이크다운 표는 'No.' 헤더 다음 줄부터 번호가 1..10 으로 붙어 있다
        first = None
        for r in range(1, aum.max_row + 1):
            if str(aum.cell(row=r, column=1).value).strip() == '1':
                first = r
                break
        if first:
            funds = list(inputs['firm_aum_funds'])
            for i in range(10):
                r = first + i
                if i < len(funds):
                    item = funds[i]
                    value = (round(50_000_000 + cum_pnl, 2)
                             if str(item.get('aum')).upper() == 'QSMA' else item['aum'])
                    put(aum, f'B{r}', item['name'])
                    put(aum, f'C{r}', value)
                    put(aum, f'F{r}', item.get('note', ''))
                else:
                    put(aum, f'B{r}', None)
                    put(aum, f'C{r}', None)
                    put(aum, f'F{r}', None)

    wb.save(out_path)
    return out_path


def write_security_level(out_path, frame, month_end):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sec = pd.DataFrame({
        'Type': frame['Type'],
        'Description': frame['Description'],
        'Bloomberg Ticker': frame['Bloomberg Ticker'],
        'Ric': frame['_ric'],
        'ISIN': frame['_isin'],
        'EOM quantity / positions': frame['EOM quantity / positions'],
        'ME price': frame['ME price'],
        'ME FX': frame['ME FX'],
        'Realized + Unrealized (USD)': frame['_equity'] + frame['_dividend'],
        'of which Dividend': frame['_dividend'],
        'Financing (account level)': frame['_financing'],
        '$ MTD P&L': frame['$ MTD P&L'],
    })
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        sec.to_excel(writer, sheet_name='Security Level P&L', index=False)
        ws = writer.book['Security Level P&L']
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor='1F3864')
            cell.font = Font(color='FFFFFF', bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for col in ws.iter_cols(min_row=1):
            header = str(ws.cell(row=1, column=col[0].column).value or '')
            letter = get_column_letter(col[0].column)
            if header in ('Description',):
                ws.column_dimensions[letter].width = 44
            elif header in ('Type', 'Bloomberg Ticker', 'Ric', 'ISIN'):
                ws.column_dimensions[letter].width = 20
            else:
                ws.column_dimensions[letter].width = 16
                for cell in col[1:]:
                    cell.number_format = '#,##0.00;[Red]-#,##0.00'
    return out_path


DEFAULT_INPUTS = {
    '_note': 'GS 리포트로 산출할 수 없는 수치. 매월 확인해 갱신하세요.',
    'total_programme_capacity': 600_000_000,
    'unused_reserved_capacity': 50_000_000,
    'dunamis_strategy_aum': 76_436_526,
    'third_party_allocations': 0,
    'firm_aum_funds': [
        {'name': 'Multi Strategy Fund 1', 'aum': 6_910_741.17, 'note': ''},
        {'name': 'Kosdaq Venture Fund 1', 'aum': 1_311_045.29, 'note': ''},
        {'name': 'Kosdaq Venture Fund 2', 'aum': 749_930.25, 'note': ''},
        {'name': 'Kosdaq Venture Fund 3', 'aum': 1_296_684.61, 'note': ''},
        {'name': 'Block Deal Fund 1', 'aum': 6_168_124.60, 'note': ''},
        {'name': 'Prelude', 'aum': 10_000_000, 'note': 'Fixed at USD 10,000,000'},
        {'name': 'QSMA', 'aum': 'QSMA', 'note': "aum='QSMA' 이면 50M+누적손익 자동계산"},
    ],
}


# --------------------------------------------------------------------------- #
def main():
    parser = argparse.ArgumentParser(description='Qube 월간 레포트 3종 생성')
    parser.add_argument('root', nargs='?', default=str(base.DEFAULT_ROOT))
    parser.add_argument('--prev', default=None)
    parser.add_argument('--cost-basis', default=None)
    parser.add_argument('--template', required=False,
                        default=r"Z:\01.공용\Ops\33. Qube-SMA (QRT)\06. Monthly Report"
                                r"\202607_report\QSMA_Qube_Reporting_2026.07 v2.xlsx",
                        help='기본값은 2026-07 제출본(Qube 가 받은 레이아웃)')
    parser.add_argument('--citco', default=None,
                        help='전월 Citco 대사파일 (Prior Month 열에 사용)')
    parser.add_argument('--prior-cum', type=float, default=None,
                        help='전월말까지의 누적손익 (미지정 시 Citco 합계 사용)')
    parser.add_argument('--inputs', default=None, help='report_inputs.json')
    parser.add_argument('--write-inputs', action='store_true',
                        help='report_inputs.json 템플릿만 생성하고 종료')
    parser.add_argument('--outdir', default=None)
    args = parser.parse_args()

    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, 'reconfigure'):
            stream.reconfigure(encoding='utf-8', errors='replace')

    if args.write_inputs:
        target = Path(args.inputs or 'report_inputs.json')
        target.write_text(json.dumps(DEFAULT_INPUTS, ensure_ascii=False, indent=2),
                          encoding='utf-8')
        print(f'입력 템플릿 생성: {target}')
        return

    root = Path(args.root)
    prev_root = Path(args.prev) if args.prev else mo.guess_prev_root(root)
    cost_basis = mo.load_cost_basis(args.cost_basis)
    inputs = dict(DEFAULT_INPUTS)
    if args.inputs and Path(args.inputs).is_file():
        inputs.update(json.loads(Path(args.inputs).read_text(encoding='utf-8')))

    current, frame, meta = build_current(root, prev_root, cost_basis)
    prior = load_citco_prior(args.citco)
    month_end = meta['month_end']

    net = (current['realized'] + current['unrealized'] + current['dividend']
           + current['commission'] + current['tax']
           + current['swap_interest'] + current['broker_interest'])
    prior_cum = (args.prior_cum if args.prior_cum is not None
                 else (prior['grand_total'] if prior else 0.0))
    cum_pnl = prior_cum + net

    outdir = Path(args.outdir) if args.outdir else root
    outdir.mkdir(parents=True, exist_ok=True)
    stamp = f'{month_end:%B %Y}'
    wb_path = fill_workbook(args.template,
                            outdir / f'QSMA_Qube_Reporting_{month_end:%Y.%m}.xlsx',
                            month_end, current, prior, inputs, cum_pnl)
    sec_path = write_security_level(
        outdir / f'(2) Dunamis Report - Monthly P&L - {stamp} (Security Level).xlsx',
        frame, month_end)

    print(f'대상 월       : {meta["month_start"]} ~ {month_end}')
    print(f'기준선        : 현물 {meta["cash_base_date"]} / 스왑 {meta["swap_base_date"]}')
    print()
    label = {'realized': 'R1 Realized P&L', 'unrealized': 'R2 Unrealized P&L',
             'dividend': 'R3 Dividend Income', 'commission': 'E1 Brokerage Commissions',
             'tax': 'E2 Stamp Duty', 'swap_interest': 'E3 Interest — Equity Swaps',
             'broker_interest': 'E4 Interest — Broker', 'other': 'E6 Other'}
    print(f'{"항목":<32}{"당월(USD)":>18}{"전월 Citco(USD)":>20}')
    for key in ('realized', 'unrealized', 'dividend', 'commission', 'tax',
                'swap_interest', 'broker_interest', 'other'):
        p = prior.get(key, 0.0) if prior else 0.0
        print(f'  {label[key]:<30}{current.get(key, 0.0):>18,.2f}{p:>20,.2f}')
    print(f'  {"(D) Monthly Net P&L":<30}{net:>18,.2f}'
          f'{prior["grand_total"] if prior else 0.0:>20,.2f}')
    print(f'\n  검증: qube_monthly_pnl Sub-total 과 차이 = '
          f'{net - current["net_check"]:,.2f}')
    print(f'  설정 후 누적손익 = {cum_pnl:,.2f}  → Current Notional '
          f'{50_000_000 + cum_pnl:,.2f}')

    print(f'\n생성: {wb_path}')
    print(f'생성: {sec_path}')
    messages = WARNINGS + mo.WARNINGS + base.WARNINGS
    if messages:
        print('\n[확인 필요]')
        for m in messages:
            print(' -', m)


if __name__ == '__main__':
    main()
