"""Qube-RT(GS PBS) 일자별 리포트로부터 일일손익 / 누적손익 엑셀을 생성합니다.

손익 breakdown
    - Cash Equity      : 현물 주식 (커스터디 계좌 08005650, Product Type = EQ)
    - Equity via Swap  : 개별주식 스왑/CFD (스왑 계좌 08005649)
    - Futures via Swap : 선물/지수 스왑 (스왑 계좌 중 Multiplier != 1 또는 기초자산명이 선물/지수)

산출 방법
    Cash Equity 손익 = 평가액(Base) 증감 + 매매 순대금(Base) + 현물 현금배당(ex-date 기준)
    Swap 손익        = Total Mark to Market(Base) 증감 + 당일 결제금액(Total Settlement)
                       (Total MTM = Equity MTM + Total Interest Accrued
                                    + Dividend Accrued + Unsettled P&L -- 데이터로 검증된 항등식)

사용법
    python qube_pnl.py [리포트_루트폴더] [-o 출력파일.xlsx]
"""

import argparse
import datetime as dt
import glob
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
import xlrd
import xlrd.sheet

# GS PBS .xls 일부 파일의 NOTE 레코드가 xlrd 파서를 죽이므로 무력화한다.
xlrd.sheet.Sheet.handle_note = lambda self, data, txos: None

DEFAULT_ROOT = Path(r"Z:\02.펀드\003.매매보고서 대사\Qube-RT\202608")

# 계좌 구분 (Account Number 앞 8자리로 정규화)
CASH_ACCT = '08005650'   # 현물 커스터디 계좌 (080056500)
SWAP_ACCT = '08005649'   # 스왑/CFD 계좌      (080056492)

# 리포트 종류별 파일 패턴 (GS AR 코드로 식별)
REPORT_PATTERNS = {
    'custody_pos':   '*DATA_Custody_Pos_301712_*.xls',   # DATA: Custody Position
    'custody_trade': '*DATA_Custody_Tra_286534_*.xls',   # DATA: Custody Transaction (일별)
    'swap_pnv':      '*Syn_Contract_PnV_303172_*.xls',   # Equity Synthetic Contract P&V
    'swap_settle':   '*MTDSynSettle_302553_*.xls',       # Equity Synthetic MTD Settlement
    'asset_serv':    '*DATA_Asset_Servi_303179_*.xls',   # DATA: Asset Servicing Announcement
}

BUCKETS = ['Cash Equity', 'Equity via Swap', 'Futures via Swap']
FUTURES_RE = re.compile(r'\b(FUT|FUTURE|FUTURES|INDEX|IDX)\b', re.I)

WARNINGS = []


# --------------------------------------------------------------------------- #
# 엑셀 파싱 유틸
# --------------------------------------------------------------------------- #
def num(value):
    """셀 값을 float으로 변환. 숫자가 아니면 0.0."""
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(',', '').strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def to_date(value, datemode=0):
    """엑셀 serial 또는 'YYYYMMDD' / 'Aug 19, 2026' 형태를 date로 변환."""
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        return dt.date(*xlrd.xldate_as_tuple(value, datemode)[:3])
    text = str(value).strip()
    if not text:
        return None
    for fmt in ('%Y%m%d', '%b %d, %Y', '%Y-%m-%d'):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def business_date_of(path):
    """리포트 상단 메타에서 Business Date를 읽는다 (폴더명은 배포일이라 신뢰할 수 없음).

    전송이 잘려 열리지 않는 파일이 섞여 들어오는 경우가 있어(다운로드 실패 등)
    예외를 삼키고 None 을 돌려준 뒤 경고로 남긴다. 여기서 죽으면 월 전체가 멈춘다.
    """
    try:
        book = xlrd.open_workbook(path, on_demand=True)
    except Exception as exc:
        WARNINGS.append(f'파일을 열 수 없어 건너뜁니다 — {os.path.basename(path)} '
                        f'({os.path.getsize(path):,} bytes, {type(exc).__name__}: '
                        f'{str(exc)[:80]}). 다시 내려받으세요.')
        return None
    sheet = book.sheet_by_index(0)
    found = None
    for r in range(min(sheet.nrows, 10)):
        if 'Business Date' in str(sheet.cell_value(r, 0)):
            found = to_date(sheet.cell_value(r, 1), book.datemode)
            break
    book.release_resources()
    return found


def read_report(path, required_cols):
    """리포트 1개를 읽어 (기준일, 컬럼인덱스, 데이터행, datemode)를 반환.

    상단 6줄이 메타(Advisor/Fund/Business Date/...)이고 헤더 행 위치가 리포트마다
    다르므로 required_cols 가 모두 존재하는 행을 헤더로 판정한다.
    """
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)

    bdate = None
    for r in range(min(sheet.nrows, 10)):
        if 'Business Date' in str(sheet.cell_value(r, 0)):
            bdate = to_date(sheet.cell_value(r, 1), book.datemode)
            break

    for r in range(min(sheet.nrows, 20)):
        header = [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
        if all(col in header for col in required_cols):
            index = {}
            for i, name in enumerate(header):
                if name and name not in index:   # 중복 컬럼은 첫 번째만 사용
                    index[name] = i
            rows = [[sheet.cell_value(rr, c) for c in range(sheet.ncols)]
                    for rr in range(r + 1, sheet.nrows)]
            return bdate, index, rows, book.datemode

    return bdate, None, [], book.datemode      # 'NO DATA' 리포트


def next_month_root(root):
    """202608 → 202609 형태의 다음 달 형제 폴더 (없으면 None).

    GS 는 배포일 기준으로 폴더를 나누는데 Consolidated Fund·Interest MTD 처럼
    직전 영업일자를 다음날 생성하는 리포트가 있어, 월 마지막 영업일자 파일이
    항상 익월 첫 폴더에 들어간다. 따라서 익월 폴더까지 훑어야 월말이 채워진다.
    """
    path = Path(root)
    name = path.name
    if not (len(name) == 6 and name.isdigit()):
        return None
    year, month = int(name[:4]), int(name[4:])
    if not 1 <= month <= 12:
        return None
    year, month = (year + 1, 1) if month == 12 else (year, month + 1)
    candidate = path.parent / f'{year}{month:02d}'
    return candidate if candidate.is_dir() else None


def collect_files(root, key):
    """리포트 종류별 {기준일: 파일경로}. 같은 기준일이면 접미사 없는 원본을 우선한다.

    root 는 폴더 하나 또는 여러 개(iterable). 월 폴더 형태면 익월 폴더도 함께 훑는다.
    기준일은 파일 내부 Business Date 를 쓰므로 대상 월 밖의 파일이 섞여도 무해하다.
    """
    def priority(path):
        name = os.path.basename(path)
        if '_APE_' in name:
            return 2
        if '_AP_' in name:
            return 1
        return 0

    if isinstance(root, (str, Path)):
        roots = [root]
        nxt = next_month_root(root)
        if nxt is not None:
            roots.append(nxt)
    else:
        roots = list(root)

    found = {}
    for one in roots:
        for path in sorted(glob.glob(os.path.join(str(one), '*', REPORT_PATTERNS[key]))):
            bdate = business_date_of(path)
            if bdate is None:
                continue
            if bdate not in found or priority(path) < priority(found[bdate]):
                found[bdate] = path
    return found


def acct8(value):
    """Account Number를 앞 8자리 숫자로 정규화 (08005649 / 08005650)."""
    return re.sub(r'\D', '', str(value))[:8]


# --------------------------------------------------------------------------- #
# Cash Equity
# --------------------------------------------------------------------------- #
def load_cash_positions(root):
    """{기준일: {'mv','rows','other_mv','fx'}} — 현물 EQ 포지션.

    'fx'는 통화별 종가 환산율 {통화: Base/Local}. 리포트의 'FX Rate' 컬럼은 소수점
    6자리로 반올림돼 있어(예: 0.000703) 거래대금 환산 시 하루 최대 $190 수준의
    오차가 생기므로, 평가액 Base/Local 합계에서 정밀 환율을 역산해 쓴다.
    """
    result = {}
    for bdate, path in sorted(collect_files(root, 'custody_pos').items()):
        _, idx, rows, _ = read_report(
            path, ['Account Number', 'Product Type', 'Ending Market Value - Base'])
        if idx is None:
            result[bdate] = {'mv': 0.0, 'rows': {}, 'other_mv': 0.0, 'fx': {}}
            continue

        total, other, detail = 0.0, 0.0, {}
        fx_base, fx_local = defaultdict(float), defaultdict(float)
        for row in rows:
            symbol = str(row[idx['Symbol']]).strip() if 'Symbol' in idx else ''
            if acct8(row[idx['Account Number']]) != CASH_ACCT or not symbol:
                continue
            product = str(row[idx['Product Type']]).strip().upper()
            mv = num(row[idx['Ending Market Value - Base']])
            if product == 'EQ':
                total += mv
                # 같은 종목이 Account Type(01/06) 별로 두 줄로 나오는 날이 있어(결제 대기
                # 이관) 종목별 집계는 합산해야 한다. 덮어쓰면 상계분이 사라져 허위 손익이 생긴다.
                bag = detail.setdefault(symbol, {
                    'symbol': symbol,
                    'name': str(row[idx['Product Description']]).strip(),
                    'ric': str(row[idx['RIC Code']]).strip() if 'RIC Code' in idx else '',
                    'qty': 0.0,
                    'mv': 0.0,
                })
                bag['qty'] += num(row[idx['Trade Date Quantity']])
                bag['mv'] += mv
                currency = str(row[idx['Ending Local Currency']]).strip().upper()
                local_mv = num(row[idx['Ending Market Value - Local']])
                if currency and local_mv:
                    fx_base[currency] += mv
                    fx_local[currency] += local_mv
            elif product != 'CA':
                other += mv      # 현물 계좌에 주식/현금 외 상품이 생기면 경고용으로 집계
        if other:
            WARNINGS.append(
                f'{bdate}: 현물계좌에 EQ/CA 이외 상품 평가액 {other:,.2f} 존재 (손익 미분류)')
        result[bdate] = {
            'mv': total, 'rows': detail, 'other_mv': other,
            'fx': {c: fx_base[c] / fx_local[c] for c in fx_local if fx_local[c]},
        }
    return result


def load_cash_trades(root, fx_by_date=None):
    """{기준일: {'net': 매매순대금(Base), 'rows': [...], 'transfers': [...]}}

    Trade Net Amount는 현금 관점 부호(매수=음수)라서 평가액 증감에 그대로 더하면
    매매 자체로는 손익이 생기지 않는다.

    환산율 우선순위 (정밀도 순)
        1) fx_by_date : 같은 날 포지션 리포트에서 역산한 통화별 종가 환율
                        — 평가액 레그와 동일 환율을 써야 허위 손익이 안 생긴다
        2) 1 / Inverse FX Rate : 거래 리포트의 역환율 (예: 1423.11)
        3) FX Rate     : 6자리 반올림 값 (예: 0.000703) — 최후 수단
    """
    fx_by_date = fx_by_date or {}
    result = {}
    for bdate, path in sorted(collect_files(root, 'custody_trade').items()):
        _, idx, rows, datemode = read_report(
            path, ['Account Number', 'Product Type', 'Trade Net Amount'])
        if idx is None:
            result[bdate] = {'net': 0.0, 'rows': [], 'transfers': []}
            continue

        fx_col = next((idx[c] for c in ('FX Rate', 'Fx Rate (Settle to Base)') if c in idx), None)
        inv_col = idx.get('Inverse FX Rate')
        day_fx = fx_by_date.get(bdate, {})
        net_total, detail, transfers = 0.0, [], []
        for row in rows:
            if acct8(row[idx['Account Number']]) != CASH_ACCT:
                continue
            if str(row[idx['Product Type']]).strip().upper() != 'EQ':
                continue
            currency = str(row[idx['Settle Currency']]).strip().upper()
            base_ccy = str(row[idx['Base Currency']]).strip().upper()
            inverse = num(row[inv_col]) if inv_col is not None else 0.0
            if currency == base_ccy:
                fx = 1.0
            elif currency in day_fx:
                fx = day_fx[currency]
            elif inverse:
                fx = 1.0 / inverse
            else:
                fx = num(row[fx_col]) if fx_col is not None else 1.0
            if fx == 0:
                fx = 1.0
            net_base = num(row[idx['Trade Net Amount']]) * fx
            net_total += net_base
            mnemonic = str(row[idx['Transaction Mnemonic']]).strip()
            record = {
                'symbol': str(row[idx['Symbol']]).strip(),
                'name': str(row[idx['Product Description']]).strip(),
                'ric': str(row[idx['RIC Code']]).strip() if 'RIC Code' in idx else '',
                'bbg': (str(row[idx['Bloomberg Ticker']]).strip()
                        if 'Bloomberg Ticker' in idx else ''),
                'ccy': str(row[idx['Issue Currency']]).strip().upper()
                       if 'Issue Currency' in idx else '',
                'mnemonic': mnemonic,
                'qty': num(row[idx['Trade Quantity']]),
                'net_base': net_base,
                'trade_date': to_date(row[idx['Trade Date']], datemode),
            }
            detail.append(record)
            # 대가 없는 수량 이동(무상입출고 FREC/FRDL, 대체입출고 등)은 매입원가가
            # 잡히지 않아 평가액/매도대금이 그대로 손익으로 인식된다. 별도 표시 대상.
            if record['qty'] and net_base == 0.0:
                transfers.append(record)
        if transfers:
            names = ', '.join(f"{t['symbol']} {t['qty']:,.0f}주({t['mnemonic']})"
                              for t in transfers)
            WARNINGS.append(
                f'{bdate}: 대가 없는 현물 수량이동 {names} — 매입원가가 없어 '
                'Cash Equity 손익이 과대/과소 계상될 수 있음 (손익상세 시트의 '
                '"CE 무상이전 손익영향" 확인)')
        result[bdate] = {'net': net_total, 'rows': detail, 'transfers': transfers}
    return result


def load_physical_dividends(root):
    """현물 현금배당을 ex-date 기준으로 집계. Announcement Id로 중복 제거.

    반환 (by_date, detail)
        by_date : {ex_date: 합계}
        detail  : {ex_date: [(종목명, 금액, RIC)]}  — RIC 로 종목별 손익에 귀속시킨다
    """
    events = {}
    for _, path in sorted(collect_files(root, 'asset_serv').items()):
        _, idx, rows, datemode = read_report(
            path, ['Account Number', 'Event Description', 'Net Amount (Base)'])
        if idx is None:
            continue
        for row in rows:
            if acct8(row[idx['Account Number']]) != CASH_ACCT:
                continue
            if 'DIVIDEND' not in str(row[idx['Event Description']]).upper():
                continue
            key = str(row[idx['Announcement Id']]).strip()
            ex_date = to_date(row[idx['Ex Date']], datemode)
            if not key or ex_date is None:
                continue
            events[key] = (ex_date, num(row[idx['Net Amount (Base)']]),
                           str(row[idx['Underlyer Description']]).strip(),
                           str(row[idx['RIC']]).strip() if 'RIC' in idx else '')

    by_date, detail = defaultdict(float), defaultdict(list)
    for ex_date, amount, name, ric in events.values():
        by_date[ex_date] += amount
        detail[ex_date].append((name, amount, ric))
    return dict(by_date), dict(detail)


# --------------------------------------------------------------------------- #
# Swap (Equity / Futures)
# --------------------------------------------------------------------------- #
def is_futures_swap(underlyer_name, multiplier):
    """스왑 기초자산이 선물/지수인지 판정."""
    if multiplier and multiplier != 1.0:
        return True
    return bool(FUTURES_RE.search(str(underlyer_name)))


def load_swap_positions(root):
    """{기준일: {bucket: {'mtm','equity','financing','dividend','settled_pnl','rows'}}}"""
    result = {}
    for bdate, path in sorted(collect_files(root, 'swap_pnv').items()):
        _, idx, rows, _ = read_report(path, ['Contract ID', 'Total Mark to Market (Base)'])
        day = {b: {'mtm': 0.0, 'equity': 0.0, 'financing': 0.0,
                   'dividend': 0.0, 'settled_pnl': 0.0, 'rows': []}
               for b in BUCKETS[1:]}
        if idx is None:
            result[bdate] = day
            continue

        for row in rows:
            contract = str(row[idx['Contract ID']]).strip()
            if not contract or acct8(row[idx['Account Number']]) != SWAP_ACCT:
                continue
            name = str(row[idx['Underlyer Name']]).strip()
            bucket = ('Futures via Swap'
                      if is_futures_swap(name, num(row[idx['Multiplier']]))
                      else 'Equity via Swap')
            # Unsettled P&L은 Contract CCY 표기이므로 Base로 환산
            fx = num(row[idx['FX Contract to Base']]) or 1.0
            values = {
                'mtm':         num(row[idx['Total Mark to Market (Base)']]),
                'equity':      num(row[idx['Equity Mark to Market (Base)']]),
                'financing':   num(row[idx['Total Interest Accrued (Base)']]),
                'dividend':    num(row[idx['Dividend Accrued (Base)']]),
                'settled_pnl': num(row[idx['Unsettled P&L']]) * fx,
            }
            bag = day[bucket]
            for key, value in values.items():
                bag[key] += value
            bag['rows'].append({
                'contract': contract,
                'name': name,
                'side': str(row[idx['Long/Short']]).strip(),
                'qty': num(row[idx['Traded Quantity']]),
                **values,
            })
        result[bdate] = day
    return result


def load_swap_settlements(root, as_of=None):
    """MTD Settlement(월 누적) 중 as_of 이하의 최신 파일 1개로 결제금액을 집계.

    폴더에 익월분 리포트가 섞여 있을 수 있다(예: 8월 폴더의 '0901' 하위폴더에
    BD 9/1 파일). MTD 리포트는 월이 바뀌면 초기화되므로 무턱대고 max() 를 쓰면
    당월 결제가 통째로 사라진다. 반드시 as_of 이하에서 고른다.

    반환 (by_bucket, by_contract)
        by_bucket   : {결제일: {bucket: 금액}}   — 버킷 합계용
        by_contract : {결제일: {contract: 금액}} — 종목별 배분용
    두 곳에서 같은 값을 쓰지 않으면 결제일에 버킷 합계와 종목별 합계가 어긋난다.
    """
    files = collect_files(root, 'swap_settle')
    if as_of is not None:
        files = {d: p for d, p in files.items() if d <= as_of}
    if not files:
        return {}, {}
    _, idx, rows, datemode = read_report(files[max(files)],
                                         ['Payment Date', 'Total Settlement'])
    if idx is None:
        return {}, {}

    by_bucket = defaultdict(lambda: defaultdict(float))
    by_contract = defaultdict(lambda: defaultdict(float))
    for row in rows:
        pay_date = to_date(row[idx['Payment Date']], datemode)
        if pay_date is None:
            continue
        name = str(row[idx['Underlyer Desc']]).strip() if 'Underlyer Desc' in idx else ''
        bucket = 'Futures via Swap' if is_futures_swap(name, 0.0) else 'Equity via Swap'
        amount = num(row[idx['Total Settlement']])
        by_bucket[pay_date][bucket] += amount
        by_contract[pay_date][str(row[idx['Contract ID']]).strip()] += amount
    return ({d: dict(v) for d, v in by_bucket.items()},
            {d: dict(v) for d, v in by_contract.items()})


# --------------------------------------------------------------------------- #
# 손익 산출
# --------------------------------------------------------------------------- #
def build_pnl(root):
    cash_pos = load_cash_positions(root)
    cash_trd = load_cash_trades(root, {d: v['fx'] for d, v in cash_pos.items()})
    div_by_date, div_detail = load_physical_dividends(root)
    swap_pos = load_swap_positions(root)
    dates = sorted(set(cash_pos) & set(swap_pos))

    # 월 폴더(YYYYMM) 안에 익월 초 리포트가 섞여 들어오는 경우가 있어 대상 월로 제한한다.
    name = Path(root).name
    if len(name) == 6 and name.isdigit() and dates:
        year, month = int(name[:4]), int(name[4:])
        if 1 <= month <= 12:
            dropped = [d for d in dates if (d.year, d.month) != (year, month)]
            if dropped:
                WARNINGS.append(
                    f'{name} 폴더에 대상 월 밖의 기준일이 섞여 있어 제외했습니다: '
                    + ', '.join(str(d) for d in dropped))
            dates = [d for d in dates if (d.year, d.month) == (year, month)]
    settle, settle_contract = load_swap_settlements(root, max(dates) if dates else None)
    if not dates:
        missing = [f'{key}({REPORT_PATTERNS[key]})'
                   for key, found in (('custody_pos', cash_pos), ('swap_pnv', swap_pos))
                   if not found]
        raise SystemExit(
            '기준일이 겹치는 포지션 리포트를 찾지 못했습니다.\n'
            f'  대상 폴더 : {root}\n'
            f'  현물 포지션 리포트 : {len(cash_pos)}일\n'
            f'  스왑 P&V 리포트    : {len(swap_pos)}일\n'
            + (f'  누락 : {", ".join(missing)}\n' if missing else '')
            + '  ※ 월에 따라 GS 리포트 구성(AR 코드)이 다를 수 있습니다. '
              'REPORT_PATTERNS 를 해당 월 파일명에 맞게 조정하세요.')

    # 한쪽 리포트만 있는 날은 손익을 낼 수 없어 제외된다 → 조용히 빠지지 않게 경고
    for missing in sorted((set(cash_pos) | set(swap_pos)) - set(dates)):
        side = '스왑 P&V' if missing in cash_pos else '현물 포지션'
        WARNINGS.append(f'{missing}: {side} 리포트가 없어 해당일을 손익 산출에서 제외함')

    # 거래 리포트만 없는 날은 매매순대금이 0으로 잡혀 손익이 왜곡된다
    for date in dates:
        if date not in cash_trd:
            WARNINGS.append(
                f'{date}: 현물 거래(Custody Transaction) 리포트가 없어 매매순대금을 '
                '0으로 처리함 — Cash Equity 손익에 매수/매도 대금이 손익으로 섞임')

    # 중간 영업일이 빠지면 다음 날 손익이 여러 날 합계로 계상된다.
    # 주말은 정상이므로 사이에 낀 평일(월~금)만 공백으로 본다. 공휴일은 오탐이 될 수 있음.
    for i in range(1, len(dates)):
        missing = []
        day = dates[i - 1] + dt.timedelta(days=1)
        while day < dates[i]:
            if day.weekday() < 5:
                missing.append(day)
            day += dt.timedelta(days=1)
        if missing:
            WARNINGS.append(
                f'{dates[i]} 일간손익에 {", ".join(str(d) for d in missing)} 손익이 합산됨 '
                '(해당일 리포트 없음 — 공휴일이면 정상)')

    summary, detail, instrument = [], [], []
    cumulative = {b: 0.0 for b in BUCKETS}

    for i, date in enumerate(dates):
        prev = dates[i - 1] if i else None
        first = prev is None

        # ---------------- Cash Equity ----------------
        mv = cash_pos[date]['mv']
        mv_prev = cash_pos[prev]['mv'] if prev else 0.0
        trade_net = cash_trd.get(date, {}).get('net', 0.0)
        dividend = div_by_date.get(date, 0.0)
        cash_pnl = None if first else (mv - mv_prev) + trade_net + dividend

        # ---------------- Swap ----------------
        swap_pnl, swap_parts = {}, {}
        for bucket in BUCKETS[1:]:
            now = swap_pos[date][bucket]
            was = swap_pos[prev][bucket] if prev else None
            paid = settle.get(date, {}).get(bucket, 0.0)
            if first:
                swap_pnl[bucket] = None
                swap_parts[bucket] = {k: None for k in ('equity', 'financing', 'dividend')}
            else:
                swap_pnl[bucket] = (now['mtm'] - was['mtm']) + paid
                swap_parts[bucket] = {
                    'equity': ((now['equity'] + now['settled_pnl'])
                               - (was['equity'] + was['settled_pnl']) + paid),
                    'financing': now['financing'] - was['financing'],
                    'dividend': now['dividend'] - was['dividend'],
                }

        daily = {'Cash Equity': cash_pnl, **swap_pnl}
        for bucket in BUCKETS:
            if daily[bucket] is not None:
                cumulative[bucket] += daily[bucket]
        total_daily = None if first else sum(v for v in daily.values() if v is not None)

        summary.append({
            '기준일': date,
            'Cash Equity 일간손익': daily['Cash Equity'],
            'Equity via Swap 일간손익': daily['Equity via Swap'],
            'Futures via Swap 일간손익': daily['Futures via Swap'],
            '일간손익 합계': total_daily,
            'Cash Equity 누적손익': None if first else cumulative['Cash Equity'],
            'Equity via Swap 누적손익': None if first else cumulative['Equity via Swap'],
            'Futures via Swap 누적손익': None if first else cumulative['Futures via Swap'],
            '누적손익 합계': None if first else sum(cumulative.values()),
            'Cash Equity 평가액': mv,
            'Equity via Swap MTM': swap_pos[date]['Equity via Swap']['mtm'],
            'Futures via Swap MTM': swap_pos[date]['Futures via Swap']['mtm'],
            '비고': '기준일(직전 영업일 자료 없음 → 손익 미산출)' if first else '',
        })

        # ---------------- 종목별 손익 (Cash Equity) ----------------
        prev_rows = cash_pos[prev]['rows'] if prev else {}
        cur_rows = cash_pos[date]['rows']
        trades = defaultdict(float)
        for record in cash_trd.get(date, {}).get('rows', []):
            trades[record['symbol']] += record['net_base']

        # 당일 ex-date 배당을 RIC 로 종목에 귀속. 못 찾으면 별도 행으로 남긴다.
        div_by_symbol, div_unmatched = defaultdict(float), []
        ric_to_symbol = {}
        for source in (cur_rows, prev_rows):
            for sym, rec in source.items():
                if rec.get('ric'):
                    ric_to_symbol.setdefault(rec['ric'], sym)
        for name, amount, ric in div_detail.get(date, []):
            symbol = ric_to_symbol.get(ric)
            if symbol:
                div_by_symbol[symbol] += amount
            else:
                div_unmatched.append((name, amount))

        cash_symbol_pnl = {}
        if not first:
            for symbol in sorted(set(prev_rows) | set(cur_rows) | set(trades) | set(div_by_symbol)):
                was, now = prev_rows.get(symbol, {}), cur_rows.get(symbol, {})
                pnl = ((now.get('mv', 0.0) - was.get('mv', 0.0)) + trades.get(symbol, 0.0)
                       + div_by_symbol.get(symbol, 0.0))
                cash_symbol_pnl[symbol] = pnl
                if abs(pnl) < 0.005 and not now.get('qty') and not was.get('qty'):
                    continue
                instrument.append({
                    '기준일': date, 'Breakdown': 'Cash Equity',
                    '종목코드': symbol, '종목명': now.get('name') or was.get('name', ''),
                    '전일수량': was.get('qty', 0.0), '당일수량': now.get('qty', 0.0),
                    '전일평가액': was.get('mv', 0.0), '당일평가액': now.get('mv', 0.0),
                    '매매순대금': trades.get(symbol, 0.0), '일간손익': pnl,
                })
            for name, amount in div_unmatched:
                instrument.append({
                    '기준일': date, 'Breakdown': 'Cash Equity',
                    '종목코드': '(배당-미매칭)', '종목명': name,
                    '전일수량': 0.0, '당일수량': 0.0,
                    '전일평가액': 0.0, '당일평가액': 0.0,
                    '매매순대금': 0.0, '일간손익': amount,
                })

        # ---------------- 구성요소 상세 ----------------
        transfers = cash_trd.get(date, {}).get('transfers', [])
        transfer_symbols = {t['symbol'] for t in transfers}
        transfer_pnl = sum(cash_symbol_pnl.get(s, 0.0) for s in transfer_symbols)
        detail.append({
            '기준일': date,
            'CE 평가액': mv,
            'CE 평가액증감': None if first else mv - mv_prev,
            'CE 매매순대금': trade_net,
            'CE 배당(ex-date)': dividend,
            'CE 무상이전 내역': ', '.join(f"{t['symbol']} {t['qty']:,.0f}주({t['mnemonic']})"
                                     for t in transfers),
            'CE 무상이전 손익영향': None if first or not transfers else transfer_pnl,
            'CE 일간손익': cash_pnl,
            'EQSW 주식MTM손익': swap_parts['Equity via Swap']['equity'],
            'EQSW 금융비용': swap_parts['Equity via Swap']['financing'],
            'EQSW 배당손익': swap_parts['Equity via Swap']['dividend'],
            'EQSW 일간손익': daily['Equity via Swap'],
            'FUSW 주식MTM손익': swap_parts['Futures via Swap']['equity'],
            'FUSW 금융비용': swap_parts['Futures via Swap']['financing'],
            'FUSW 배당손익': swap_parts['Futures via Swap']['dividend'],
            'FUSW 일간손익': daily['Futures via Swap'],
            '스왑결제금액': sum(settle.get(date, {}).values()),
            '일간손익 합계': total_daily,
        })

        # ---------------- 종목별 손익 (Swap) ----------------
        if first:
            continue

        for bucket in BUCKETS[1:]:
            prev_rows = {r['contract']: r for r in swap_pos[prev][bucket]['rows']}
            cur_rows = {r['contract']: r for r in swap_pos[date][bucket]['rows']}
            for contract in sorted(set(prev_rows) | set(cur_rows)):
                was, now = prev_rows.get(contract, {}), cur_rows.get(contract, {})
                pnl = (now.get('mtm', 0.0) - was.get('mtm', 0.0)
                       + settle_contract.get(date, {}).get(contract, 0.0))
                if abs(pnl) < 0.005 and not now.get('qty') and not was.get('qty'):
                    continue
                instrument.append({
                    '기준일': date, 'Breakdown': bucket,
                    '종목코드': contract, '종목명': now.get('name') or was.get('name', ''),
                    '전일수량': was.get('qty', 0.0), '당일수량': now.get('qty', 0.0),
                    '전일평가액': was.get('mtm', 0.0), '당일평가액': now.get('mtm', 0.0),
                    '매매순대금': 0.0, '일간손익': pnl,
                })

    return (pd.DataFrame(summary), pd.DataFrame(detail),
            pd.DataFrame(instrument), dates, div_detail)


# --------------------------------------------------------------------------- #
# 엑셀 출력
# --------------------------------------------------------------------------- #
def build_notes(summary, dates, div_detail, root):
    rows = [
        ('대상 폴더', str(root)),
        ('기준일 범위', f'{dates[0]} ~ {dates[-1]} ({len(dates)}영업일)'),
        ('통화', 'USD (Base Currency)'),
        ('Cash Equity 정의', '커스터디 계좌 080056500, Product Type = EQ (현물 주식)'),
        ('Cash Equity 산식',
         '당일 평가액(Ending Market Value - Base) - 전일 평가액 + 매매 순대금(Base)'
         ' + 현금배당(ex-date 기준)'),
        ('Equity via Swap 정의', '스왑 계좌 080056492 의 개별주식 스왑/CFD 계약'),
        ('Futures via Swap 정의',
         '스왑 계좌 계약 중 Multiplier != 1 또는 기초자산명이 선물/지수인 계약'),
        ('Swap 산식',
         'Total Mark to Market(Base) 증감 + 당일 결제금액(MTD Settlement 의 Total Settlement)'),
        ('Swap 항등식(검증됨)',
         'Total MTM(Base) = Equity MTM(Base) + Total Interest Accrued(Base)'
         ' + Dividend Accrued(Base) + Unsettled P&L'),
        ('사용 리포트',
         'DATA Custody Position(AR=301712), DATA Custody Transaction(AR=286534), '
         'Syn Contract P&V(AR=303172), MTD SynSettle(AR=302553), '
         'DATA Asset Servicing Announcement(AR=303179)'),
        ('기준일 처리',
         f'{dates[0]}은 직전 영업일(전월 말) 리포트가 없어 손익을 산출하지 않고 기준선으로만 사용'),
        ('누적손익', f'{dates[1]} 이후 일간손익의 누적합 (당월 MTD)'),
        ('미포함 항목',
         '현금 잔고 이자, KRW/USD 환전(FX) 손익, 계좌 간 자금이동은 손익에 포함하지 않음'),
    ]
    if not (summary['Futures via Swap 일간손익'].fillna(0) != 0).any():
        rows.append(('Futures via Swap', '해당 기간 선물/지수 스왑 계약 없음 → 전 구간 0'))
    for ex_date, items in sorted(div_detail.items()):
        rows.append((f'현물 현금배당 ex-date {ex_date}',
                     ', '.join(f'{n}({r}) {a:,.2f}' for n, a, r in items)))
    for message in WARNINGS:
        rows.append(('확인 필요', message))
    return pd.DataFrame(rows, columns=['구분', '내용'])


def write_excel(out_path, summary, detail, instrument, dates, div_detail, root):
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    notes = build_notes(summary, dates, div_detail, root)
    for frame in (summary, detail, instrument):
        if '기준일' in frame.columns:
            frame['기준일'] = pd.to_datetime(frame['기준일'])

    sheets = {'손익요약': summary, '손익상세': detail,
              '종목별손익': instrument, '산출기준': notes}

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)

        book = writer.book
        head_fill = PatternFill('solid', fgColor='1F3864')
        head_font = Font(color='FFFFFF', bold=True, size=10)
        thin = Side(style='thin', color='BFBFBF')
        text_headers = {'종목코드', '종목명', 'Breakdown', '비고',
                    'CE 무상이전 내역', '구분', '내용'}

        for name in sheets:
            ws = book[name]
            for cell in ws[1]:
                cell.fill, cell.font = head_fill, head_font
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)
                cell.border = Border(bottom=thin)
            ws.row_dimensions[1].height = 30
            ws.freeze_panes = 'B2'
            ws.auto_filter.ref = ws.dimensions

            for column in ws.iter_cols(min_row=1):
                header = str(ws.cell(row=1, column=column[0].column).value or '')
                letter = get_column_letter(column[0].column)
                if header == '기준일':
                    width, fmt = 12, 'yyyy-mm-dd'
                elif name == '산출기준':
                    width, fmt = (26, None) if header == '구분' else (120, None)
                elif header in text_headers:
                    width = 34 if header in ('종목명', '비고', 'CE 무상이전 내역') else 16
                    fmt = None
                elif '수량' in header:
                    width, fmt = 12, '#,##0;[Red]-#,##0'
                else:
                    width, fmt = max(13, min(len(header) + 2, 22)), '#,##0.00;[Red]-#,##0.00'
                ws.column_dimensions[letter].width = width
                for cell in column[1:]:
                    if fmt:
                        cell.number_format = fmt
                    if name == '산출기준':
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

        # 누적손익 추이 차트
        ws = book['손익요약']
        chart = LineChart()
        chart.title = '누적손익 추이 (USD)'
        chart.height, chart.width = 9, 24
        chart.y_axis.numFmt = '#,##0'
        for label in ('Cash Equity 누적손익', 'Equity via Swap 누적손익',
                      'Futures via Swap 누적손익', '누적손익 합계'):
            col = summary.columns.get_loc(label) + 1
            chart.add_data(Reference(ws, min_col=col, min_row=1, max_row=ws.max_row),
                           titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, f'A{ws.max_row + 3}')

    return out_path


# --------------------------------------------------------------------------- #
def main():
    parser = argparse.ArgumentParser(description='Qube-RT 일일/누적 손익 엑셀 생성')
    parser.add_argument('root', nargs='?', default=str(DEFAULT_ROOT),
                        help=r'일자별 리포트 폴더가 들어있는 월 폴더 (예: ...\Qube-RT\202608)')
    parser.add_argument('-o', '--output', default=None, help='출력 엑셀 경로')
    args = parser.parse_args()

    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, 'reconfigure'):
            stream.reconfigure(encoding='utf-8', errors='replace')

    root = Path(args.root)
    if not root.is_dir():
        raise SystemExit(f'폴더를 찾을 수 없습니다: {root}')

    summary, detail, instrument, dates, div_detail = build_pnl(root)
    output = Path(args.output) if args.output else root / f'손익요약_{root.name}.xlsx'
    write_excel(output, summary, detail, instrument, dates, div_detail, root)

    pd.set_option('display.width', 220)
    pd.set_option('display.max_columns', 30)
    pd.set_option('display.float_format', lambda v: f'{v:,.0f}')
    print(summary[['기준일', 'Cash Equity 일간손익', 'Equity via Swap 일간손익',
                   'Futures via Swap 일간손익', '일간손익 합계',
                   '누적손익 합계']].to_string(index=False))
    if WARNINGS:
        print('\n[확인 필요]')
        for message in WARNINGS:
            print(' -', message)
    print(f'\n저장 완료: {output}')


if __name__ == '__main__':
    main()
