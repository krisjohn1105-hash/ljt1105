"""Qube-RT(QSMA) 월별 Manager's P&L 리포트 생성 — Qube 제출 양식/기준.

Qube 요구사항 (Edward Hougasian, Qube Fund Controller / 2026-08-05, 08-25 메일)
    - 계정 단위가 아니라 security/ticker 단위로 분해할 것
    - KRWUSD 환율은 런던 16:00 스냅 (Refinitiv). Citco는 자체 월말 환율로
      KRW 종가를 USD로 환산함
    - 실현손익(Realised P&L)은 trade date 기준 인식
    - 커미션은 instrument P&L 에 포함 (매매파일에 포함돼 있거나 체결가에 내재된 경우)
    - 스왑 financing 은 종목별로 쪼개지 않고 통화/계정 단위 1줄로 계상
    - 이자는 계정 성격별로 구분: Int Exp/Inc Broker(GS PB 계좌) vs
      Int Exp/Inc Equity Swaps(스왑 계좌)
    - 월별 비용(Part B)도 같은 표에 포함 가능

Citco(공식 장부) P&L 구성 — 2026년 7월 대사표에서 검증한 항등식
    Gross P&L = OTE Change + P&S           (미실현 증감 + 실현)
    P&L       = Gross P&L + Dividends      (배당은 종목 P&L 에 포함)
    FA P&L    = P&L
    financing/이자는 종목 P&L 에서 제외되고 GL 계정 단위로 별도 계상
    지수선물/개별주식선물 스왑도 Type = 'Equity Swap' 로 분류
    FX 손익은 Type = 'Cross Rate', ticker 'KRWUSD CURNCY'

출력 양식 (QSMA - Manager's P&L example.xlsx)
    Type | Description | Bloomberg Ticker | EOM quantity / positions
         | ME price | ME FX | $ MTD P&L

사용법
    python qube_monthly_pnl.py [월폴더] [--prev 전월폴더] [--ric-map 대사파일.xlsx]
                               [--expenses expenses.json] [-o 출력.xlsx]
"""

import argparse
import datetime as dt
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd

BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

import qube_pnl as base                                    # noqa: E402
from qube_pnl import CASH_ACCT, SWAP_ACCT, acct8, collect_files, num, read_report, to_date  # noqa: E402

# 월별 리포트에서 추가로 쓰는 GS 리포트
base.REPORT_PATTERNS.update({
    'cash_bal': '*Custody_Cash_Bal_302239_*.xls',    # Custody Cash Balances (FX 손익용)
    'int_mtd':  '*Interest_MTD_Acc_302415_*.xls',    # Interest MTD Accrual (PB 계좌 이자)
})

BASE_CCY = 'USD'

# Citco 분류에 맞춘 Type. 지수/개별주식 선물 스왑도 Citco 는 'Equity Swap' 으로 본다.
TYPE_EQUITY = 'Equity'
TYPE_SWAP = 'Equity Swap'
TYPE_FX = 'Cross Rate'
TYPE_INTEREST = 'Interest'
TYPE_EXPENSE = 'Expense'

# Citco GL 계정명 (2026-08-25 메일에서 확인)
GL_SWAP_INTEREST = 'Int Exp / Inc Equity Swaps'
GL_BROKER_INTEREST = 'Int Exp / Inc Broker'

WARNINGS = []


def cell(row, idx, name, default=''):
    """리포트마다 컬럼 구성이 조금씩 달라 없는 컬럼은 기본값으로 넘긴다."""
    pos = idx.get(name)
    return default if pos is None else row[pos]


def warn(message):
    if message not in WARNINGS:
        WARNINGS.append(message)


# --------------------------------------------------------------------------- #
# 티커 매핑
# --------------------------------------------------------------------------- #
def load_ric_map(path):
    """Qube 대사파일의 Ric_Ticker 시트 → {RIC: (BloombergCode, InstrumentType)}."""
    if not path:
        return {}
    try:
        frame = pd.read_excel(path, 'Ric_Ticker')
    except Exception as exc:
        warn(f'Ric_Ticker 시트를 읽지 못했습니다 ({path}): {exc}')
        return {}
    mapping = {}
    for _, row in frame.iterrows():
        ric = str(row.get('Ric', '')).strip()
        code = str(row.get('BloombergCode', '')).strip()
        if ric and code and code.lower() != 'nan':
            mapping[ric] = (code, str(row.get('InstrumentType', '')).strip())
    return mapping


def load_cost_basis(path):
    """무상입고(FREC) 종목의 취득원가를 주입한다 — {symbol: {...}}.

    GS 는 IPO 청약분을 대가 0 의 FREE RECEIVE 로 입고시키므로 취득원가가 리포트에
    전혀 없다. 그대로 두면 매도대금 전액이 손익으로 잡히므로 청약내역에서 원가를
    받아 차감한다.

    JSON 형식 (금액은 음수 = 취득원가 지출)
        {"KQ950260": {"cost_local": -363600000, "ccy": "KRW",
                      "fx_date": "2026-08-04",
                      "note": "공모가 x 배정수량 + 수수료 1%"},
         "KQ417030": {"cost_usd": -95000}}

    fx_date 를 주면 그 날짜 환율로 환산한다 (청약대금이 실제로 출금된 날).
    생략하면 무상입고일 환율을 쓴다.
    """
    if not path:
        return {}
    try:
        raw = json.loads(Path(path).read_text(encoding='utf-8'))
    except Exception as exc:
        warn(f'취득원가 파일을 읽지 못했습니다 ({path}): {exc}')
        return {}
    basis = {}
    for symbol, spec in raw.items():
        if isinstance(spec, (int, float)):
            spec = {'cost_usd': float(spec)}
        basis[str(symbol).strip()] = spec
    return basis


SUBSCRIPTION_PATH = Path(r"Z:\02.펀드\002.청약\청약내역_펀드_v2.xlsx")

base.REPORT_PATTERNS.setdefault('mtd_txn', '*DATA_Custody_MTD_303209_*.xls')


def load_subscription_prices(path=None):
    """청약내역 엑셀 → {6자리 티커: (공모단가, 수수료율)}.

    이 장부에는 국내 펀드(코벤/멀티/블록딜) 배정분만 있고 QSMA 배정분은 없다.
    다만 공모단가·수수료율은 전 펀드 공통이므로 단가만 가져와 GS 의 무상입고
    수량에 곱한다. 금액은 반드시 GS 현금 PAYMENT 와 교차검증한다.
    """
    path = Path(path or SUBSCRIPTION_PATH)
    if not path.is_file():
        warn(f'청약내역 파일을 찾을 수 없어 IPO 취득원가를 자동 도출하지 못했습니다: {path}')
        return {}
    try:
        raw = pd.read_excel(path, sheet_name='Sheet1', header=None)
        header = None
        for i in range(min(len(raw), 12)):
            values = [str(v).strip() for v in raw.iloc[i].tolist()]
            if 'Ticker' in values and '배정수량' in values:
                header = i
                break
        if header is None:
            warn(f'청약내역 파일에서 헤더(Ticker/배정수량)를 찾지 못했습니다: {path}')
            return {}
        frame = pd.read_excel(path, sheet_name='Sheet1', header=header)
    except Exception as exc:
        warn(f'청약내역 파일을 읽지 못했습니다 ({path}): {exc}')
        return {}

    frame.columns = [str(c).strip().replace('\n', ' ') for c in frame.columns]
    out = {}
    for _, row in frame.iterrows():
        ticker = re.sub(r'\D', '', str(row.get('Ticker', ''))).zfill(6)
        price = num(row.get('단가'))
        gross = num(row.get('청약금액'))
        fee = num(row.get('수수료'))
        if len(ticker) == 6 and price:
            out[ticker] = (price, (fee / gross) if gross else 0.0)
    return out


def load_cash_payments(root, month_start, month_end):
    """당월 현금 PAYMENT(매매 외 출금) — [(기준일, 현지금액, 통화)].

    IPO 청약대금이 이 경로로 빠져나가므로 취득원가 교차검증에 쓴다.
    """
    files = {d: p for d, p in collect_files(root, 'mtd_txn').items() if d <= month_end}
    if not files:
        return []
    _, idx, rows, datemode = read_report(
        files[max(files)], ['Account Number', 'Product Type', 'Trade Net Amount'])
    if idx is None:
        return []
    seen, out = set(), []
    for row in rows:
        if acct8(cell(row, idx, 'Account Number')) != CASH_ACCT:
            continue
        if 'PAYMENT' not in str(cell(row, idx, 'Transaction Mnemonic')).upper():
            continue
        bdate = to_date(cell(row, idx, 'Business Date'), datemode)
        amount = num(cell(row, idx, 'Trade Net Amount'))
        if bdate is None or not (month_start <= bdate <= month_end):
            continue
        key = (bdate, amount)
        if key in seen:
            continue
        seen.add(key)
        out.append((bdate, amount, str(cell(row, idx, 'Settle Currency')).strip().upper()))
    return out


def derive_cost_basis(root, transfers, month_start, month_end, subscription_path=None):
    """무상입고(FREC) 종목의 취득원가를 자동 도출 — {symbol: spec}.

    청약 공모단가 x 무상입고 수량 x (1+수수료율) 로 계산한 뒤,
    같은 금액의 GS 현금 PAYMENT 가 실제로 있는지 확인한 것만 채택한다.
    (2026-08 기준으로 3건 모두 원 단위까지 일치함을 확인)
    """
    prices = load_subscription_prices(subscription_path)
    if not prices:
        return {}
    payments = load_cash_payments(root, month_start, month_end)

    by_symbol = defaultdict(float)
    meta = {}
    for bdate, record in transfers:
        by_symbol[record['symbol']] += record['qty']
        meta.setdefault(record['symbol'], (bdate, record['name']))

    derived = {}
    for symbol, qty in by_symbol.items():
        ticker = re.sub(r'\D', '', symbol).zfill(6)
        if ticker not in prices or qty <= 0:
            continue
        price, fee_rate = prices[ticker]
        gross = qty * price
        expected = -(gross + round(gross * fee_rate))
        match = next(((d, amt, ccy) for d, amt, ccy in payments
                      if abs(amt - expected) <= 1.0), None)
        bdate, name = meta[symbol]
        if match is None:
            warn(f'{symbol} {name}: 청약단가 {price:,.0f}원 x {qty:,.0f}주 = '
                 f'{expected:,.0f} 인데 같은 금액의 GS 현금 PAYMENT 를 찾지 못했습니다 '
                 '— 자동 도출을 보류했습니다. --cost-basis 로 직접 넣으세요.')
            continue
        pay_date, amount, ccy = match
        derived[symbol] = {
            'cost_local': amount,
            'ccy': ccy or 'KRW',
            'fx_date': pay_date.isoformat(),
            'note': (f'{name} 공모단가 {price:,.0f} x {qty:,.0f}주 + 수수료 '
                     f'{fee_rate:.2%} = {amount:,.0f} {ccy}. '
                     f'GS 현금 PAYMENT {pay_date} 금액과 일치(자동 도출).'),
        }
        warn(f'{symbol} {name}: IPO 취득원가 {amount:,.0f} {ccy} 자동 도출 '
             f'(청약단가 {price:,.0f} x {qty:,.0f}주, GS PAYMENT {pay_date} 일치)')
    return derived


def resolve_cost(spec, transfer_date, fx_by_date, me_fx):
    """원가 스펙을 Base(USD) 금액으로 환산. 반환 (금액, 적용환율일)."""
    if 'cost_usd' in spec:
        return float(spec['cost_usd']), None
    amount = float(spec.get('cost_local', 0.0))
    currency = str(spec.get('ccy', '')).strip().upper()
    if not currency or currency == BASE_CCY:
        return amount, None

    fx_date = to_date(spec.get('fx_date')) or transfer_date
    # 지정일에 환율이 없으면(주말/리포트 공백) 그 이전 가장 가까운 영업일로 대체
    candidates = sorted(d for d in fx_by_date if d <= fx_date and fx_by_date[d].get(currency))
    used = candidates[-1] if candidates else None
    rate = fx_by_date[used][currency] if used else (me_fx or {}).get(currency)
    if not rate:
        warn(f'{currency} 환율을 찾지 못해 취득원가 {amount:,.0f} 를 환산하지 못했습니다.')
        return 0.0, None
    if used and used != fx_date:
        warn(f'{fx_date} 환율이 없어 {used} 환율로 취득원가를 환산했습니다.')
    return amount * rate, (used or 'ME')


def bloomberg_ticker(ric, gs_code, ric_map, is_index=False):
    """Bloomberg 티커 결정. Qube 매핑표 우선, 없으면 GS 코드에 접미사를 붙인다.

    GS 는 'Underlyer Bloomberg Code' 를 '005930 KP' / 'KMU6 Index' 형태로 주고
    Citco 는 '005930 KP Equity' / 'KMU6 INDEX' 를 쓴다. 대사 매칭키는 RIC 이므로
    티커는 표시용이며, Qube 매핑표가 있으면 그쪽을 따른다.
    """
    ric = (ric or '').strip()
    if ric in ric_map:
        return ric_map[ric][0]
    code = (gs_code or '').strip()
    if not code:
        return ric
    upper = code.upper()
    if upper.endswith(' EQUITY') or upper.endswith(' INDEX') or upper.endswith(' CURNCY'):
        return code
    if is_index or upper.endswith(' INDEX'):
        return code if upper.endswith('INDEX') else f'{code} Index'
    return f'{code} Equity'


# --------------------------------------------------------------------------- #
# 스왑 (Equity Swap) — 종목별
# --------------------------------------------------------------------------- #
def load_swap_snapshots(root):
    """{기준일: {contract_id: dict}} — 스왑 계약별 스냅샷.

    Unsettled P&L 은 Contract CCY 표기이고 'FX Contract to Base' 는 소수점 6자리로
    반올림돼 있어(KRW 계약은 오차가 수만 KRW 단위로 커진다) 곱셈 대신
    항등식에서 역산한다.
        Total MTM(Base) = Equity MTM(Base) + Total Interest Accrued(Base)
                          + Dividend Accrued(Base) + Unsettled P&L(Base)
    """
    snapshots = {}
    for bdate, path in sorted(collect_files(root, 'swap_pnv').items()):
        _, idx, rows, _ = read_report(path, ['Contract ID', 'Total Mark to Market (Base)'])
        day = {}
        if idx is None:
            snapshots[bdate] = day
            continue
        for row in rows:
            contract = str(cell(row, idx, 'Contract ID')).strip()
            if not contract or acct8(cell(row, idx, 'Account Number')) != SWAP_ACCT:
                continue
            total = num(cell(row, idx, 'Total Mark to Market (Base)'))
            equity_mtm = num(cell(row, idx, 'Equity Mark to Market (Base)'))
            financing = num(cell(row, idx, 'Total Interest Accrued (Base)'))
            dividend = num(cell(row, idx, 'Dividend Accrued (Base)'))
            unsettled = total - equity_mtm - financing - dividend   # 항등식 역산 (환율 오차 없음)
            multiplier = num(cell(row, idx, 'Multiplier')) or 1.0
            local_fx = num(cell(row, idx, 'FX Contract to Base')) or 1.0
            day[contract] = {
                'name': str(cell(row, idx, 'Underlyer Name')).strip(),
                'side': str(cell(row, idx, 'Long/Short')).strip(),
                'qty': num(cell(row, idx, 'Traded Quantity')),
                'multiplier': multiplier,
                'price': num(cell(row, idx, 'Current Price (Underlyer)')) or num(cell(row, idx, 'Current Price')),
                'ccy': str(cell(row, idx, 'Underlyer Currency')).strip().upper(),
                'fx_contract_base': local_fx,
                'bbg': str(cell(row, idx, 'Underlyer Bloomberg Code')).strip(),
                'ric': str(cell(row, idx, 'Underlyer RIC')).strip(),
                'isin': str(cell(row, idx, 'Underlyer ISIN')).strip(),
                # Consolidated Fund Summary 의 Synthetic MTM 은 Equity MTM 단독값이므로
                # 검산 대조를 위해 따로 보관한다. price_contract 는 Contract CCY 표시가격.
                'equity_mtm': equity_mtm,
                'unsettled': unsettled,
                'price_contract': num(cell(row, idx, 'Current Price')),
                # 종목 단위 P&L 구성 (Qube 기준: financing 제외)
                'equity': equity_mtm + unsettled,
                'dividend': dividend,
                'financing': financing,
                'total': total,
            }
        snapshots[bdate] = day
    return snapshots


def load_swap_settlements(root, month_start, month_end):
    """{contract_id: {'equity','dividend'}} 및 financing 합계 — 당월 결제분.

    MTDSynSettle 은 월 누적이므로 파일 1개만 쓰되, 폴더에 익월분(월이 바뀌면
    초기화된 파일)이 섞여 있을 수 있으므로 month_end 이하에서 최신본을 고른다.
    결제일이 당월 범위 밖(익월 예정분)인 행은 제외한다.
    """
    files = {d: p for d, p in collect_files(root, 'swap_settle').items()
             if d <= month_end}
    per_contract = defaultdict(lambda: {'equity': 0.0, 'dividend': 0.0, 'financing': 0.0})
    financing_total = 0.0
    if not files:
        warn('스왑 결제(MTD SynSettle) 리포트가 없어 당월 결제분을 반영하지 못했습니다.')
        return per_contract, financing_total, None

    path = files[max(files)]
    _, idx, rows, datemode = read_report(path, ['Payment Date', 'Total Settlement'])
    if idx is None:
        return per_contract, financing_total, path

    for row in rows:
        pay_date = to_date(cell(row, idx, 'Payment Date'), datemode)
        if pay_date is None or not (month_start <= pay_date <= month_end):
            continue
        contract = str(cell(row, idx, 'Contract ID')).strip()
        settled_ccy = str(cell(row, idx, 'Settled CCY')).strip().upper()
        fx = 1.0 if settled_ccy in ('', BASE_CCY) else (num(cell(row, idx, 'Settlement FX Rate')) or 1.0)
        if settled_ccy not in ('', BASE_CCY) and fx == 1.0:
            warn(f'스왑 결제 {contract} 의 결제통화가 {settled_ccy} 인데 환율이 1.0 입니다 '
                 '(Settlement FX Rate 확인 필요).')
        per_contract[contract]['equity'] += num(cell(row, idx, 'Equity Leg')) * fx
        per_contract[contract]['dividend'] += num(cell(row, idx, 'Dividend Leg')) * fx
        leg = num(cell(row, idx, 'Financing Leg')) * fx
        per_contract[contract]['financing'] += leg
        financing_total += leg
    return per_contract, financing_total, path


# --------------------------------------------------------------------------- #
# 현물 (Equity) — 종목별
# --------------------------------------------------------------------------- #
def load_cash_snapshots(root):
    """{기준일: {symbol: dict}} — 현물 주식 스냅샷 + 통화별 정밀 환율."""
    snapshots, fx_by_date = {}, {}
    for bdate, path in sorted(collect_files(root, 'custody_pos').items()):
        _, idx, rows, _ = read_report(
            path, ['Account Number', 'Product Type', 'Ending Market Value - Base'])
        day = {}
        fx_base, fx_local = defaultdict(float), defaultdict(float)
        if idx is not None:
            for row in rows:
                symbol = str(cell(row, idx, 'Symbol')).strip()
                if not symbol or acct8(cell(row, idx, 'Account Number')) != CASH_ACCT:
                    continue
                product = str(cell(row, idx, 'Product Type')).strip().upper()
                mv = num(cell(row, idx, 'Ending Market Value - Base'))
                if product != 'EQ':
                    if product != 'CA' and mv:
                        warn(f'{bdate}: 현물계좌 미분류 상품(Product Type={product}) '
                             f'평가액 {mv:,.2f} — 손익에서 제외됨')
                    continue
                currency = str(cell(row, idx, 'Ending Local Currency')).strip().upper()
                local_mv = num(cell(row, idx, 'Ending Market Value - Local'))
                if currency and local_mv:
                    fx_base[currency] += mv
                    fx_local[currency] += local_mv
                # 같은 종목이 Account Type(01/06) 별로 두 줄로 쪼개져 나오는 날이 있다
                # (결제 대기 이관). 상계되어 순포지션이 0 인 경우가 있으므로 합산해야 한다.
                bag = day.setdefault(symbol, {
                    'name': str(cell(row, idx, 'Product Description')).strip(),
                    'qty': 0.0,
                    'mv': 0.0,
                    'local_price': num(cell(row, idx, 'Ending Local Price')),
                    'ccy': currency,
                    'bbg': str(cell(row, idx, 'Bloomberg Ticker')).strip(),
                    'ric': str(cell(row, idx, 'RIC Code')).strip(),
                    'isin': str(cell(row, idx, 'ISIN')).strip(),
                })
                bag['qty'] += num(cell(row, idx, 'Trade Date Quantity'))
                bag['mv'] += mv
        snapshots[bdate] = day
        fx_by_date[bdate] = {c: fx_base[c] / fx_local[c] for c in fx_local if fx_local[c]}
    return snapshots, fx_by_date


def load_cash_trades_month(root, fx_by_date, month_start, month_end):
    """당월 현물 매매 집계. 커미션/세금 포함(Qube 기준).

    반환 (per_symbol, transfers, meta)
        meta : {symbol: {'name','ric','bbg','ccy'}} — 월말 포지션에 없는 종목
               (당월 중 전량 매도/무상입고 후 매도)의 표시 정보를 보완하는 데 쓴다
    """
    per_symbol = defaultdict(float)
    transfers, meta = [], {}
    trades = base.load_cash_trades(root, fx_by_date)
    for bdate, bag in trades.items():
        if not (month_start <= bdate <= month_end):
            continue
        for record in bag['rows']:
            per_symbol[record['symbol']] += record['net_base']
            if record['name']:
                meta.setdefault(record['symbol'], {
                    'name': record['name'],
                    'ric': record.get('ric', ''),
                    'bbg': record.get('bbg', ''),
                    'ccy': record.get('ccy', ''),
                })
        for record in bag['transfers']:
            transfers.append((bdate, record))
    return per_symbol, transfers, meta


def load_dividends_month(root, month_start, month_end):
    """{symbol 또는 이름: 금액} — 당월 ex-date 현물 현금배당."""
    by_date, detail = base.load_physical_dividends(root)
    total, items = 0.0, []
    for ex_date, amount in by_date.items():
        if month_start <= ex_date <= month_end:
            total += amount
            items.extend((ex_date, n, a, r) for n, a, r in detail.get(ex_date, []))
    return total, items


# --------------------------------------------------------------------------- #
# FX 손익 / PB 이자
# --------------------------------------------------------------------------- #
def load_fx_pnl(root, month_start, month_end, prev_root=None):
    """비-USD 현금잔고의 일별 환평가손익 합계 → Citco 'Cross Rate' 대응.

    FX 손익 = Σ_일 Σ_통화  전일잔고(Local) x (당일환율 - 전일환율)
    일별로 누적하므로 월 중 자금이동이 있어도 자연히 반영된다.

    당월 첫 영업일의 환변동은 전월말 잔고에 붙으므로, 전월 폴더의 월말 잔고를
    시작점으로 이어 붙인다(AR=302239 는 월 첫 영업일자가 당월 폴더에 있어도
    그 하루 전 변동은 전월말 잔고 기준이라 여기서 빠진다).
    """
    files = collect_files(root, 'cash_bal')
    if prev_root and Path(prev_root).is_dir():
        prev_files = {d: p for d, p in collect_files(prev_root, 'cash_bal').items()
                      if d < month_start}
        if prev_files:
            anchor = max(prev_files)
            files[anchor] = prev_files[anchor]
    if not files:
        warn('현금잔고(Custody Cash Balances) 리포트가 없어 FX 손익을 산출하지 못했습니다.')
        return 0.0, {}, None

    daily = {}
    for bdate, path in sorted(files.items()):
        _, idx, rows, _ = read_report(path, ['Account Number', 'Currency', 'Trade Date Qty (Local)'])
        if idx is None:
            continue
        balances, rates = defaultdict(float), {}
        for row in rows:
            if not str(cell(row, idx, 'Account Number')).strip():
                continue
            currency = str(cell(row, idx, 'Currency')).strip().upper()
            if not currency or currency == BASE_CCY:
                continue
            local = num(cell(row, idx, 'Trade Date Qty (Local)'))
            base_amt = num(cell(row, idx, 'Trade Date Qty (Base)'))
            balances[currency] += local
            if local:
                rates[currency] = base_amt / local      # 정밀 환율 (반올림 컬럼 대신 역산)
        daily[bdate] = (dict(balances), rates)

    dates = [d for d in sorted(daily) if d <= month_end]
    total, by_ccy = 0.0, defaultdict(float)
    for i in range(1, len(dates)):
        prev, cur = dates[i - 1], dates[i]
        if cur < month_start:
            continue
        prev_bal, prev_rate = daily[prev]
        _, cur_rate = daily[cur]
        for currency, balance in prev_bal.items():
            if currency in prev_rate and currency in cur_rate:
                amount = balance * (cur_rate[currency] - prev_rate[currency])
                total += amount
                by_ccy[currency] += amount

    # 시작점이 전월말이면 첫 영업일 환변동까지 잡힌 것이므로 경고 불필요
    if dates and dates[0] >= month_start:
        warn(f'현금잔고 시작점이 {dates[0]} 입니다 — 전월말 잔고가 없어 '
             f'{dates[0]} 하루의 환변동이 FX 손익에서 빠졌습니다.')
    return total, dict(by_ccy), max(files)


def load_broker_interest(root, month_start, month_end):
    """GS PB 계좌 MTD 이자 → Citco 'Int Exp / Inc Broker' 대응.

    이 리포트는 'MTD 누적'이지만 월 마지막 영업일자 파일이 하루치만 담고 초기화되는
    경우가 있다(2026-08 기준: BD 8/28 파일 = 7/31~8/30 누적 -8,226, BD 8/31 파일 =
    8/31 하루치 -696). 최신 파일 하나만 쓰면 8월 이자가 통째로 사라지므로,
    당월 파일 전체에서 (계좌·기간) 구간을 모아 중복 제거 후 합산한다.
    """
    files = {d: p for d, p in collect_files(root, 'int_mtd').items()
             if month_start <= d <= month_end}
    if not files:
        warn('Interest MTD Accrual 리포트가 없어 PB 계좌 이자를 반영하지 못했습니다.')
        return 0.0, None

    spans = {}
    for bdate in sorted(files):                 # 뒤에 오는 파일이 같은 구간을 덮어쓴다
        _, idx, rows, datemode = read_report(files[bdate],
                                             ['Account Number', 'Debit Interest Base'])
        if idx is None:
            continue
        for row in rows:
            account = str(cell(row, idx, 'Account Number')).strip()
            if not account:
                continue
            key = (acct8(account),
                   to_date(cell(row, idx, 'From Date'), datemode),
                   to_date(cell(row, idx, 'To Date'), datemode),
                   str(cell(row, idx, 'Product Description')).strip())
            spans[key] = (num(cell(row, idx, 'Debit Interest Base'))
                          + num(cell(row, idx, 'Credit Interest Base')))

    # 구간이 월경계를 걸치는 경우가 있다(예: 2026-07-31 ~ 08-02 3일치 -88.40).
    # 전월분은 전월 리포트에 이미 계상돼 있으므로 일수 비례로 당월분만 취한다.
    total, spilled = 0.0, 0.0
    for (_acct, from_dt, to_dt, _desc), amount in spans.items():
        if from_dt is None or to_dt is None:
            total += amount
            continue
        days = (to_dt - from_dt).days + 1
        inside = sum(1 for i in range(days)
                     if month_start <= from_dt + dt.timedelta(days=i) <= month_end)
        if inside == days:
            total += amount
        else:
            total += amount * inside / days
            spilled += amount * (days - inside) / days
    if spilled:
        warn(f'PB 이자 중 당월 밖 구간 {spilled:,.2f} 는 일수 비례로 제외했습니다 '
             '(월경계 걸친 누적구간 — 전월 리포트에 이미 계상됨).')
    return total, max(files)


# --------------------------------------------------------------------------- #
# 월별 손익 조립
# --------------------------------------------------------------------------- #
def build_monthly(root, prev_root, ric_map, cost_basis=None,
                  subscription_path=None):
    swap_snaps = load_swap_snapshots(root)
    cash_snaps, fx_by_date = load_cash_snapshots(root)

    dates = sorted(set(swap_snaps) | set(cash_snaps))
    if not dates:
        raise SystemExit(f'{root} 에서 포지션 리포트를 찾지 못했습니다.')

    # 대상 연월을 폴더명으로 확정하고 그 달 안에서만 기준일을 고른다.
    year, month = month_of(root, max(dates))
    dates = [d for d in dates if (d.year, d.month) == (year, month)]
    if not dates:
        raise SystemExit(f'{root}: {year}-{month:02d} 에 해당하는 기준일이 없습니다.')

    # 월말 평가 시점은 현물/스왑 스냅샷이 모두 존재하는 마지막 날이어야 한다.
    # 한쪽만 먼저 도착한 날을 월말로 잡으면 반대편 포지션이 통째로 0 으로 평가돼
    # 수백만 달러짜리 허위 손익이 난다. (비어 있는 'NO DATA' 스냅샷도 제외)
    usable_cash = {d for d, v in cash_snaps.items() if v and (d.year, d.month) == (year, month)}
    usable_swap = {d for d, v in swap_snaps.items() if v and (d.year, d.month) == (year, month)}
    common = sorted(usable_cash & usable_swap)
    if not common:
        raise SystemExit(
            f'{root}: 현물/스왑 포지션이 같은 기준일에 모두 존재하는 날이 없습니다.\n'
            f'  현물 스냅샷 {len(usable_cash)}일, 스왑 스냅샷 {len(usable_swap)}일')
    month_end = common[-1]
    month_start = dt.date(year, month, 1)

    # 리포트별 최신 기준일을 모아 '앞서 온 것 / 밀려 있는 것' 을 함께 알려준다.
    # 무엇을 기다려야 하는지(AR 코드 포함)가 실제로 필요한 정보다.
    latest = {}
    for label, key in (('현물 포지션', 'custody_pos'), ('스왑 P&V', 'swap_pnv'),
                       ('현물 거래', 'custody_trade'), ('현금잔고', 'cash_bal'),
                       ('스왑 결제', 'swap_settle'), ('자산관리', 'asset_serv'),
                       ('PB 이자', 'int_mtd')):
        found = [d for d in collect_files(root, key) if (d.year, d.month) == (year, month)]
        if found:
            ar = re.search(r'_(\d{6})_', base.REPORT_PATTERNS[key])
            latest[label] = (max(found), ar.group(1) if ar else '?')

    if latest:
        newest = max(v[0] for v in latest.values())
        lagging = {k: v for k, v in latest.items() if v[0] < newest}
        if lagging:
            need = ', '.join(f'{k}(AR={ar}) {d}' for k, (d, ar) in lagging.items())
            warn(f'리포트 도착일이 어긋납니다. 가장 최신은 {newest} 인데 다음이 밀려 '
                 f'있습니다: {need}. 월말 평가 시점은 현물/스왑이 모두 있는 '
                 f'{month_end} 를 사용했습니다 — 밀린 리포트가 도착하면 재실행하세요.')

    # 월 마지막 영업일(주말 제외) 리포트가 아직 없으면 MTD 가 월말 기준이 아니다.
    last_day = (dt.date(month_end.year + (month_end.month == 12),
                        month_end.month % 12 + 1, 1) - dt.timedelta(days=1))
    while last_day.weekday() >= 5:
        last_day -= dt.timedelta(days=1)
    if month_end < last_day:
        warn(f'월 마지막 영업일({last_day}) 리포트가 없어 MTD 를 {month_end} 까지만 '
             '집계했습니다 — 제출 전 해당일 리포트를 받아 재실행하세요.')

    # ---- 월초 기준선 (전월말 스냅샷) ----
    cash_base, cash_base_date = {}, None
    if prev_root and Path(prev_root).is_dir():
        # 전월 폴더를 읽으면 익월(=당월) 폴더까지 함께 훑기 때문에 당월 날짜가 섞여 들어온다.
        # 기준선은 반드시 당월 시작 이전의 마지막 스냅샷이어야 한다.
        def before_month(snaps):
            usable = [d for d in snaps if d < month_start and snaps[d]]
            return max(usable) if usable else None

        prev_cash, _ = load_cash_snapshots(prev_root)
        cash_base_date = before_month(prev_cash)
        if cash_base_date:
            cash_base = prev_cash[cash_base_date]
        else:
            cash_base_date, cash_base = dates[0], cash_snaps.get(dates[0], {})
            warn(f'전월 폴더({Path(prev_root).name})에 전월말 현물 포지션이 없어 '
                 f'기준선을 당월 첫 영업일({dates[0]})로 사용했습니다 → '
                 f'전월말~{dates[0]} 구간의 현물 손익이 MTD 에서 누락됩니다.')
        prev_swap = load_swap_snapshots(prev_root)
        swap_base_date = before_month(prev_swap)
        if swap_base_date:
            swap_base = prev_swap[swap_base_date]
        else:
            swap_base_date, swap_base = dates[0], swap_snaps.get(dates[0], {})
            warn(f'전월 폴더({Path(prev_root).name})에 스왑 Position & Valuation 리포트가 없어 '
                 f'스왑 기준선을 당월 첫 영업일({dates[0]})로 사용했습니다 → '
                 f'전월말~{dates[0]} 구간의 스왑 손익이 MTD 에서 누락됩니다.')
    else:
        cash_base_date, cash_base = dates[0], cash_snaps.get(dates[0], {})
        swap_base_date, swap_base = dates[0], swap_snaps.get(dates[0], {})
        warn(f'전월 폴더가 지정되지 않아 기준선을 당월 첫 영업일({dates[0]})로 사용했습니다 → '
             f'전월말~{dates[0]} 구간 손익이 MTD 에서 누락됩니다.')

    if cash_base_date and cash_base_date >= month_start:
        warn(f'현물 기준선이 {cash_base_date} (당월) 입니다 — 전월말 스냅샷이 아니므로 '
             'MTD 현물 손익에 첫 영업일 변동이 빠집니다.')

    cash_end = cash_snaps.get(month_end, {})
    swap_end = swap_snaps.get(month_end, {})

    # ---- 당월 거래/배당/결제 ----
    trades, transfers, trade_meta = load_cash_trades_month(
        root, fx_by_date, month_start, month_end)
    me_fx = fx_by_date.get(month_end, {})
    krw_per_usd = (1.0 / me_fx['KRW']) if me_fx.get('KRW') else None
    dividend_total, dividend_items = load_dividends_month(root, month_start, month_end)
    settle_by_contract, settle_financing, settle_path = load_swap_settlements(
        root, month_start, month_end)

    # ---- 무상입고 종목의 취득원가 주입 (없으면 매도대금 전액이 손익이 됨) ----
    cost_basis = cost_basis or {}
    cost_adjust, missing_cost, cost_fx_date = defaultdict(float), {}, {}
    for bdate, record in transfers:
        symbol = record['symbol']
        spec = cost_basis.get(symbol)
        if spec:
            amount, used = resolve_cost(spec, bdate, fx_by_date, me_fx)
            cost_adjust[symbol] += amount
            cost_fx_date[symbol] = used
        else:
            missing_cost[symbol] = (bdate, record['qty'], record['mnemonic'],
                                    record['name'])
    # 수동 입력이 없으면 청약내역 + GS 현금 PAYMENT 로 자동 도출을 시도한다
    if missing_cost:
        auto = derive_cost_basis(root, transfers, month_start, month_end,
                                 subscription_path)
        for symbol, spec in auto.items():
            if symbol in missing_cost:
                amount, used = resolve_cost(spec, missing_cost[symbol][0],
                                            fx_by_date, me_fx)
                cost_adjust[symbol] += amount
                cost_fx_date[symbol] = used
                cost_basis[symbol] = spec
                missing_cost.pop(symbol, None)
    for symbol, (bdate, qty, mnemonic, name) in missing_cost.items():
        warn(f'{bdate}: {symbol} {name} {qty:,.0f}주 무상입고({mnemonic}) — '
             '취득원가가 GS 리포트에 없고 청약내역에서도 찾지 못했습니다. '
             '--cost-basis 로 직접 넣지 않으면 매도대금 전액이 손익으로 계상됩니다.')
    for symbol, amount in cost_adjust.items():
        warn(f'{symbol}: 취득원가 {amount:,.2f} USD 차감 '
             f'(환율기준일 {cost_fx_date.get(symbol)}) — --cost-basis 반영')

    rows = []

    # ---- Equity (현물) ----
    for symbol in sorted(set(cash_base) | set(cash_end) | set(trades)):
        was, now = cash_base.get(symbol, {}), cash_end.get(symbol, {})
        adjust = cost_adjust.get(symbol, 0.0)
        pnl = ((now.get('mv', 0.0) - was.get('mv', 0.0)) + trades.get(symbol, 0.0)
               + adjust)
        if abs(pnl) < 0.005 and not now.get('qty') and not was.get('qty'):
            continue
        meta = now or was or {}
        if not meta.get('name'):        # 월말/기준선 스냅샷에 없는 종목은 거래에서 보완
            meta = {**trade_meta.get(symbol, {}), **{k: v for k, v in meta.items() if v}}
        rows.append({
            'Type': TYPE_EQUITY,
            'Description': meta.get('name', symbol),
            'Bloomberg Ticker': bloomberg_ticker(meta.get('ric'), meta.get('bbg'), ric_map),
            'EOM quantity / positions': now.get('qty', 0.0),
            'ME price': now.get('local_price', 0.0),
            'ME FX': (1.0 / me_fx[meta.get('ccy', '')]
                      if me_fx.get(meta.get('ccy', '')) else 1.0),
            '$ MTD P&L': pnl,
            '_ric': meta.get('ric', ''),
            '_isin': meta.get('isin', ''),
            '_equity': pnl - adjust,
            '_dividend': 0.0,
            '_financing': 0.0,
            '_cost_adj': adjust,
            '_key': symbol,
        })

    # 현물 현금배당은 종목별 배분 정보가 ex-date 공시에만 있어 Equity 합계에 별도 행으로 둔다
    if dividend_total:
        rows.append({
            'Type': TYPE_EQUITY,
            'Description': 'Physical equity cash dividends (ex-date basis)',
            'Bloomberg Ticker': '',
            'EOM quantity / positions': 0.0,
            'ME price': 0.0,
            'ME FX': krw_per_usd or 1.0,
            '$ MTD P&L': dividend_total,
            '_ric': '', '_isin': '',
            '_equity': 0.0, '_dividend': dividend_total, '_financing': 0.0,
            '_cost_adj': 0.0, '_key': 'PHYS_DIV',
        })

    # ---- Equity Swap (개별주식 + 지수/개별주식 선물 스왑) ----
    swap_financing = 0.0
    for contract in sorted(set(swap_base) | set(swap_end)):
        was, now = swap_base.get(contract, {}), swap_end.get(contract, {})
        settled = settle_by_contract.get(
            contract, {'equity': 0.0, 'dividend': 0.0, 'financing': 0.0})
        equity_pnl = (now.get('equity', 0.0) - was.get('equity', 0.0)) + settled['equity']
        dividend_pnl = (now.get('dividend', 0.0) - was.get('dividend', 0.0)) + settled['dividend']
        # financing 은 종목 P&L 에서 빼되(Qube 기준) Detail 시트 대조를 위해 계약별로 기록
        financing_pnl = (now.get('financing', 0.0) - was.get('financing', 0.0)
                         + settled['financing'])
        swap_financing += financing_pnl
        pnl = equity_pnl + dividend_pnl                     # Qube 기준: financing 제외
        # 당월 중 전량 종결된 계약은 종목 P&L 이 0 이라도 financing 이 남아있어
        # Detail 시트가 이자 계정라인과 대조되도록 행을 유지한다.
        if (abs(pnl) < 0.005 and abs(financing_pnl) < 0.005
                and not now.get('qty') and not was.get('qty')):
            continue
        meta = now or was
        is_index = meta.get('multiplier', 1.0) != 1.0
        rows.append({
            'Type': TYPE_SWAP,
            'Description': meta.get('name', contract),
            'Bloomberg Ticker': bloomberg_ticker(meta.get('ric'), meta.get('bbg'),
                                                 ric_map, is_index),
            'EOM quantity / positions': now.get('qty', 0.0),
            'ME price': now.get('price', 0.0),
            'ME FX': (1.0 / me_fx[meta.get('ccy', '')]
                      if me_fx.get(meta.get('ccy', '')) else (krw_per_usd or 1.0)),
            '$ MTD P&L': pnl,
            '_ric': meta.get('ric', ''),
            '_isin': meta.get('isin', ''),
            '_equity': equity_pnl,
            '_dividend': dividend_pnl,
            '_financing': financing_pnl,
            '_cost_adj': 0.0,
            '_key': contract,
        })

    # ---- FX (Cross Rate) ----
    fx_pnl, fx_by_ccy, _ = load_fx_pnl(root, month_start, month_end, prev_root)
    rows.append({
        'Type': TYPE_FX,
        'Description': 'KRW [USD] FX CROSS',
        'Bloomberg Ticker': 'KRWUSD CURNCY',
        'EOM quantity / positions': 0.0,
        'ME price': 0.0,
        'ME FX': krw_per_usd or 1.0,
        '$ MTD P&L': fx_pnl,
        '_ric': 'KRWUSD=R', '_isin': '',
        '_equity': fx_pnl, '_dividend': 0.0, '_financing': 0.0,
        '_cost_adj': 0.0, '_key': 'FX',
    })

    # ---- 이자 (Qube 기준: 종목별 배분 없이 계정 단위 1줄) ----
    broker_interest, _ = load_broker_interest(root, month_start, month_end)
    for label, amount in ((GL_SWAP_INTEREST, swap_financing),
                          (GL_BROKER_INTEREST, broker_interest)):
        rows.append({
            'Type': TYPE_INTEREST,
            'Description': label,
            'Bloomberg Ticker': '',
            'EOM quantity / positions': 0.0,
            'ME price': 0.0,
            'ME FX': 1.0,
            '$ MTD P&L': amount,
            '_ric': '', '_isin': '',
            '_equity': 0.0, '_dividend': 0.0, '_financing': amount,
            '_cost_adj': 0.0, '_key': label,
        })

    meta = {
        'month_start': month_start,
        'month_end': month_end,
        'dates': dates,
        'cash_base_date': cash_base_date,
        'swap_base_date': swap_base_date,
        'krw_per_usd': krw_per_usd,
        'dividend_items': dividend_items,
        'fx_by_ccy': fx_by_ccy,
        'settle_path': settle_path,
        'transfers': transfers,
        'missing_cost': missing_cost,
        'cost_adjust': dict(cost_adjust),
    }
    return pd.DataFrame(rows), meta


# --------------------------------------------------------------------------- #
# 엑셀 출력
# --------------------------------------------------------------------------- #
QUBE_COLUMNS = ['Type', 'Description', 'Bloomberg Ticker', 'EOM quantity / positions',
                'ME price', 'ME FX', '$ MTD P&L']


def collect_warnings(meta):
    """월별 리포트용 경고 목록.

    취득원가를 이미 반영한 종목에 대해서는 일별엔진(qube_pnl)의 '매입원가 없음'
    경고를 빼서 같은 리포트 안에서 모순된 안내가 나오지 않게 한다.
    """
    applied = set(meta.get('cost_adjust') or {})
    messages = list(WARNINGS)
    for message in base.WARNINGS:
        if '무상' in message and any(symbol in message for symbol in applied):
            continue
        if message not in messages:
            messages.append(message)
    return messages


def write_excel(out_path, frame, meta, expenses, root):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    month_end = meta['month_end']
    positions = frame[frame['Type'].isin([TYPE_EQUITY, TYPE_SWAP, TYPE_FX])]
    interest = frame[frame['Type'] == TYPE_INTEREST]
    subtotal = float(frame['$ MTD P&L'].sum())
    expense_total = float(sum(expenses.values()))

    # ---- Manager P&L (Qube 제출 시트) ----
    # financing 만 남은 종결 계약(P&L=0)은 이자 계정라인에 이미 포함되므로 제출표에서 제외
    blocks = []
    visible = positions[positions['$ MTD P&L'].abs() >= 0.005]
    for _, row in pd.concat([visible, interest]).iterrows():
        blocks.append([row['Type'], row['Description'], row['Bloomberg Ticker'],
                       row['EOM quantity / positions'], row['ME price'],
                       row['ME FX'], row['$ MTD P&L']])
    blocks.append([None] * 7)
    blocks.append([None, 'Sub-total', None, None, None, None, subtotal])
    if expenses:
        blocks.append([None] * 7)
        blocks.append([None, 'Monthly expenses', None, None, None, None, None])
        for label, amount in expenses.items():
            blocks.append([TYPE_EXPENSE, label, None, None, None, None, amount])
    blocks.append([None] * 7)
    blocks.append([None, 'TOTAL', None, None, None, None, subtotal + expense_total])
    manager = pd.DataFrame(blocks, columns=QUBE_COLUMNS)

    # ---- Reconciliation (Qube Summary 시트에 붙일 형태) ----
    recon = pd.DataFrame({
        'Description': frame['Description'],
        'Bloomberg-Ticker': frame['Bloomberg Ticker'],
        'Ric': frame['_ric'],
        'ISIN': frame['_isin'],
        'Mgr P&L': frame['$ MTD P&L'],
        'Citco (Admin P&L)': None,
        'Difference (Mgr to Citco)': None,
    })

    # ---- Detail (내부 검증용) ----
    detail = pd.DataFrame({
        'Type': frame['Type'],
        'Description': frame['Description'],
        'Key(Contract/Symbol)': frame['_key'],
        'Ric': frame['_ric'],
        'EOM qty': frame['EOM quantity / positions'],
        'ME price(Local)': frame['ME price'],
        'ME FX(Local per USD)': frame['ME FX'],
        'Equity P&L': frame['_equity'],
        'Dividend P&L': frame['_dividend'],
        'Financing': frame['_financing'],
        '무상입고 취득원가': frame['_cost_adj'],
        '$ MTD P&L': frame['$ MTD P&L'],
    })

    # ---- Basis (산출기준 + Qube 기준 인용 + 경고) ----
    notes = [
        ('대상 폴더', str(root)),
        ('MTD 기간', f"{meta['month_start']} ~ {month_end} (리포트 {len(meta['dates'])}영업일)"),
        ('통화', f'{BASE_CCY} (Base Currency)'),
        ('월말 KRW/USD',
         f"{meta['krw_per_usd']:,.4f}" if meta['krw_per_usd'] else '산출 불가'),
        ('현물 기준선', f"{meta['cash_base_date']} 스냅샷"),
        ('스왑 기준선', f"{meta['swap_base_date']} 스냅샷"),
        ('', ''),
        ('[Qube 기준] 요구 형식',
         'security/ticker 단위 분해 (계정 단위 아님) — 2026-08-05 Edward Hougasian'),
        ('[Qube 기준] 환율',
         'KRWUSD 는 런던 16:00 스냅(Refinitiv). Citco 는 자체 월말 환율로 KRW 종가를 '
         'USD 환산 → 본 리포트는 GS 리포트 환율을 역산해 사용하므로 소수 차이 발생 가능'),
        ('[Qube 기준] 실현손익', 'trade date 기준 인식 (GS 리포트도 trade-date 기반)'),
        ('[Qube 기준] 커미션', 'instrument P&L 에 포함 — 현물은 Trade Net Amount, '
                            '스왑은 Net Price 에 내재'),
        ('[Qube 기준] 스왑 financing',
         f'종목별 배분 없이 계정 단위 1줄 ({GL_SWAP_INTEREST})'),
        ('[Qube 기준] 이자 계정 구분',
         f'{GL_BROKER_INTEREST} = GS PB 계좌 / {GL_SWAP_INTEREST} = 스왑 계좌'),
        ('', ''),
        ('[Citco 항등식] 종목 P&L', 'P&L = Gross P&L + Dividends, Gross P&L = OTE Change + P&S'),
        ('[Citco 분류] 선물',
         '지수/개별주식 선물 스왑도 Type = Equity Swap (Price Factor = Multiplier)'),
        ('[Citco 분류] FX', "Type = Cross Rate, ticker 'KRWUSD CURNCY'"),
        ('', ''),
        ('산식: Equity',
         '월말 평가액(Base) - 기준선 평가액(Base) + 당월 매매순대금(Base) + 현금배당(ex-date)'),
        ('산식: Equity Swap',
         '(Equity MTM + Unsettled P&L + Dividend Accrued) 증감 + 당월 결제(Equity/Dividend Leg). '
         'financing 은 제외하고 별도 계정 라인으로 계상'),
        ('산식: Unsettled P&L(Base)',
         'Total MTM(Base) - Equity MTM(Base) - Interest(Base) - Dividend(Base) 로 역산 '
         '(FX Contract to Base 가 6자리 반올림이라 KRW 계약에서 오차가 커짐)'),
        ('산식: FX', 'Σ_일 Σ_통화 전일 현금잔고(Local) x (당일환율 - 전일환율)'),
        ('사용 리포트',
         'Custody Position(AR=301712), Custody Transaction(AR=286534), '
         'Syn Contract P&V(AR=303172), MTD SynSettle(AR=302553), '
         'Asset Servicing(AR=303179), Custody Cash Balances(AR=302239), '
         'Interest MTD Accrual(AR=302415), Custody MTD Transaction(AR=303209)'),
        ('', ''),
        ('[Citco 와 다를 수 있는 항목] 1. 월말 환율',
         'Qube/Citco 는 KRWUSD 를 런던 16:00 Refinitiv 스냅으로 사용. 본 리포트는 '
         'GS 포지션 리포트의 Base/Local 평가액에서 역산한 환율을 쓴다. 2026-07 대사에서 '
         'Dunamis 1,424 vs Citco 1,438.23 (약 1%) 차이 발생 — Citco 월말 환율을 받으면 '
         '고정값으로 대체 가능.'),
        ('[Citco 와 다를 수 있는 항목] 2. 실현/미실현 배분',
         'Citco 는 OTE Change(자기 장부원가 대비) / P&S 로 구분. 본 리포트의 현물은 '
         '월초 평가액을 기초원가로 보는 이동평균법이라 합계는 같아도 배분이 다를 수 있다. '
         '스왑은 Equity MTM / Unsettled P&L 로 나눠 Citco 개념과 정렬됨.'),
        ('[Citco 와 다를 수 있는 항목] 3. 선물환·현금 평가',
         '미결제 FX(Forward Cash) 포지션을 스팟으로 재평가해 Cross Rate 에 포함했다. '
         'Citco 는 Cross Rate 와 Cash Balance 를 분리하고 결제일 기준 선물환율을 쓸 '
         '가능성이 있어 줄 단위로는 어긋날 수 있다.'),
        ('미포함 항목',
         '계좌 간 자금이동(IPO 청약대금 등)은 손익이 아니므로 제외. '
         '871m 원천징수·대차·Stock Loan 리포트는 당월 NO DATA 로 확인됨.'),
    ]
    for ex_date, name, amount, ric in meta['dividend_items']:
        notes.append((f'현물 배당 ex-date {ex_date}',
                      f'{name} ({ric}) {amount:,.2f}'))
    for currency, amount in meta['fx_by_ccy'].items():
        notes.append((f'FX 손익 내역 {currency}', f'{amount:,.2f}'))
    for message in collect_warnings(meta):
        notes.append(('확인 필요', message))
    basis = pd.DataFrame(notes, columns=['구분', '내용'])

    sheets = {'Manager P&L': manager, 'Reconciliation': recon,
              'Detail': detail, 'Basis': basis}

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for name, sheet in sheets.items():
            sheet.to_excel(writer, sheet_name=name, index=False)

        book = writer.book
        head_fill = PatternFill('solid', fgColor='1F3864')
        head_font = Font(color='FFFFFF', bold=True, size=10)
        thin = Side(style='thin', color='BFBFBF')
        money = '#,##0.00;[Red]-#,##0.00'

        for name, sheet in sheets.items():
            ws = book[name]
            for cell in ws[1]:
                cell.fill, cell.font = head_fill, head_font
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)
                cell.border = Border(bottom=thin)
            ws.row_dimensions[1].height = 30
            ws.freeze_panes = 'A2'
            if name != 'Basis':
                ws.auto_filter.ref = ws.dimensions

            for column in ws.iter_cols(min_row=1):
                header = str(ws.cell(row=1, column=column[0].column).value or '')
                letter = get_column_letter(column[0].column)
                if name == 'Basis':
                    width, fmt = (26, None) if header == '구분' else (120, None)
                elif header in ('Description', 'Instrument Description'):
                    width, fmt = 44, None
                elif header in ('Type', 'Bloomberg Ticker', 'Ric', 'ISIN',
                                'Key(Contract/Symbol)'):
                    width, fmt = 20, None
                elif 'qty' in header.lower() or 'quantity' in header.lower():
                    width, fmt = 16, '#,##0;[Red]-#,##0'
                elif header in ('ME price', 'ME FX', 'ME price(Local)',
                                'ME FX(Local per USD)'):
                    width, fmt = 14, '#,##0.0000'
                else:
                    width, fmt = max(14, min(len(header) + 2, 24)), money
                ws.column_dimensions[letter].width = width
                for cell in column[1:]:
                    if fmt:
                        cell.number_format = fmt
                    if name == 'Basis':
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Sub-total / TOTAL 행 강조
        ws = book['Manager P&L']
        for row in ws.iter_rows(min_row=2):
            label = str(row[1].value or '')
            if label in ('Sub-total', 'TOTAL', 'Monthly expenses'):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.border = Border(top=thin)

    return subtotal, expense_total


# --------------------------------------------------------------------------- #
def month_of(root, fallback):
    """대상 연월을 폴더명(YYYYMM)에서 확정한다.

    월 폴더 안에 익월 초 리포트가 섞여 들어오는 일이 잦다(예: 202608 폴더의
    '0901' 하위폴더에 BD 9/1 파일). max(기준일) 로 월을 정하면 8월 리포트가
    9월 하루짜리로 둔갑하므로 폴더명을 우선한다.
    """
    name = Path(root).name
    if len(name) == 6 and name.isdigit():
        year, month = int(name[:4]), int(name[4:])
        if 1 <= month <= 12:
            return year, month
    return fallback.year, fallback.month


def guess_prev_root(root):
    """202608 → 202607 형태의 형제 폴더를 추정."""
    root = Path(root)
    if not root.name.isdigit() or len(root.name) != 6:
        return None
    year, month = int(root.name[:4]), int(root.name[4:])
    year, month = (year - 1, 12) if month == 1 else (year, month - 1)
    candidate = root.parent / f'{year}{month:02d}'
    return candidate if candidate.is_dir() else None


def main():
    parser = argparse.ArgumentParser(description='Qube 제출용 월별 Manager P&L 생성')
    parser.add_argument('root', nargs='?', default=str(base.DEFAULT_ROOT),
                        help=r'월 폴더 (예: ...\Qube-RT\202608)')
    parser.add_argument('--prev', default=None,
                        help='전월 폴더 (생략 시 형제 폴더에서 자동 추정)')
    parser.add_argument('--ric-map', default=None,
                        help='Qube 대사파일 경로 (Ric_Ticker 시트를 티커 매핑에 사용)')
    parser.add_argument('--expenses', default=None,
                        help='월별 비용 JSON (예: {"Other Data costs": -3150})')
    parser.add_argument('--cost-basis', default=None,
                        help='무상입고(IPO 청약) 종목의 취득원가 JSON '
                             '(생략하면 청약내역에서 자동 도출)')
    parser.add_argument('--subscriptions', default=None,
                        help=f'청약내역 엑셀 (기본 {SUBSCRIPTION_PATH})')
    parser.add_argument('-o', '--output', default=None, help='출력 엑셀 경로')
    args = parser.parse_args()

    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, 'reconfigure'):
            stream.reconfigure(encoding='utf-8', errors='replace')

    root = Path(args.root)
    if not root.is_dir():
        raise SystemExit(f'폴더를 찾을 수 없습니다: {root}')
    prev_root = Path(args.prev) if args.prev else guess_prev_root(root)

    expenses = {}
    if args.expenses:
        expenses = json.loads(Path(args.expenses).read_text(encoding='utf-8'))

    ric_map = load_ric_map(args.ric_map)
    cost_basis = load_cost_basis(args.cost_basis)
    frame, meta = build_monthly(root, prev_root, ric_map, cost_basis,
                                args.subscriptions)

    output = Path(args.output) if args.output else \
        root / f"Dunamis - Manager's P&L {root.name}.xlsx"
    subtotal, expense_total = write_excel(output, frame, meta, expenses, root)

    print(f"MTD 기간      : {meta['month_start']} ~ {meta['month_end']}")
    print(f"현물 기준선   : {meta['cash_base_date']}")
    print(f"스왑 기준선   : {meta['swap_base_date']}")
    print(f"전월 폴더     : {prev_root if prev_root else '(없음)'}")
    print()
    summary = frame.groupby('Type')['$ MTD P&L'].agg(['size', 'sum'])
    for type_name, row in summary.iterrows():
        print(f"  {type_name:<14} {int(row['size']):>3}건  {row['sum']:>16,.2f}")
    print(f"  {'Sub-total':<14} {'':>5}  {subtotal:>16,.2f}")
    if expenses:
        print(f"  {'Expenses':<14} {'':>5}  {expense_total:>16,.2f}")
        print(f"  {'TOTAL':<14} {'':>5}  {subtotal + expense_total:>16,.2f}")

    if meta['missing_cost']:
        template = output.with_name(f'cost_basis_template_{root.name}.json')
        sample = {
            symbol: {'cost_local': 0, 'ccy': 'KRW',
                     'note': f'{name} {qty:,.0f}주 {bdate} 무상입고 — '
                             '공모가 x 배정수량 + 수수료를 음수로 입력'}
            for symbol, (bdate, qty, _mnem, name) in meta['missing_cost'].items()
        }
        template.write_text(json.dumps(sample, ensure_ascii=False, indent=2),
                            encoding='utf-8')
        print(f'\n취득원가 미입력 종목이 있어 템플릿을 만들었습니다: {template}')
        print('  값을 채운 뒤 --cost-basis 로 지정해 재실행하세요.')

    messages = collect_warnings(meta)
    if messages:
        print('\n[확인 필요]')
        for message in messages:
            print(' -', message)
    print(f'\n저장 완료: {output}')


if __name__ == '__main__':
    main()
