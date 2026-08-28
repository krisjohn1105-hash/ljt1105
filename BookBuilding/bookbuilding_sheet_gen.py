# -*- coding: utf-8 -*-
"""
38.co.kr 수요예측 일정에서 종목을 가져와
'01_수요예측_사전수량배분(YYYY).xlsx' 에 종목별 시트를 자동 생성한다.

- 템플릿(가장 최근 시트)을 Excel COM 으로 그대로 복사하므로
  서식 / 수식 / 병합 / 메모 / 열너비가 100% 유지된다.
- 종목 정보(종목명, 종목코드, 희망공모가 밴드, 기관배정 최대수량)는 웹에서 채운다.
- 자산총액 3개월평균 / 확약기간 / 참여수량 / 단가 는 공백으로 남긴다.
- 설정금액(D열) / 순자산 전일·3개월(O·P열) 은 템플릿 값을 그대로 두고 직접 수정한다.
- 시트명 주관사는 국내기관 수요예측을 접수하는 대표주관사를 쓴다.
- 같은 종목이 재수요예측을 하면 날짜가 다르므로 시트를 새로 만든다.
- 실행할 때마다 BACKUP_DIR 에 '원본파일명_YYYYMMDD.xlsx' 로 백업한다.

사용법:
    python bookbuilding_sheet_gen.py --dry-run     # 생성 대상만 확인
    python bookbuilding_sheet_gen.py               # 실제 시트 생성
    python bookbuilding_sheet_gen.py --all         # 수요예측 완료 종목까지 포함
    python bookbuilding_sheet_gen.py --file "D:/other.xlsx"
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import shutil
import ssl
import sys
from dataclasses import dataclass, field
from pathlib import Path

import requests
import urllib3
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.poolmanager import PoolManager

urllib3.disable_warnings()

# ──────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────

DEFAULT_XLSX = r"Z:\02.펀드\001.수요예측\01_수요예측_사전수량배분(2026).xlsx"

# 실행할 때마다 '원본파일명_YYYYMMDD.xlsx' 로 여기에 백업한다.
BACKUP_DIR = r"Z:\02.펀드\001.수요예측\Backup"

LIST_URL = "https://www.38.co.kr/html/fund/?o=r"
DETAIL_URL = "https://www.38.co.kr/html/fund/{href}"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

# 시트명 = {수요예측 시작일:YYYYMMDD}_{종목명}_{주관사 약칭}
SHEET_NAME_FMT = "{date}_{name}_{broker}"

# 주관사 → 시트명 약칭(영문). 여기만 고치면 표기가 바뀐다.
BROKER_ABBR = {
    "NH투자증권": "NH",
    "한국투자증권": "KIS",
    "삼성증권": "Samsung",
    "미래에셋증권": "Mirae",
    "KB증권": "KB",
    "대신증권": "Daishin",
    "유진투자증권": "Eugene",
    "신한투자증권": "Shinhan",
    "신한금융투자": "Shinhan",
    "키움증권": "Kiwoom",
    "IBK투자증권": "IBK",
    "DB증권": "DB",
    "DB금융투자": "DB",
    "메리츠증권": "Meritz",
    "SK증권": "SK",
    "하나증권": "Hana",
    "신영증권": "Shinyoung",
    "한화투자증권": "Hanwha",
    "교보증권": "Kyobo",
    "상상인증권": "Sangsangin",
    "유안타증권": "Yuanta",
    "BNK투자증권": "BNK",
    "다올투자증권": "Daol",
    "현대차증권": "HMSEC",
    "LS증권": "LS",
    "이베스트투자증권": "LS",
    "부국증권": "Bookook",
    "흥국증권": "Heungkuk",
    "한양증권": "Hanyang",
    "케이비증권": "KB",
}

# 인수회사 역할 우선순위 — 국내기관 수요예측은 대표주관사가 접수한다.
# 38.co.kr 은 '접수 창구'를 따로 표기하지 않으므로 이 순위를 대용으로 쓴다.
ROLE_RANK = {"대표주관": 0, "대표": 1, "공동주관": 2, "공동": 3, "인수": 4}

# 예외적으로 대표주관사가 아닌 곳이 국내기관 접수를 받는 종목은 여기에 적는다.
#   예) "빅웨이브로보틱스": "미래에셋증권"
BROKER_OVERRIDE: dict[str, str] = {}

# 종목 정보 입력 셀 (템플릿 레이아웃 기준)
CELL_STOCK_NAME = "C5"   # 종목명
CELL_STOCK_CODE = "D5"   # 종목코드
CELL_PRICE_MAX = "K4"    # 희망공모가 상단 (Max.)
CELL_PRICE_MIN = "K5"    # 희망공모가 하단 (Min.)
CELL_MAX_QTY = "K6"      # 최대가능수량 = 기관투자자등 배정물량 상단

# 무조건 비우는 셀 — 참여수량 / 단가 / 확약기간 (딜마다 직접 입력)
FORCE_BLANK = ["G9", "H9", "M9"]

# 값(상수)만 지우고 수식은 그대로 두는 범위 — 자산총액 3개월평균
BLANK_IF_LITERAL = ["E9:E13", "E18:E19", "E24:E28"]

INVALID_SHEET_CHARS = "[]:*?/\\'"
MAX_SHEET_NAME = 31

_WARNED_BROKERS: set[str] = set()


# ──────────────────────────────────────────────────────────────────────────
# 38.co.kr 스크래핑
# ──────────────────────────────────────────────────────────────────────────

class _LegacyTLSAdapter(HTTPAdapter):
    """38.co.kr 은 구형 TLS/cipher 를 쓰므로 보안수준을 낮춘 컨텍스트가 필요하다."""

    def init_poolmanager(self, connections, maxsize, block=False, **kw):
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        ctx.options |= 0x4  # OP_LEGACY_SERVER_CONNECT
        try:
            ctx.minimum_version = ssl.TLSVersion.TLSv1
        except (AttributeError, ValueError):
            pass
        ctx.set_ciphers("DEFAULT@SECLEVEL=0")
        kw["ssl_context"] = ctx
        self.poolmanager = PoolManager(num_pools=connections, maxsize=maxsize,
                                       block=block, **kw)


def make_session() -> requests.Session:
    s = requests.Session()
    s.mount("https://", _LegacyTLSAdapter())
    s.headers.update({"User-Agent": UA})
    return s


def _get(session: requests.Session, url: str) -> BeautifulSoup:
    r = session.get(url, timeout=30, verify=False)
    r.raise_for_status()
    r.encoding = "euc-kr"
    return BeautifulSoup(r.text, "html.parser")


def _num(text: str | None) -> int | None:
    if not text:
        return None
    t = re.sub(r"[^\d]", "", text)
    return int(t) if t else None


@dataclass
class Deal:
    name: str                    # 표기 종목명 (원문)
    bb_start: dt.date            # 수요예측 시작일
    bb_end: dt.date | None
    price_min: int | None
    price_max: int | None
    price_fixed: int | None      # 확정공모가 (없으면 None)
    offer_amount_mn: int | None  # 공모금액(백만)
    brokers: list[str] = field(default_factory=list)
    href: str = ""
    # 상세 페이지에서 채움
    code: str | None = None
    total_shares: int | None = None
    inst_qty_max: int | None = None
    underwriters: list[tuple[str, str]] = field(default_factory=list)  # (회사, 역할)

    @property
    def lead_broker(self) -> str | None:
        """국내기관 수요예측 접수 주관사 = 대표주관사."""
        if (ov := BROKER_OVERRIDE.get(self.clean_name)):
            return ov
        if self.underwriters:
            return min(self.underwriters,
                       key=lambda x: ROLE_RANK.get(x[1], 9))[0]
        return self.brokers[0] if self.brokers else None

    @property
    def lead_role(self) -> str:
        if BROKER_OVERRIDE.get(self.clean_name):
            return "수동지정"
        lead = self.lead_broker
        for name, role in self.underwriters:
            if name == lead:
                return role
        return "단독" if len(self.brokers) <= 1 else "목록순"

    @property
    def broker_abbr(self) -> str:
        raw = self.lead_broker
        if not raw:
            return "NA"
        if BROKER_ABBR.get(raw):
            return BROKER_ABBR[raw]
        # 미등록 주관사: '○○투자증권' / '○○증권' 에서 접미사만 떼고 사용
        guess = re.sub(r"(투자)?증권$|금융투자$", "", raw).strip()
        if raw not in _WARNED_BROKERS:
            _WARNED_BROKERS.add(raw)
            print(f"  [주의] 주관사 약칭 미등록: {raw!r} -> {guess!r} "
                  f"(BROKER_ABBR 에 추가하세요)")
        return guess or raw

    @property
    def clean_name(self) -> str:
        """'덕산넵코어스(구.넵코어스)' -> '덕산넵코어스'"""
        n = re.sub(r"\s*\([^)]*\)\s*", "", self.name).strip()
        n = n.replace(" ", "")
        for ch in INVALID_SHEET_CHARS:
            n = n.replace(ch, "")
        return n or self.name

    def sheet_name(self) -> str:
        broker = self.broker_abbr
        date = self.bb_start.strftime("%Y%m%d")
        name = self.clean_name
        sheet = SHEET_NAME_FMT.format(date=date, name=name, broker=broker)
        if len(sheet) > MAX_SHEET_NAME:  # 종목명을 잘라 31자에 맞춘다
            over = len(sheet) - MAX_SHEET_NAME
            name = name[: max(1, len(name) - over)]
            sheet = SHEET_NAME_FMT.format(date=date, name=name, broker=broker)
        return sheet

    def summary(self) -> str:
        band = ("-" if self.price_min is None
                else f"{self.price_min:,}~{self.price_max:,}")
        inst = "-" if self.inst_qty_max is None else f"{self.inst_qty_max:,}"
        return (f"{self.sheet_name():<32} 공모가 {band:>17} / "
                f"코드 {self.code or '-':<8} / 기관배정 {inst:>11} / "
                f"주관 {self.lead_broker}({self.lead_role})")


def _parse_range(text: str) -> tuple[int | None, int | None]:
    """'13,000~16,000' -> (13000, 16000)"""
    nums = [_num(p) for p in text.split("~") if p.strip()]
    nums = [n for n in nums if n is not None]
    if not nums:
        return None, None
    return min(nums), max(nums)


def _parse_bb_dates(text: str) -> tuple[dt.date | None, dt.date | None]:
    """'2026.09.02~09.08' 또는 '2026.09.02 ~ 2026.09.08' -> (start, end)"""
    text = re.sub(r"\s+", "", text)
    m = re.match(r"(\d{4})\.(\d{1,2})\.(\d{1,2})~(?:(\d{4})\.)?(\d{1,2})\.(\d{1,2})",
                 text)
    if not m:
        m2 = re.match(r"(\d{4})\.(\d{1,2})\.(\d{1,2})", text)
        if not m2:
            return None, None
        y, mo, d = map(int, m2.groups())
        return dt.date(y, mo, d), None
    y, mo, d, y2, mo2, d2 = m.groups()
    start = dt.date(int(y), int(mo), int(d))
    end_year = int(y2) if y2 else int(y)
    end = dt.date(end_year, int(mo2), int(d2))
    if end < start:  # 연말 → 연초 롤오버
        end = dt.date(end_year + 1, int(mo2), int(d2))
    return start, end


def fetch_deals(session: requests.Session) -> list[Deal]:
    """수요예측 일정 목록을 긁어온다."""
    soup = _get(session, LIST_URL)

    target = None
    for table in soup.find_all("table"):
        head = table.find("tr")
        if not head:
            continue
        cells = [td.get_text(strip=True) for td in head.find_all(["td", "th"])]
        if "종목명" in cells and any("수요예측일" in c for c in cells):
            target = table
            break
    if target is None:
        raise RuntimeError(
            "수요예측 일정 표를 찾지 못했습니다. 페이지 구조가 바뀐 것 같습니다.")

    deals: list[Deal] = []
    for tr in target.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 6:
            continue
        cells = [td.get_text(" ", strip=True).replace("\xa0", " ") for td in tds]
        if cells[0] in ("종목명", ""):
            continue
        start, end = _parse_bb_dates(cells[1])
        if start is None:
            continue
        pmin, pmax = _parse_range(cells[2])
        link = tr.find("a")
        deals.append(Deal(
            name=cells[0],
            bb_start=start,
            bb_end=end,
            price_min=pmin,
            price_max=pmax,
            price_fixed=_num(cells[3]) if cells[3] not in ("-", "") else None,
            offer_amount_mn=_num(cells[4]),
            brokers=[b.strip() for b in cells[5].split(",") if b.strip()],
            href=(link.get("href") or "").lstrip("./") if link else "",
        ))
    return deals


def _parse_underwriters(text: str) -> list[tuple[str, str]]:
    """상세 페이지의 '인수회사 | 주식수 | 청약한도 | 기타' 표를 (회사, 역할)로 파싱.

    공동주관 딜에만 이 표가 있고, 단독주관이면 빈 리스트가 돌아온다.
    """
    m = re.search(
        r"인수회사\s*\|\s*주식수\s*\|\s*청약한도\s*\|\s*기타\s*\|(.*?)\|\s*주요일정", text)
    if not m:
        return []
    parts = [p.strip() for p in m.group(1).split("|")]
    if not parts or len(parts) % 4:
        print(f"  [주의] 인수회사 표 형식이 예상과 다릅니다: {parts}")
        return []
    return [(parts[i], parts[i + 3]) for i in range(0, len(parts), 4)]


def enrich_detail(session: requests.Session, deal: Deal) -> None:
    """상세 페이지에서 종목코드 / 총공모주식수 / 기관배정물량을 채운다."""
    if not deal.href:
        return
    try:
        soup = _get(session, DETAIL_URL.format(href=deal.href))
    except Exception as exc:  # noqa: BLE001
        print(f"  [주의] {deal.name} 상세 조회 실패: {exc}")
        return

    text = re.sub(r"\s+", " ", soup.get_text("|", strip=True))

    deal.underwriters = _parse_underwriters(text)

    m = re.search(r"종목코드\s*\|\s*([0-9A-Za-z]+)", text)
    if m:
        deal.code = m.group(1)

    m = re.search(r"총공모주식수\s*\|\s*([\d,]+)", text)
    if m:
        deal.total_shares = _num(m.group(1))

    # '기관투자자등 | 1,400,000~1,500,000 주  (70.0~75.0%)'
    m = re.search(r"기관투자자등\s*\|\s*([\d,]+)(?:\s*~\s*([\d,]+))?\s*주", text)
    if m:
        vals = [v for v in (_num(m.group(1)), _num(m.group(2))) if v]
        if vals:
            deal.inst_qty_max = max(vals)

    # 희망공모가는 상세 페이지 값을 우선 (목록보다 최신인 경우가 있다)
    m = re.search(r"희망공모가액\s*\|\s*([\d,]+)\s*~\s*([\d,]+)", text)
    if m:
        deal.price_min = _num(m.group(1))
        deal.price_max = _num(m.group(2))


# ──────────────────────────────────────────────────────────────────────────
# Excel 시트 생성 (COM — 서식/수식/메모 완전 보존)
# ──────────────────────────────────────────────────────────────────────────

SHEET_DATE_RE = re.compile(r"^(\d{8})_")


def pick_template(sheet_names: list[str]) -> str:
    """시트명 앞 YYYYMMDD 가 가장 큰 시트를 템플릿으로 쓴다."""
    dated = [(m.group(1), n) for n in sheet_names
             if (m := SHEET_DATE_RE.match(n))]
    if not dated:
        raise RuntimeError(
            "YYYYMMDD_ 로 시작하는 시트가 없어 템플릿을 정할 수 없습니다.")
    return max(dated)[1]


def _clear(ws, ref: str) -> None:
    """셀 내용을 지운다. 병합 셀이면 병합 영역 전체를 대상으로 한다."""
    rng = ws.Range(ref)
    if rng.MergeCells:
        rng.MergeArea.ClearContents()
    else:
        rng.ClearContents()


def _clear_literals(ws, ref: str) -> None:
    """범위 안에서 수식이 아닌 셀만 지운다."""
    for cell in ws.Range(ref):
        f = cell.Formula
        if isinstance(f, str) and f.startswith("="):
            continue
        if f in (None, ""):
            continue
        _clear(ws, cell.Address)


def make_backup(xlsx: Path, overwrite: bool) -> Path | None:
    """BACKUP_DIR 에 '원본파일명_YYYYMMDD.xlsx' 로 백업한다."""
    bak_dir = Path(BACKUP_DIR)
    bak_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.date.today().strftime("%Y%m%d")
    bak = bak_dir / f"{xlsx.stem}_{stamp}{xlsx.suffix}"

    if bak.exists() and not overwrite:
        # 같은 날 두 번째 실행. 덮어쓰면 '변경 전 원본'이 사라지므로 기존 것을 지킨다.
        print(f"백업 유지: {bak.name} (오늘자 백업이 이미 있어 덮어쓰지 않음. "
              f"교체하려면 --overwrite-backup)")
        return bak

    shutil.copy2(xlsx, bak)
    print(f"백업 생성: {bak}")
    return bak


def create_sheets(xlsx: Path, deals: list[Deal], template_name: str | None,
                  backup: bool, overwrite_backup: bool = False) -> list[str]:
    import win32com.client as win32

    if backup:
        make_backup(xlsx, overwrite_backup)

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    created: list[str] = []
    try:
        wb = excel.Workbooks.Open(str(xlsx))
        try:
            names = [wb.Worksheets(i + 1).Name for i in range(wb.Worksheets.Count)]
            tpl_name = template_name or pick_template(names)
            if tpl_name not in names:
                raise RuntimeError(f"템플릿 시트 {tpl_name!r} 를 찾을 수 없습니다.")
            print(f"템플릿 시트: {tpl_name}")
            template = wb.Worksheets(tpl_name)

            # 오래된 딜부터 넣어 맨 앞이 가장 최신이 되도록
            for deal in sorted(deals, key=lambda d: d.bb_start):
                sheet_name = deal.sheet_name()
                template.Copy(Before=wb.Worksheets(1))
                ws = wb.Worksheets(1)
                ws.Name = sheet_name

                for ref in FORCE_BLANK:
                    _clear(ws, ref)
                for ref in BLANK_IF_LITERAL:
                    _clear_literals(ws, ref)

                ws.Range(CELL_STOCK_NAME).Value = deal.clean_name
                ws.Range(CELL_STOCK_CODE).Value = deal.code or ""
                ws.Range(CELL_PRICE_MAX).Value = deal.price_max
                ws.Range(CELL_PRICE_MIN).Value = deal.price_min
                ws.Range(CELL_MAX_QTY).Value = deal.inst_qty_max

                created.append(sheet_name)
                print(f"  + {sheet_name}")

            if created:
                wb.Save()
        finally:
            wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
    return created


# ──────────────────────────────────────────────────────────────────────────
# main
# ──────────────────────────────────────────────────────────────────────────

def existing_sheet_names(xlsx: Path) -> list[str]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="38.co.kr 수요예측 종목 시트 자동 생성")
    ap.add_argument("--file", default=DEFAULT_XLSX, help="대상 엑셀 파일")
    ap.add_argument("--template", default=None,
                    help="복사할 템플릿 시트명 (기본: 날짜가 가장 최근인 시트)")
    ap.add_argument("--all", action="store_true",
                    help="확정공모가가 나온 종목까지 포함")
    ap.add_argument("--dry-run", action="store_true", help="생성 대상만 출력")
    ap.add_argument("--no-backup", action="store_true",
                    help="백업 파일을 만들지 않음")
    ap.add_argument("--overwrite-backup", action="store_true",
                    help="같은 날짜 백업이 이미 있어도 덮어씀")
    args = ap.parse_args(argv)

    xlsx = Path(args.file)
    if not xlsx.exists():
        print(f"[에러] 파일이 없습니다: {xlsx}")
        return 1

    session = make_session()
    print("38.co.kr 수요예측 일정 조회 …")
    deals = fetch_deals(session)
    print(f"  목록 {len(deals)}건")

    if not args.all:
        deals = [d for d in deals if d.price_fixed is None]
        print(f"  확정공모가 미정 {len(deals)}건")

    have = existing_sheet_names(xlsx)
    have_lower = {n.lower() for n in have}

    todo: list[Deal] = []
    for deal in deals:
        enrich_detail(session, deal)
        name = deal.sheet_name()
        if name.lower() in have_lower:
            print(f"  = 이미 존재: {name}")
            continue
        dup = [n for n in have if deal.clean_name in n]
        if dup:
            print(f"  * 재수요예측: 기존 {dup} 와 별도로 {name} 를 만듭니다")
        todo.append(deal)

    if not todo:
        print("\n생성할 시트가 없습니다.")
        return 0

    print(f"\n생성 대상 {len(todo)}건")
    for d in sorted(todo, key=lambda x: x.bb_start):
        print("  " + d.summary())

    if args.dry_run:
        print("\n--dry-run: 엑셀 파일은 변경하지 않았습니다.")
        return 0

    created = create_sheets(xlsx, todo, args.template,
                            backup=not args.no_backup,
                            overwrite_backup=args.overwrite_backup)
    print(f"\n완료: {len(created)}개 시트 생성 → {xlsx}")
    print("공백으로 남긴 항목: 자산총액 3개월평균 / 확약기간 / 참여수량 / 단가")
    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    raise SystemExit(main())
