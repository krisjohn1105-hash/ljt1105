import os
import shutil
import datetime
import openpyxl
from crawl_ipo import run_ipo_crawler

def get_quarterly_filename():
    """현재 날짜를 기준으로 YYYYQn 포맷의 파일명을 반환합니다."""
    today = datetime.date.today()
    year = today.year
    quarter = (today.month - 1) // 3 + 1
    return f"{year}Q{quarter}_result.xlsx"

def apply_data_to_sheet(ws, data_dict):
    """
    시트 내의 지정된 키워드를 찾아, 바로 오른쪽(열+1) 셀에 데이터를 입력합니다.
    """
    # 탐색할 키워드들 목록
    keywords = ['종목명', '종목코드', '시장구분', '희망공모가액', '수요예측일', '공모청약일', '기관투자자 최대배정수량', '주간사']
    
    # openpyxl 시트 순회
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                if val in keywords and val in data_dict:
                    # 해당 키워드의 우측 셀을 선택
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    target_cell.value = data_dict[val]

def main():
    template_path = 'IPO_Template.xlsx'
    result_filename = get_quarterly_filename()
    
    # 1. 분기별 결과 파일이 없으면 템플릿 복사
    if not os.path.exists(result_filename):
        if not os.path.exists(template_path):
            print(f"[에러] {template_path} 원본 템플릿 파일이 존재하지 않습니다.")
            return
        
        print(f"[알림] 새로운 분기 엑셀 파일 생성: {result_filename}")
        shutil.copy(template_path, result_filename)
        
    try:
        wb = openpyxl.load_workbook(result_filename)
    except Exception as e:
        print(f"[에러] {result_filename} 엑셀 파일을 여는 중 오류 발생: {e}")
        return
        
    # 2. 존재하는 시트 이름 가져오기 (existing_items 지정)
    existing_items = wb.sheetnames
    
    # 3. 크롤링 스크립트 실행 (존재하는 종목들은 스킵)
    print("IPO 수요예측 크롤링을 시작합니다...\n")
    ipo_list = run_ipo_crawler(existing_items)
    
    if not ipo_list:
        print("\n새로 추가할 종목 정보가 없습니다.")
        return
    
    # 4. 대상 종목들에 대해 시트 복사 및 데이터 입력 (오픈파이셀)
    # 기준 시트 찾기 (첫 번째 시트를 템플릿으로 가정, 가급적 'Template' 등 우선)
    base_sheet_name = 'Template' if 'Template' in wb.sheetnames else wb.sheetnames[0]
    base_sheet = wb[base_sheet_name]
    
    added_count = 0
    for item_data in ipo_list:
        item_name = item_data.get('종목명')
        if not item_name:
            continue
            
        print(f"[{item_name}] 시트 생성 및 데이터 매핑 중...")
        
        # 새 시트 복사
        new_ws = wb.copy_worksheet(base_sheet)
        new_ws.title = str(item_name).replace('/', '_').replace('\\', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_').replace('?', '_')[:31]
        
        # 키워드 기반 값 주입
        apply_data_to_sheet(new_ws, item_data)
        added_count += 1
        
    # 5. 저장
    try:
        wb.save(result_filename)
        print(f"\n작업 완료! {added_count}개 종목이 {result_filename} 에 성공적으로 추가되었습니다.")
    except Exception as e:
        print(f"[에러] 파일을 저장하는 중 오류 발생: {e}\n(엑셀 파일이 열려있는지 확인하세요)")
        
if __name__ == '__main__':
    main()
