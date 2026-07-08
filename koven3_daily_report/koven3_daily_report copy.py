import os
import sys
import time
import datetime
import glob
import shutil
import pandas as pd
from openpyxl import load_workbook
import win32com.client as win32
from apscheduler.schedulers.blocking import BlockingScheduler

# PROJECT_ROOT 설정
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)

import constants
import input_database
from common_function import get_fund_pnl_database

# ════════════════════════════════════════════════════════
# 설정 및 토글 (개발 중에는 TEST_MODE = True 상태로 둡니다)
# ════════════════════════════════════════════════════════
TEST_MODE = False  # True: 메일을 보내지 않고 화면에 드래프트(Display) 창을 띄웁니다. False: 실제 이메일을 발송합니다.

# 수신자 목록
RECIPIENTS = ['dhong@truspringpartners.com', 'khpark@dunamiscap.com', 'wkim@dunamiscap.com', 'jtlee@dunamiscap.com']  # 대석, 기흥
# RECIPIENTS = ["jtlee@dunamiscap.com"]
# TEST_RECIPIENT = 'jtlee@dunamiscap.com'  # 테스트 모드 시 수신자 (필요 시 수정)

# ════════════════════════════════════════════════════════
# 1. 엑셀 및 DB 데이터 조회 함수들
# ════════════════════════════════════════════════════════
def get_prev_nav_and_aum():
    """
    10304 일자별기준가격 조회 오늘날짜 엑셀 파일에서 코벤3호의 전일 수정기준가와 AUM을 가져옵니다.
    오늘날짜 파일이 없을 경우 최신 일일 파일을 찾아 안전하게 로드합니다.
    """
    import re
    today_str = datetime.datetime.now().strftime("%Y%m%d")
    today_file = os.path.join(constants.LONGSHORT_DIR, f"10304_일자별기준가격 조회_{today_str}.xlsx")
    
    selected_file = None
    if os.path.exists(today_file):
        selected_file = today_file
        print(f"[*] 오늘 날짜 AUM 엑셀 파일 로드: {selected_file}")
    else:
        file_pattern = os.path.join(constants.LONGSHORT_DIR, "10304_일자별기준가격 조회_*.xlsx")
        files = glob.glob(file_pattern)
        
        # 8자리 숫자로 이루어진 일일 데이터 파일만 필터링 (EOM 'X월' 파일 제외)
        daily_files = []
        for f in files:
            basename = os.path.basename(f)
            if re.search(r"\d{8}", basename):
                daily_files.append(f)
                
        if not daily_files:
            raise FileNotFoundError(f"AUM 일일 엑셀 파일을 찾을 수 없습니다. (경로: {constants.LONGSHORT_DIR})")
        
        daily_files.sort()
        selected_file = daily_files[-1]
        print(f"[!] 오늘 날짜 엑셀이 존재하지 않아, 가장 최근의 일일 파일 로드: {selected_file}")
        
    # 파일 잠금(Lock) 에러 방지를 위해 임시 파일로 복사하여 읽기
    temp_file = os.path.join(PROJECT_ROOT, "temp_koven3_aum.xlsx")
    try:
        shutil.copy2(selected_file, temp_file)
        wb = load_workbook(temp_file, data_only=True)
        sheet = wb["22204_기준가격표"]
        
        aum_pre = None
        nav_pre = None
        # E열(5)부터 J열(10)까지 순회
        for row in sheet.iter_rows(min_row=2, max_row=100, min_col=5, max_col=10):
            fund_name_e = row[0].value
            if fund_name_e == "두나미스 코스닥벤처 일반사모투자신탁 3호":
                aum_pre = row[2].value  # G열
                nav_pre = row[5].value  # J열
                break
                
        if aum_pre is None or nav_pre is None:
            raise ValueError("엑셀 파일 내에서 '두나미스 코스닥벤처 일반사모투자신탁 3호'의 데이터를 찾을 수 없습니다.")
            
        return float(aum_pre), float(nav_pre)
    finally:
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass

def get_eom_nav():
    """
    MTD(월간) 수익률 계산을 위해 전월말 수정기준가를 가져옵니다.
    """
    try:
        from etc_module.hints_pnl import hints_month_excel
        eom_dict = hints_month_excel()
        val = eom_dict.get('kv_3')
        if val is not None:
            return float(val)
    except FileNotFoundError:
        print("[!] 이번 달 EOM AUM 엑셀 파일이 존재하지 않아, 이전 달 파일을 탐색합니다.")
        try:
            prev_month = (datetime.datetime.now().replace(day=1) - datetime.timedelta(days=1)).month
            end_of_month_excel = os.path.join(constants.LONGSHORT_DIR, f"10304_일자별기준가격 조회_{prev_month}월.xlsx")
            if os.path.exists(end_of_month_excel):
                eom_df = pd.read_excel(end_of_month_excel)
                df_cleaned = eom_df.dropna(subset=['No', '통화코드'])
                val = df_cleaned.loc[df_cleaned['펀드코드'] == 'DM13030', 'Unnamed: 9'].tolist()[0]
                return float(val)
        except Exception as e:
            print(f"[!] 이전 달 EOM AUM 조회 실패: {e}")
    except Exception as e:
        print(f"[!] EOM NAV 조회 중 오류 발생: {e}")
    return 1000.0  # 기본값

# ════════════════════════════════════════════════════════
# 2. 포맷 및 헬퍼 함수
# ════════════════════════════════════════════════════════
def format_krw(amount):
    """
    금액을 억원/만원 단위의 예쁜 문자열로 포맷팅합니다.
    """
    manwon = amount / 10000
    if abs(manwon) >= 10000:
        eok = manwon / 10000
        if eok.is_integer():
            return f"{'+' if amount > 0 else ''}{int(eok)}억원"
        return f"{'+' if amount > 0 else ''}{eok:.1f}억원"
    return f"{'+' if amount > 0 else ''}{int(manwon):,}만원"

def format_pct(value):
    """
    수익률 퍼센트를 포맷팅합니다.
    """
    pct = value * 100
    return f"{'+' if pct > 0 else ''}{pct:.2f}%"

# ════════════════════════════════════════════════════════
# 3. 리포트 생성 및 이메일 전송
# ════════════════════════════════════════════════════════
def run_report():
    print(f"\n==================================================")
    print(f"[*] 코벤3호 Daily NAV 및 수익률 보고 처리 시작 ({datetime.datetime.now()})")
    print(f"==================================================")
    
    try:
        # 데이터 수집
        aum_pre, nav_pre = get_prev_nav_and_aum()
        daily_pnl = get_fund_pnl_database("venture-3")
        eom_nav = get_eom_nav()
        
        # 지표 산출
        chg_pct = daily_pnl / aum_pre if aum_pre > 0 else 0
        today_nav = nav_pre * (1 + chg_pct)
        today_aum = aum_pre + daily_pnl
        
        # MTD, YTD, Cumulative
        total_pct = (today_nav - 1000) / 1000
        mtd_pct = today_nav / eom_nav - 1 if eom_nav > 0 else 0
        ytd_pct = today_nav / 1329.7383 - 1  # 2026년 기준 ytd base
        
        # 콘솔 출력 및 로깅
        print(f"[*] 전일 AUM: {aum_pre/100000000:.2f}억원 / 전일 NAV: {nav_pre:.4f}")
        print(f"[*] 당일 P&L: {format_krw(daily_pnl)} (수익률: {format_pct(chg_pct)})")
        print(f"[*] 당일 AUM: {today_aum/100000000:.2f}억원 / 당일 NAV: {today_nav:.4f}")
        print(f"[*] MTD: {format_pct(mtd_pct)} / YTD: {format_pct(ytd_pct)} / Cumulative: {format_pct(total_pct)}")
        
        # HTML 메일 본문 템플릿 생성
        now = datetime.datetime.now()
        today_str = now.strftime("%Y-%m-%d")
        time_str = f"{now.month}월 {now.day}일 {now.hour}시 {now.minute}분 기준"
        
        pnl_color = "#ef4444" if daily_pnl >= 0 else "#3b82f6"
        mtd_color = "#ef4444" if mtd_pct >= 0 else "#3b82f6"
        ytd_color = "#ef4444" if ytd_pct >= 0 else "#3b82f6"
        cml_color = "#ef4444" if total_pct >= 0 else "#3b82f6"
        
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
        <meta charset="utf-8">
        </head>
        <body style="font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, Roboto, sans-serif; background-color: #f4f6f9; margin: 0; padding: 0; color: #333333;">
        <div class="container" style="max-width: 600px; margin: 20px auto; background-color: #ffffff; border-radius: 12px; box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05); overflow: hidden; border: 1px solid #e1e5eb;">
          <div class="header" style="background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%); color: #ffffff; padding: 30px 24px; text-align: center;">
            <h1 style="margin: 0; font-size: 24px; font-weight: 600; letter-spacing: -0.5px;">두나미스 코스닥벤처 3호</h1>
            <p style="margin: 8px 0 0 0; font-size: 14px; opacity: 0.9;">Daily NAV & 수익률 일일 보고 ({today_str})</p>
          </div>
          <div class="content" style="padding: 24px;">
            <div class="pnl-hero" style="text-align: center; background-color: #f8fafc; border-radius: 8px; padding: 20px; margin-bottom: 24px; border: 1px solid #edf2f7;">
              <div class="label" style="font-size: 12px; text-transform: uppercase; letter-spacing: 1px; color: #64748b; margin-bottom: 6px;">{time_str} 당일 손익 (DAILY P&L)</div>
              <div class="value" style="font-size: 32px; font-weight: 700; color: {pnl_color};">{format_krw(daily_pnl)}</div>
              <div class="sub-value" style="font-size: 14px; font-weight: 500; margin-top: 6px; color: {pnl_color};">일일 수익률: {format_pct(chg_pct)}</div>
            </div>
            
            <div class="grid" style="margin-bottom: 24px;">
              <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%">
                <tr>
                  <td width="50%" style="padding-right: 8px;">
                    <div class="grid-card" style="background-color: #ffffff; border: 1px solid #edf2f7; border-radius: 8px; padding: 16px; text-align: center;">
                      <div class="card-label" style="font-size: 11px; color: #64748b; margin-bottom: 4px;">수정기준가 (NAV)</div>
                      <div class="card-value" style="font-size: 18px; font-weight: 600; color: #1e293b;">{today_nav:.4f}</div>
                    </div>
                  </td>
                  <td width="50%" style="padding-left: 8px;">
                    <div class="grid-card" style="background-color: #ffffff; border: 1px solid #edf2f7; border-radius: 8px; padding: 16px; text-align: center;">
                      <div class="card-label" style="font-size: 11px; color: #64748b; margin-bottom: 4px;">순자산 (AUM)</div>
                      <div class="card-value" style="font-size: 18px; font-weight: 600; color: #1e293b;">{today_aum/100000000:.2f}억원</div>
                    </div>
                  </td>
                </tr>
              </table>
            </div>
            
            <div class="table-container" style="margin-top: 24px;">
              <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                <thead>
                  <tr style="background-color: #f8fafc;">
                    <th style="color: #64748b; font-weight: 600; text-align: left; padding: 10px 12px; border-bottom: 1px solid #edf2f7;">구분</th>
                    <th style="color: #64748b; font-weight: 600; text-align: right; padding: 10px 12px; border-bottom: 1px solid #edf2f7; text-align: right;">수익률</th>
                  </tr>
                </thead>
                <tbody>
                  <tr>
                    <td style="padding: 12px; border-bottom: 1px solid #f1f5f9; color: #334155;">당월 수익률 (MTD)</td>
                    <td style="padding: 12px; border-bottom: 1px solid #f1f5f9; text-align: right; font-weight: 600; color: {mtd_color};">{format_pct(mtd_pct)}</td>
                  </tr>
                  <tr>
                    <td style="padding: 12px; border-bottom: 1px solid #f1f5f9; color: #334155;">금년 수익률 (YTD)</td>
                    <td style="padding: 12px; border-bottom: 1px solid #f1f5f9; text-align: right; font-weight: 600; color: {ytd_color};">{format_pct(ytd_pct)}</td>
                  </tr>
                  <tr>
                    <td style="padding: 12px; border-bottom: none; color: #334155;">누적 수익률 (Cumulative)</td>
                    <td style="padding: 12px; border-bottom: none; text-align: right; font-weight: 600; color: {cml_color};">{format_pct(total_pct)}</td>
                  </tr>
                </tbody>
              </table>
            </div>
          </div>
          <div class="footer" style="background-color: #f8fafc; padding: 16px; text-align: center; font-size: 11px; color: #94a3b8; border-top: 1px solid #edf2f7;">
            본 메일은 두나미스 자동화 시스템에 의해 생성 및 발송되었습니다.
          </div>
        </div>
        </body>
        </html>
        """
        
        # Outlook API 연동
        outlook = None
        try:
            outlook = win32.GetActiveObject("Outlook.Application")
            print("[*] 기존에 실행 중인 Outlook 인스턴스를 사용합니다.")
        except:
            print("[*] Outlook이 백그라운드에서 감지되지 않아 새로 구동합니다.")
            outlook = win32.Dispatch("Outlook.Application")
            time.sleep(3)
            
        mail = outlook.CreateItem(0)
        
        # 수신인 지정 (테스트 모드 분기)
        if TEST_MODE:
            mail.To = RECIPIENT
            mail.Subject = f"[TEST MODE] {today_str} // 코벤3호 Daily NAV 및 수익률 보고"
            print(f"[*] [TEST_MODE] 활성화 상태: 임시 수신인({RECIPIENT})으로 설정됩니다.")
        else:
            mail.To = "; ".join(RECIPIENTS)
            mail.Subject = f"{today_str} // 코벤3호 Daily NAV 및 수익률 보고"
            print(f"[*] 실제 이메일 발송 모드: 수신인({', '.join(RECIPIENTS)})으로 설정됩니다.")
            
        mail.HTMLBody = html_body
        
        # 발송 또는 드래프트 디스플레이
        if TEST_MODE:
            # 드래프트 창 띄우기 (실제 발송하지 않고 사용자가 확인 가능하게 함)
            mail.Display()
            print("[*] Outlook 드래프트 창을 정상적으로 표시했습니다. 메일을 확인하십시오.")
        else:
            mail.Send()
            print("[*] 메일을 성공적으로 발송했습니다.")
            
    except Exception as e:
        print(f"[⚠️ 에러] 보고서 작성 및 메일 발송 처리 중 오류 발생: {e}")

# ════════════════════════════════════════════════════════
# 4. 스케줄러 설정
# ════════════════════════════════════════════════════════
def start_scheduler():
    """
    매일 오후 5시 30분(17:30)에 코벤3호 Daily NAV 및 수익률 보고서를 
    자동으로 생성하고 메일을 발송하는 스케줄러를 구동합니다.
    """
    print("[시작] 스케줄러 구동 및 설정을 시작합니다.")
    scheduler = BlockingScheduler()
    
    # 매일 오후 5:30(17:30)에 실행
    scheduler.add_job(run_report, "cron", hour="17", minute="30", id="koven3_daily_report")
    
    print("\n" + "=" * 55)
    print("  📧 코벤3호 일일 NAV & 수익률 자동 이메일 보고 스케줄러")
    print("  └─ 매일 오후 5:30 (17:30) 자동 실행 및 메일 발송")
    print("  └─ 설정 모드: TEST_MODE = " + str(TEST_MODE))
    print("=" * 55)
    
    print("\n등록된 예약 잡(Jobs):")
    for job in scheduler.get_jobs():
        next_run = getattr(job, "next_run_time", None)
        next_run_str = str(next_run) if next_run is not None else "스케줄러 구동 후 결정"
        print(f"  · [{job.id}] {job.name}  →  다음 실행 예정: {next_run_str}")
        
    print("\n스케줄러 작동을 위해 본 창을 계속 켜두십시오. (종료: Ctrl+C)\n")
    try:
        scheduler.start()
    except (KeyboardInterrupt, SystemExit):
        print("\n[!] 스케줄러가 종료되었습니다.")
    print("[종료] 스케줄러 작동이 종료되었습니다.")

if __name__ == "__main__":
    # 초기 1회 즉시 실행 (테스트 용도)
    run_report()
    
    # '--once' 아규먼트가 없는 경우에만 백그라운드 스케줄러 구동
    if "--once" not in sys.argv:
        start_scheduler()


# ════════════════════════════════════════════════════════════════════════════════════════════════════
# 📑 [코벤3호 Daily NAV & 수익률 보고 시스템] 운영 가이드 및 데이터 연동 명세서
# ════════════════════════════════════════════════════════════════════════════════════════════════════
# 
# 1. 시스템 개요
#    - 본 스크립트는 두나미스 코스닥벤처 일반사모투자신탁 3호(코벤3호)의 일일 성과와 기간별 수익률을 산출하여 
#      매일 오후 5시 30분(17:30)에 메일 수신자(대석, 기흥)에게 Outlook을 통해 자동으로 발송/작성합니다.
# 
# 2. 연동 API 명세 (ReactDB API)
#    - 호출 주소: http://43.202.36.216:13070/long-short/account/ (Dunamis Portal Backend)
#    - 데이터 성격: Enfusion(OMS/PMS)의 포지션 실시간 평가 손익 데이터가 캐싱되어 있는 ReactDB의 API입니다.
#    - 집계 대상 계좌: 'venture-3' 문자열이 들어가는 총 17개의 모든 서브북 계좌를 자동으로 빠짐없이 집계합니다.
#      (L/S Stock, L/S Futures, Block Stock, Block Futures, Event Stock, Event Futures 뿐만 아니라
#       ipo-new-stock(신주), ipo-exist-stock(구주), ipo-ipo-stock, macro, post 등 모든 서브북 계좌 포함)
#    - 파싱 데이터: 각 포지션별 'dayToDayPositionalPnl' (전일비 평가손익) 및 'intraDayPnl' (당일 매매손익)의 총합.
# 
# 3. 엑셀 파일 다운로드 및 저장 경로 명세 (HINTS 연동)
#    - 저장 위치: C:\Users\park\PycharmProjects\Automation\resources\longshort (로컬 longshort 디렉토리)
# 
#    ==========================================================================================
#    💡 [타 PC 구동 시 필수 사전 다운로드 및 환경 세팅 가이드]
#    ------------------------------------------------------------------------------------------
#    다른 사용자의 PC에서 본 코드를 처음 구동하거나 연동 엑셀이 유실되었을 경우, HINTS에서 아래 두 가지
#    유형의 엑셀 파일을 수동 다운로드하여 아래 로컬 경로에 반드시 배치해야 기간별 P&L이 정상 산출됩니다.
# 
#    1. 경로 생성: Pycharm 프로젝트 루트 폴더 하위에 아래와 같이 폴더 구조를 만듭니다.
#       └─ [Project Root] / resources / longshort /
# 
#    2. [필수 1] 당일 데일리 기준가 파일 (당일 손익 & YTD & 누적 수익률 계산용 기초자산 매핑)
#       - HINTS 메뉴: [일자별 기준가격 조회] (또는 기준가격표 조회)
#       - 조회 대상일: 당일(T0) 날짜 (T0 아침 SOD 기준, 전일(T-1) 마감 최종 확정 데이터가 적혀있음)
#       - 다운로드 방법: HINTS에서 당일 날짜로 조회하여 엑셀 저장
#       - 파일명 이름변경 규칙: 10304_일자별기준가격 조회_YYYYMMDD.xlsx (반드시 8자리 날짜 포맷)
#         * 예: 2026년 5월 18일에 돌릴 경우 -> 10304_일자별기준가격 조회_20260518.xlsx
#       - 파일 내 필수 데이터: 
#         * 시트명: "22204_기준가격표"
#         * '두나미스 코스닥벤처 일반사모투자신탁 3호' 행의 G열 (전일 AUM), J열 (전일 수정기준가)
# 
#    3. [필수 2] 전월말 EOM 기준가 파일 (당월 MTD 수익률 계산용 기점 데이터)
#       - HINTS 메뉴: [일자별 기준가격 조회]
#       - 조회 대상일: 전월(M-1)의 마지막 영업일 (최종 거래일 마감 확정 데이터)
#       - 다운로드 방법: HINTS에서 전월말일 날짜로 조회하여 엑셀 저장
#       - 파일명 이름변경 규칙: 10304_일자별기준가격 조회_{Month}월.xlsx
#         * 예: 현재 5월 영업 중 MTD 계산 시 -> 10304_일자별기준가격 조회_5월.xlsx
#       - 파일 내 필수 데이터:
#         * 펀드코드 'DM13030' (코벤3호)의 최종 확정 수정기준가 (Unnamed: 9열)
# 
#    ※ [참고] YTD(금년) 및 Cumulative(누적) 수익률은 2026년 시작가(1329.7383)와 최초 설정가(1000.0) 상수가
#      코드 내에 하드코딩되어 있으므로, 별도의 과거 연도/설정일 기준가 파일을 추가로 다운로드하실 필요가 없습니다.
#    ==========================================================================================
# 
# 4. 기간별 수익률 계산 공식 및 상수 매핑
#    - 일일 수익률 (Daily Chg%): 당일 실시간 P&L / 전일 AUM
#    - 당일 수정기준가 (NAV): 전일 NAV * (1 + 일일 수익률)
#    - 당일 순자산 (AUM): 전일 AUM + 당일 실시간 P&L
#    - 당월 수익률 (MTD): 당일 NAV / 전월말 NAV - 1
#    - 금년 수익률 (YTD): 당일 NAV / YTD Base - 1  (※ 2026년 코벤3호 YTD Base 상수: 1329.7383 적용)
#    - 누적 수익률 (Cumulative): (당일 NAV - 초기 Base) / 초기 Base  (※ 초기 Base 설정가 상수: 1000.0 적용)
# 
# 5. 비상시 조치 및 예외 처리 (Robustness)
#    - 파일 잠금 방지: 엑셀 파일이 타인에 의해 열려 있어 발생하는 PermissionError를 막고자, shutil 복사본을 떠서 읽은 뒤 자동 삭제합니다.
#    - 월초 폴백 로직: 월초에 신규 당월 EOM 파일이 생성되지 않았을 경우, 자동으로 이전 달 EOM 엑셀 파일을 찾아 전월말 기준가를 역탐색합니다.
#    - 데일리 폴백 로직: 정규표현식(\d{8})으로 일일 파일들만 엄선하여 역정렬하므로, 오늘 날짜 파일이 누락되더라도 가장 최근 일자의 데일리 기준가 데이터를 안전하게 호출해 냅니다.
# 
# 6. 백그라운드 스케줄러 무중단 자동화 (터미널/PyCharm이 꺼져 있어도 자동 발송하는 법)
#    - 스크립트를 터미널에 상시 켜두지 않더라도, PC만 켜져 있다면 윈도우 OS가 백그라운드에서 지정된 시간에 메일을 발송합니다.
#    - 방법: koven3_daily_report.bat 배치 파일을 아래 PowerShell 커맨드를 사용해 [윈도우 작업 스케줄러]에 등록합니다.
# 
#    [PowerShell 관리자 권한 초고속 등록 커맨드 (복사 후 파워쉘 창에 마우스 우클릭으로 실행)]
#    ------------------------------------------------------------------------------------------
#    # 1) 매일 오후 5:30 자동 발송 등록
#    Register-ScheduledTask -TaskName "Koven3_Daily_Report_1730" -Trigger (New-ScheduledTaskTrigger -Daily -At 5:30PM) -Action (New-ScheduledTaskAction -Execute "c:\Users\park\PycharmProjects\Automation\automate\koven3_daily_report\koven3_daily_report.bat") -Description "코벤3호 일일 성과 보고 자동화 (17:30)" -Force
#    ------------------------------------------------------------------------------------------
# 
#    [PowerShell 관리자 권한 등록 스케줄러 삭제(해제) 커맨드]
#    ------------------------------------------------------------------------------------------
#    # 등록된 작업 스케줄러 항목을 영구적으로 조용하고 안전하게 제거합니다.
#    Unregister-ScheduledTask -TaskName "Koven3_Daily_Report_1730" -Confirm:$false
#    ------------------------------------------------------------------------------------------
# ════════════════════════════════════════════════════════════════════════════════════════════════════
