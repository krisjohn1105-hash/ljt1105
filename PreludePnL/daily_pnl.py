"""
Prelude 일일손익 누적 관리 파일 생성기
======================================

매일 실행하면 Prelude 리포트에서 그날의 손익을 계산해 **하나의 엑셀 파일에 누적**한다.

  * 2026-05-29 까지의 누적손익은 기존 EQSWAP.xlsx 의 Summary 시트에서 가져온다(시드).
  * 2026-06-01 부터는 Prelude_new 의 CSV 로 스왑 / 현물(Cash) / FX / IPO / 현금·이자
    5개 자산군의 일일손익·누적손익을 계산한다.

기존 EQSWAP.xlsx 는 스왑 전용이라 6월 Cash trade 개시 이후 일일손익이 부정확해졌다.
이 파일은 계좌 전체 평가액 변동을 기준으로 하므로 모든 자산군을 빠짐없이 담는다.

사용 예)
    python daily_pnl.py --src "Z:/02.펀드/003.매매보고서 대사/Prelude_new"
    python daily_pnl.py --src ... --rebuild          # 기존 누적분 무시하고 전량 재계산
    python daily_pnl.py --src ... --cutover 2026-06-01 --seed-date 2026-05-29

손익 계산 원리
--------------
    일일손익(자산군) = 평가액(t) - 평가액(t-1) + 해당 자산군 귀속 현금흐름(t)

  * 평가액   : MAC001X 'Market Value / Net Equity (Base)' (USD)
  * 현금흐름 : MAC002TDX 'Net Amt Base' 중 현금원장(Position Type = PB / COLCASH) 행
  * 현금·이자 자산군은 잔여항 -> 환평가손익 + 이자 + 배당 + 수수료가 모인다.
  => Σ 자산군 일일손익 = Δ 총평가액 - 외부 자금이동  (완전 일치, '검증' 시트 참조)

IPO 처리
--------
Prelude 상 IPO 배정주는 '대금 0원 Buy Long'(무상입고)으로 들어오고,
청약대금은 별도 'Wires' 로 빠져나간다. 따라서

    IPO 손익 = (배정주 평가액 변동 + 매도대금) - 청약대금

으로 계산된다. IPO 종목은 EQSWAP.xlsx 의 IPO 시트를 마스터로 삼고,
청약대금 Wire 는 그 시트의 Payment Amount(원화)와 금액을 대조해 찾아낸다.
시트에 없는 무상입고 건은 'IPO 후보' 로 보고만 하고 현물에 남긴다(--ipo-auto 로 포함 가능).

작성: Claude Code / 2026-08
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
from typing import Dict, List, Optional, Sequence, Set, Tuple

import pandas as pd

from prelude_pnl import (  # 같은 폴더의 모듈 재사용
    ASSET_CLASS_TO_BUCKET, CASH_LEDGER_POSITION_TYPES,
    build_activity, build_cash_balance, build_positions,
    index_by_code, load_code, long_path, scan_files, to_num,
)

# ---------------------------------------------------------------------------
# 설정
# ---------------------------------------------------------------------------

DEFAULT_CUTOVER = dt.date(2026, 6, 1)      # 이 날짜부터 Prelude 기반으로 계산
DEFAULT_PRINCIPAL = 10_000_000.0           # 기준가 산출용 원금 (EQSWAP.xlsx 와 동일)
DEFAULT_EQSWAP = "EQSWAP.xlsx"
DEFAULT_IPO_FEE_RATE = 0.01                # 청약대금 = 공모확정가 x 배정수량 x 1.01
DEFAULT_OUTPUT = "Prelude_Daily_PnL.xlsx"
MAX_GAP_DAYS = 5

BUCKETS = ["Swap", "Cash Equity", "FX", "IPO", "Cash"]
BUCKET_EN = {
    "Swap": "Swap",
    "Cash Equity": "Cash Equity",
    "FX": "FX",
    "IPO": "IPO",
    "Cash": "Cash & Interest",
}

DAILY_SHEET = "01_Daily_PnL"
LEGACY_DAILY_SHEET = "01_일일손익"      # 이전 버전(한글) 파일과의 호환용
IPO_WIRE_CATEGORIES = ("Wires",)

# 이전 한글 산출물을 읽어 이어붙일 때 쓰는 컬럼명 매핑
LEGACY_COLUMN_MAP = {
    "기준일": "Report Date", "직전 기준일": "Prior Report Date", "경과일수": "Days Elapsed",
    "스왑": "Swap", "현물": "Cash Equity", "현금·이자·기타": "Cash & Interest",
    "일일손익 합계": "Daily PnL Total", "누적손익": "Cumulative PnL",
    "기준가(달러)": "NAV per Unit (USD)", "AUM(원금+누적손익)": "AUM (Principal + Cum PnL)",
    "일일수익률(%)": "Daily Return (%)", "MS계좌 총평가액": "MS Account Market Value",
    "외부 자금이동": "External Cash Movement", "산출대상": "Computed", "비고": "Note",
    "스왑 누적": "Swap Cum.", "현물 누적": "Cash Equity Cum.", "FX 누적": "FX Cum.",
    "IPO 누적": "IPO Cum.", "현금·이자·기타 누적": "Cash & Interest Cum.",
}

# prelude_pnl 모듈이 만들어 주는 한글 컬럼 -> 영문 (출력 직전에만 적용)
OUTPUT_COLUMN_MAP = {
    "기준일": "Report Date", "통화": "Currency", "매매일": "Trade Date", "결제일": "Settle Date",
    "구분": "Type", "중분류": "Category L2", "소분류": "Category L3", "버킷": "Bucket",
    "포지션유형": "Position Type", "상품유형": "Product Type", "종목명": "Security",
    "종목코드": "Symbol", "매매구분": "Buy/Sell", "수량": "Quantity", "단가_USD": "Price (USD)",
    "약정금액_USD": "Principal (USD)", "정산금액_USD": "Net Amount (USD)",
    "결제통화": "Settle CCY", "정산금액_결제통화": "Net Amount (Settle CCY)",
    "현금원장": "Cash Ledger", "자산군": "Asset Class", "입고일": "Delivery Date",
    "등록여부": "Registered",
    "현재잔고(매매기준)": "Current Balance (Trade Date)",
    "현재잔고(결제기준)": "Current Balance (Settled)",
    "현재잔고_USD": "Current Balance (USD)",
    "현재잔고(결제기준)_USD": "Current Balance Settled (USD)",
    "결제예정 수취(D+1~D+4)": "Incoming Settlements (D+1~D+4)",
    "결제필요현금(D+1~D+4)": "Cash Required for Settlement (D+1~D+4)",
    "결제예정 순액(D+1~D+4)": "Net Settlements (D+1~D+4)",
    "미도래 결제 순액(D+5 이후)": "Net Settlements (D+5 onward)",
    "기준일 결제후잔고": "Balance After Settlement (Report Date)",
    "D+1 예상잔고": "Projected Balance D+1", "D+2 예상잔고": "Projected Balance D+2",
    "D+3 예상잔고": "Projected Balance D+3", "D+4 예상잔고": "Projected Balance D+4",
    "전체 결제후 잔고": "Final Projected Balance",
    "USD환산율(나누기)": "FX Rate to USD (divide)",
    "결제필요현금(D+1~D+4)_USD": "Cash Required for Settlement (USD)",
    "결제예정 순액(D+1~D+4)_USD": "Net Settlements (USD)",
    "전체 결제후 잔고_USD": "Final Projected Balance (USD)",
}

# 셀 값 번역
NOTE_SEED = "Handover baseline from EQSWAP.xlsx (seed)"
NOTE_PRE_CUTOVER = "Before cutover (EQSWAP seed period)"
NOTE_GAP = "Excluded - gap from prior report date too large"
TYPE_SECURITY = "Security"
TYPE_JOURNAL = "Cash flow (no security)"
TYPE_RESIDUAL = "Residual (not attributable per security)"
CASH_RESIDUAL_LABEL = "Cash & Interest (FX revaluation, interest, dividends, fees)"
IPO_PAYMENT_LABEL = "IPO subscription payment"
REGISTERED_YES = "Registered"
REGISTERED_NO = "Not registered"

# 이전 한글 산출물의 비고 값 -> 영문
LEGACY_NOTE_MAP = {
    "EQSWAP.xlsx 인수 기준점(시드)": NOTE_SEED,
    "컷오버 이전(EQSWAP 시드 구간)": NOTE_PRE_CUTOVER,
    "직전 리포트일과 간격이 커서 산출 제외": NOTE_GAP,
}


# 시트를 최신 날짜가 맨 위로 오도록 정렬할 때 기준으로 삼는 컬럼(먼저 걸리는 것 사용)
DATE_SORT_COLUMNS = ("Report Date", "Year-Month", "Delivery Date")


def sort_newest_first(df: pd.DataFrame) -> pd.DataFrame:
    """최신 날짜가 맨 위로 오도록 정렬한다(같은 날짜 안의 순서는 유지)."""
    if df is None or df.empty:
        return df
    for c in DATE_SORT_COLUMNS:
        if c in df.columns:
            return df.sort_values(c, ascending=False, kind="mergesort").reset_index(drop=True)
    return df


def to_english(df: pd.DataFrame) -> pd.DataFrame:
    """출력 직전에 한글 컬럼명을 영문으로 바꾼다."""
    if df is None or df.empty:
        return df
    return df.rename(columns=OUTPUT_COLUMN_MAP)


# ---------------------------------------------------------------------------
# EQSWAP.xlsx 읽기 (시드 누적손익 + IPO 마스터)
# ---------------------------------------------------------------------------

def read_eqswap_summary(path: str) -> pd.DataFrame:
    """EQSWAP.xlsx Summary 시트를 (기준일, 누적손익) 형태로 읽는다."""
    import openpyxl
    wb = openpyxl.load_workbook(long_path(path), read_only=True, data_only=True)
    try:
        ws = wb["Summary"]
        recs = []
        for row in ws.iter_rows(values_only=True):
            d, cum = row[0], row[1]
            if isinstance(d, dt.datetime):
                d = d.date()
            if isinstance(d, dt.date) and isinstance(cum, (int, float)):
                recs.append({
                    "기준일": d,
                    "누적손익": float(cum),
                    "미실현_평가": _f(row[2]), "미실현_대기": _f(row[3]), "미실현_이자": _f(row[4]),
                    "실현_누적": _f(row[5]), "실현_당일": _f(row[6]),
                    "금융비용": _f(row[7]), "IPO": _f(row[8]), "FX": _f(row[9]),
                })
    finally:
        wb.close()
    return pd.DataFrame(recs).sort_values("기준일").reset_index(drop=True)


def read_eqswap_ipo(path: str) -> pd.DataFrame:
    """EQSWAP.xlsx IPO 시트(청약 마스터)를 읽는다."""
    import openpyxl
    wb = openpyxl.load_workbook(long_path(path), read_only=True, data_only=True)
    try:
        if "IPO" not in wb.sheetnames:
            return pd.DataFrame()
        rows = list(wb["IPO"].iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return pd.DataFrame()
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    df = pd.DataFrame(rows[1:], columns=hdr)
    df = df[df["Stock Description"].notna()].copy()
    for c in ("Listing Date", "Trade Date", "Settle Date"):
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce").dt.date
    for c in ("Allocation Shares", "IPO Price", "Payment Amount",
              "Settled Net Amount", "Settled Net Amount ($)", "PnL", "PnL ($)"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df["ticker"] = df.get("ticker", "").astype(str).str.strip()
    return df.reset_index(drop=True)


def _f(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# ---------------------------------------------------------------------------
# IPO 식별
# ---------------------------------------------------------------------------

def _norm_name(s) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def _norm_ticker(s) -> str:
    return str(s).strip().upper().split(".")[0]


def detect_allotments(activity: pd.DataFrame) -> pd.DataFrame:
    """
    Prelude 상 '대금 0원 매수' 로 들어온 배정주를 찾는다.

    IPO 청약 배정주는 주식만 먼저 입고되고(대금 0), 청약대금은 며칠 뒤 Wire 로
    빠져나간다. 따라서 이 행 자체는 무상입고가 아니라 '원가 미확정 입고' 다.
    """
    if activity.empty:
        return pd.DataFrame()
    a = activity[
        activity["포지션유형"].eq("PB")
        & activity["소분류"].isin(["Buy Long", "Buy"])
        & (activity["수량"] > 0)
        & (activity["정산금액_USD"].abs() < 0.005)
        & (activity["정산금액_결제통화"].abs() < 1.0)
    ]
    if a.empty:
        return pd.DataFrame()
    # 같은 종목·같은 수량이 여러 파일에 재등장하는 경우(재부킹)는 한 건으로 본다
    out = (a.groupby(["종목명", "수량"], as_index=False)
           .agg(배정인식일=("기준일", "min"), 배정입력일=("입력일", "min"),
                종목코드=("종목코드", "last")))
    return out.rename(columns={"수량": "배정수량"}).sort_values("배정입력일")


def detect_subscription_wires(activity: pd.DataFrame) -> pd.DataFrame:
    """청약대금 후보 = 원화 출금 Wire."""
    if activity.empty:
        return pd.DataFrame()
    w = activity[activity["소분류"].isin(IPO_WIRE_CATEGORIES)
                 & activity["현금원장"]
                 & (activity["정산금액_결제통화"] < 0)]
    if w.empty:
        return pd.DataFrame()
    return (w.groupby(["입력일", "정산금액_결제통화"], as_index=False)
            .agg(Wire인식일=("기준일", "min"), 대금_USD=("정산금액_USD", "first"),
                 결제통화=("결제통화", "first"))
            .rename(columns={"입력일": "Wire입력일", "정산금액_결제통화": "대금_결제통화"})
            .sort_values("Wire입력일"))


def match_allotment_costs(allotments: pd.DataFrame, wires: pd.DataFrame,
                          fee_rate: float = 0.01,
                          window_before: int = 8, window_after: int = 30,
                          price_tolerance: float = 1e-5) -> pd.DataFrame:
    """
    배정주와 청약대금 Wire 를 짝지어 원가를 확정한다.

    청약대금 = 공모확정가 x 배정수량 x (1 + 수수료율)  이므로
        내재 공모가 = |Wire| / (배정수량 x (1 + 수수료율))
    가 '깔끔한 값'(10원 단위)으로 떨어지는 조합을 찾는다.
    실제 데이터에서 이 값은 오차 없이 정확히 일치한다.
    """
    if allotments.empty:
        return allotments
    out = allotments.copy()
    for c in ("Wire입력일", "Wire인식일", "공모가", "청약대금_결제통화", "청약대금_USD"):
        out[c] = None
    if wires is None or wires.empty:
        out["원가확정"] = False
        return out

    used: Dict[int, Tuple[str, str]] = {}     # wire row -> (종목명, 종목코드)
    for i in out.index:
        qty = float(out.at[i, "배정수량"])
        d = out.at[i, "배정입력일"]
        best = None
        for j in wires.index:
            if j in used:
                continue
            wd = wires.at[j, "Wire입력일"]
            if wd is None or d is None:
                continue
            if not (d - dt.timedelta(days=window_before) <= wd
                    <= d + dt.timedelta(days=window_after)):
                continue
            price = abs(float(wires.at[j, "대금_결제통화"])) / (qty * (1.0 + fee_rate))
            # 공모가는 10원 단위 -> 라운드 오차가 가장 작은 Wire 를 채택
            err = abs(price - round(price / 10.0) * 10.0) / max(price, 1.0)
            if best is None or err < best[0]:
                best = (err, j, price)
        if best is not None and best[0] <= price_tolerance:
            _, j, price = best
            used[j] = (out.at[i, "종목명"], out.at[i, "종목코드"])
            out.at[i, "Wire입력일"] = wires.at[j, "Wire입력일"]
            out.at[i, "Wire인식일"] = wires.at[j, "Wire인식일"]
            out.at[i, "공모가"] = round(price / 10.0) * 10.0
            out.at[i, "청약대금_결제통화"] = abs(float(wires.at[j, "대금_결제통화"]))
            out.at[i, "청약대금_USD"] = abs(float(wires.at[j, "대금_USD"]))
    out["원가확정"] = out["청약대금_USD"].notna()
    out.attrs["matched_wire_rows"] = used
    return out


def apply_manual_costs(allotments: pd.DataFrame, overrides: Sequence[str]) -> pd.DataFrame:
    """--ipo-cost "종목키워드=USD금액" 으로 원가를 수동 지정한다."""
    if allotments.empty or not overrides:
        return allotments
    out = allotments.copy()
    for spec in overrides:
        if "=" not in spec:
            continue
        key, amt = spec.split("=", 1)
        try:
            amt = abs(float(amt.replace(",", "").strip()))
        except ValueError:
            continue
        k = _norm_name(key)
        hit = out["종목명"].apply(lambda s: k in _norm_name(s)) | \
            out["종목코드"].apply(lambda s: k in _norm_ticker(s))
        out.loc[hit, "청약대금_USD"] = amt
        out.loc[hit, "원가확정"] = True
        out.loc[hit, "Wire인식일"] = out.loc[hit, "Wire인식일"].fillna(
            out.loc[hit, "배정인식일"])
    return out


def ipo_security_matcher(allotments: pd.DataFrame):
    """배정 이력이 있는 종목인지 판별하는 함수를 만든다."""
    names = {_norm_name(n) for n in allotments["종목명"]} if len(allotments) else set()
    codes = {_norm_ticker(c) for c in allotments["종목코드"]} if len(allotments) else set()
    codes.discard("")

    def is_ipo(name, code) -> bool:
        return _norm_name(name) in names or _norm_ticker(code) in codes
    return is_ipo


def build_ipo_synthetic(allotments: pd.DataFrame, positions: pd.DataFrame,
                        dates: Sequence[dt.date]) -> Tuple[pd.Series, pd.DataFrame]:
    """
    IPO 청약 회계를 맞추기 위한 합성 평가액을 만든다.

    (1) 청약미지급금 : 배정 인식일 ~ Wire 인식일 직전까지 -원가
        (주식은 들어왔는데 대금은 아직 안 나간 구간)
    (2) 미평가 배정주 : Prelude 가 아직 가격을 안 매긴 배정주(MV=0, 수량≠0)를 원가로 평가

    이렇게 하면
        배정일   손익 = 0                      (자산 +원가, 부채 -원가)
        최초평가일 손익 = 시가 - 원가           (무상입고 전액 인식 X)
        Wire일   손익 = 순수 시가변동          (부채 소멸 +원가, 현금 -원가)
    이 되어 실제 청약 손익과 일치한다.
    """
    idx = pd.Index(dates, name="Report Date")
    synth = pd.Series(0.0, index=idx)
    rows = []
    if allotments.empty:
        return synth, pd.DataFrame()

    for _, a in allotments.iterrows():
        cost = a.get("청약대금_USD")
        if cost is None or pd.isna(cost):
            continue
        cost = float(cost)
        d0 = a["배정인식일"]
        wd = a.get("Wire인식일") or d0
        qty0 = float(a["배정수량"])

        # (1) 청약미지급금
        payable = pd.Series(0.0, index=idx)
        payable[(idx >= d0) & (idx < wd)] = -cost

        # (2) 미평가 배정주를 원가로 평가 (잔여수량 비례)
        held = positions[(positions["종목명"] == a["종목명"])
                         & (positions["기준일"] >= d0)]
        atcost = pd.Series(0.0, index=idx)
        if len(held):
            g = held.groupby("기준일").agg(qty=("수량", "sum"), mv=("평가액_USD", "sum"))
            for d, r in g.iterrows():
                if d in atcost.index and r["mv"] == 0 and r["qty"] != 0 and qty0:
                    atcost[d] = cost * (r["qty"] / qty0)

        synth = synth.add(payable, fill_value=0.0).add(atcost, fill_value=0.0)
        for d in idx:
            if payable[d] or atcost[d]:
                rows.append({"Report Date": d, "Security": a["종목명"],
                             "Symbol": a["종목코드"],
                             "Unpaid Subscription": payable[d],
                             "Allotment at Cost": atcost[d],
                             "Total Adjustment": payable[d] + atcost[d]})
    return synth, pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# 자산군 재분류 (IPO 분리)
# ---------------------------------------------------------------------------

def apply_ipo_buckets(positions: pd.DataFrame, activity: pd.DataFrame,
                      allotments: pd.DataFrame,
                      ipo_wire_map: Dict[int, Tuple[str, str]]
                      ) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    배정 이력이 있는 종목과 청약대금 Wire 를 IPO 자산군으로 옮긴다.
    청약대금 Wire 는 해당 배정 종목에 귀속시켜(종목명 부여) 종목별 손익이 맞게 한다.
    """
    pos = positions.copy()
    act = activity.copy()
    is_ipo = ipo_security_matcher(allotments)

    if not pos.empty:
        mask = (pos["버킷"] == "Cash Equity") & pos.apply(
            lambda r: is_ipo(r["종목명"], r["종목코드"]), axis=1)
        pos.loc[mask, "버킷"] = "IPO"

    if not act.empty:
        mask = act["버킷"].eq("Cash Equity") & act.apply(
            lambda r: is_ipo(r["종목명"], r["종목코드"]), axis=1)
        act.loc[mask, "버킷"] = "IPO"
        for idx, (nm, code) in (ipo_wire_map or {}).items():
            if idx in act.index:
                act.at[idx, "버킷"] = "IPO"
                act.at[idx, "종목명"] = nm
                act.at[idx, "종목코드"] = code
    return pos, act


# ---------------------------------------------------------------------------
# 일일손익 계산
# ---------------------------------------------------------------------------

def compute_daily(positions: pd.DataFrame, activity: pd.DataFrame,
                  external_index: Set[int], cutover: dt.date,
                  max_gap_days: int = MAX_GAP_DAYS,
                  synthetic_ipo: Optional[pd.Series] = None) -> Dict[str, pd.DataFrame]:
    dates = sorted(positions["기준일"].unique())
    if not dates:
        raise SystemExit("MAC001X 포지션 데이터가 없습니다.")

    mv = (positions.groupby(["기준일", "버킷"], as_index=False)["평가액_USD"].sum()
          .pivot(index="기준일", columns="버킷", values="평가액_USD")
          .reindex(dates).fillna(0.0))
    for b in BUCKETS:
        if b not in mv.columns:
            mv[b] = 0.0
    mv = mv[BUCKETS + [c for c in mv.columns if c not in BUCKETS]]

    raw_total_mv = mv.sum(axis=1)
    # IPO 청약미지급금 / 미평가 배정주 원가 반영 (Prelude 장부에는 없는 항목)
    synth = (pd.Series(0.0, index=mv.index) if synthetic_ipo is None
             else synthetic_ipo.reindex(mv.index).fillna(0.0))
    mv["IPO"] = mv["IPO"] + synth

    cash_rows = activity[activity["현금원장"]] if not activity.empty else pd.DataFrame()
    if len(cash_rows):
        ext_mask = cash_rows.index.isin(external_index)
        ext = (cash_rows[ext_mask].groupby("기준일")["정산금액_USD"].sum()
               .reindex(mv.index).fillna(0.0))
        flow = (cash_rows[~ext_mask].groupby(["기준일", "버킷"], as_index=False)["정산금액_USD"].sum()
                .pivot(index="기준일", columns="버킷", values="정산금액_USD")
                .reindex(mv.index).fillna(0.0))
    else:
        ext = pd.Series(0.0, index=mv.index)
        flow = pd.DataFrame(0.0, index=mv.index, columns=mv.columns)
    for c in mv.columns:
        if c not in flow.columns:
            flow[c] = 0.0
    flow = flow[mv.columns]

    idx = list(mv.index)
    prev_date = pd.Series([None] + idx[:-1], index=idx)
    gap = pd.Series([float("nan")] + [(idx[i] - idx[i - 1]).days for i in range(1, len(idx))], index=idx)
    computable = pd.Series([d >= cutover and (i > 0 and gap.iloc[i] <= max_gap_days)
                            for i, d in enumerate(idx)], index=idx)

    first = next((d for d in idx if computable[d]), None)
    d_mv = mv.diff()
    pnl = pd.DataFrame(0.0, index=mv.index, columns=mv.columns)
    others = [c for c in mv.columns if c != "Cash"]
    for b in others:
        pnl[b] = d_mv[b] + flow[b]
    pnl["Cash"] = d_mv["Cash"] - flow[others].sum(axis=1) - ext
    pnl = pnl.where(computable, 0.0)

    note = pd.Series("", index=idx)
    note[[d for d in idx if d < cutover]] = NOTE_PRE_CUTOVER
    note[[d for d in idx if d >= cutover and not computable[d]]] = NOTE_GAP

    total_mv = mv.sum(axis=1)
    daily = pd.DataFrame(index=mv.index)
    daily.index.name = "Report Date"
    daily["Prior Report Date"] = prev_date
    daily["Days Elapsed"] = gap
    for b in BUCKETS:
        daily[BUCKET_EN[b]] = pnl[b]
    daily["Daily PnL Total"] = pnl[BUCKETS].sum(axis=1)
    daily["MS Account Market Value"] = raw_total_mv
    daily["IPO Subscription Adjustment"] = synth
    daily["Total Market Value (adj.)"] = total_mv
    daily["External Cash Movement"] = ext.where(computable, 0.0)
    daily["Computed"] = computable
    daily["Note"] = note

    detail = []
    for i, d in enumerate(idx):
        for b in BUCKETS:
            detail.append({
                "Report Date": d, "Asset Class": BUCKET_EN[b],
                "Prior Market Value": mv[b].shift().iloc[i] if i else float("nan"),
                "Market Value": mv[b].iloc[i],
                "MV Change": d_mv[b].iloc[i],
                "Cash Flow": flow[b].iloc[i],
                "Daily PnL": pnl[b].iloc[i],
                "Computed": bool(computable.iloc[i]),
            })

    recon = pd.DataFrame(index=mv.index)
    recon.index.name = "Report Date"
    recon["Computed"] = computable
    recon["MS Account Market Value"] = raw_total_mv
    recon["IPO Subscription Adjustment"] = synth
    recon["Total MV Change"] = total_mv.diff().where(computable)
    recon["External Cash Movement"] = ext.where(computable, 0.0)
    recon["Sum of Asset Class PnL"] = pnl[BUCKETS].sum(axis=1)
    recon["Difference (Check)"] = (recon["Total MV Change"]
                                   - recon["External Cash Movement"]
                                   - recon["Sum of Asset Class PnL"])
    recon["Note"] = note

    return {"daily": daily.reset_index(), "detail": pd.DataFrame(detail),
            "recon": recon.reset_index(), "mv": mv, "flow": flow, "pnl": pnl,
            "computable": computable}


def compute_security_pnl(positions: pd.DataFrame, activity: pd.DataFrame,
                         computable: pd.Series,
                         synth_detail: Optional[pd.DataFrame] = None,
                         bucket_pnl: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    """
    자산군·종목 단위 일일손익.

    현금·이자 자산군은 정의상 잔여항(Δ현금평가액 − 타 자산군 현금흐름)이라
    종목 단위로 분해되지 않는다. 종목별로 쪼개면 버킷 합계와 어긋나므로
    한 줄의 잔여항 행으로 표시한다.
    """
    pos = positions.copy()
    act = activity[activity["현금원장"]].copy() if not activity.empty else pd.DataFrame()
    key_cols = ["버킷", "종목명"]
    mv = (pos.groupby(["기준일"] + key_cols, as_index=False)
          .agg(평가액=("평가액_USD", "sum"), 수량=("수량", "sum"),
               종목코드=("종목코드", "last"))
          if not pos.empty else pd.DataFrame())
    fl = (act.groupby(["기준일"] + key_cols, as_index=False)
          .agg(현금흐름=("정산금액_USD", "sum"))
          if len(act) else pd.DataFrame(columns=["기준일"] + key_cols + ["현금흐름"]))

    dates = sorted(positions["기준일"].unique())
    keys = pd.concat([mv[key_cols] if len(mv) else pd.DataFrame(columns=key_cols),
                      fl[key_cols] if len(fl) else pd.DataFrame(columns=key_cols)]
                     ).drop_duplicates()
    if keys.empty:
        return pd.DataFrame()
    grid = keys.merge(pd.DataFrame({"기준일": dates}), how="cross")
    out = grid.merge(mv, on=["기준일"] + key_cols, how="left") \
              .merge(fl, on=["기준일"] + key_cols, how="left")
    for c in ("평가액", "현금흐름", "수량"):
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    # IPO 청약 합성 조정(미지급금 / 미평가 배정주 원가)을 종목 평가액에 반영
    if synth_detail is not None and len(synth_detail):
        adj = (synth_detail.rename(columns={"Report Date": "기준일", "Security": "종목명",
                                            "Total Adjustment": "_adj"})
               .groupby(["기준일", "종목명"], as_index=False)["_adj"].sum())
        out = out.merge(adj, on=["기준일", "종목명"], how="left")
        out["_adj"] = out["_adj"].fillna(0.0)
        out["평가액"] = out["평가액"] + out["_adj"]
        out = out.drop(columns="_adj")

    out = out.sort_values(key_cols + ["기준일"])
    out["전일평가액"] = out.groupby(key_cols)["평가액"].shift()
    out["일일손익"] = (out["평가액"] - out["전일평가액"].fillna(out["평가액"])) + out["현금흐름"]
    bad = set(computable.index[~computable.astype(bool)])
    out.loc[out["기준일"].isin(bad), "일일손익"] = 0.0
    out["누적손익"] = out.groupby(key_cols)["일일손익"].cumsum()
    out["자산군"] = out["버킷"].map(BUCKET_EN).fillna(out["버킷"])
    # 종목이 붙지 않는 현금원장 저널(스왑 리셋/파이낸싱 정산, 담보이체 등) 구분.
    # 포지션 리포트에 한 번도 등장하지 않은 이름만 저널로 본다.
    held_names = (set(pos.loc[pos["상품유형"].astype(str).str.upper() != "CASH", "종목명"])
                  if not pos.empty else set())
    out["구분"] = out["종목명"].apply(
        lambda n: TYPE_SECURITY if n in held_names else TYPE_JOURNAL)
    out = out[(out["평가액"] != 0) | (out["현금흐름"] != 0) | (out["일일손익"] != 0)]
    cols = ["기준일", "자산군", "구분", "종목명", "종목코드", "수량",
            "전일평가액", "평가액", "현금흐름", "일일손익", "누적손익"]
    out = out[cols]

    # 현금·이자 버킷은 종목 분해분을 버리고 잔여항 한 줄로 대체한다
    out = out[out["자산군"] != BUCKET_EN["Cash"]]
    if bucket_pnl is not None and "Cash" in bucket_pnl.columns:
        res = pd.DataFrame({
            "기준일": list(bucket_pnl.index),
            "자산군": BUCKET_EN["Cash"],
            "구분": TYPE_RESIDUAL,
            "종목명": CASH_RESIDUAL_LABEL,
            "종목코드": "",
            "수량": 0.0, "전일평가액": float("nan"), "평가액": 0.0, "현금흐름": 0.0,
            "일일손익": bucket_pnl["Cash"].values,
        })
        res["누적손익"] = res["일일손익"].cumsum()
        res = res[res["일일손익"] != 0]
        out = pd.concat([out, res[cols]], ignore_index=True, sort=False)

    out = out.sort_values(["기준일", "자산군", "구분", "종목명"])
    return out.rename(columns={
        "기준일": "Report Date", "자산군": "Asset Class", "구분": "Type",
        "종목명": "Security", "종목코드": "Symbol", "수량": "Quantity",
        "전일평가액": "Prior Market Value", "평가액": "Market Value",
        "현금흐름": "Cash Flow", "일일손익": "Daily PnL", "누적손익": "Cumulative PnL"})


# ---------------------------------------------------------------------------
# 누적 병합 (기존 파일 + 신규 계산분)
# ---------------------------------------------------------------------------

def merge_history(new_rows: pd.DataFrame, out_path: str, rebuild: bool) -> pd.DataFrame:
    """기존 산출 파일의 일일손익을 읽어 신규 계산분과 병합한다(같은 날짜는 신규가 우선)."""
    if rebuild or not os.path.exists(long_path(out_path)):
        return new_rows.copy()
    old = None
    for sheet in (DAILY_SHEET, LEGACY_DAILY_SHEET):
        try:
            old = pd.read_excel(out_path, sheet_name=sheet)
            break
        except Exception:
            continue
    if old is None:
        print("  ! 기존 파일에서 일일손익 시트를 찾지 못했습니다 - 신규 계산분만 사용합니다.",
              file=sys.stderr)
        return new_rows.copy()
    old = old.rename(columns=LEGACY_COLUMN_MAP)   # 이전 한글 산출물 호환
    if "Note" in old.columns:
        old["Note"] = old["Note"].replace(LEGACY_NOTE_MAP)
    if old.empty or "Report Date" not in old.columns:
        return new_rows.copy()
    old["Report Date"] = pd.to_datetime(old["Report Date"], errors="coerce").dt.date
    old = old[old["Report Date"].notna()]

    # 새로 계산했지만 산출 불가(직전 기준일 없음 등)인 날은 기존 값을 덮어쓰지 않는다.
    # 과거 원본을 다른 곳으로 옮겨도 이미 쌓아둔 손익이 0 으로 지워지지 않게 하는 안전장치.
    if "Computed" in new_rows.columns:
        usable = new_rows[new_rows["Computed"].astype(bool)]
        dropped = new_rows[~new_rows["Computed"].astype(bool)]
        overwritten = set(usable["Report Date"])
        readd = dropped[~dropped["Report Date"].isin(set(old["Report Date"]))]
        new_rows = pd.concat([usable, readd], ignore_index=True, sort=False)
    else:
        overwritten = set(new_rows["Report Date"])

    keep = old[~old["Report Date"].isin(overwritten | set(new_rows["Report Date"]))]
    cols = [c for c in new_rows.columns if c in keep.columns]
    merged = pd.concat([keep[cols], new_rows], ignore_index=True, sort=False)
    merged = merged.drop_duplicates(subset="Report Date", keep="last")
    return merged.sort_values("Report Date").reset_index(drop=True)


def add_cumulative(daily: pd.DataFrame, seed_cum: float, seed_date: Optional[dt.date],
                   principal: float) -> pd.DataFrame:
    d = daily.sort_values("Report Date").reset_index(drop=True).copy()
    # 인수인계 기준점(EQSWAP 시드)을 첫 행으로 넣어 누적 추이가 이어지게 한다
    if seed_date is not None and seed_date not in set(d["Report Date"]):
        seed_row = {c: 0.0 for c in d.columns if d[c].dtype.kind in "if"}
        seed_row.update({"Report Date": seed_date, "Prior Report Date": pd.NaT,
                         "Days Elapsed": float("nan"), "Computed": False,
                         "Note": NOTE_SEED})
        d = pd.concat([pd.DataFrame([seed_row]), d], ignore_index=True, sort=False)
        d = d.sort_values("Report Date").reset_index(drop=True)
    if seed_date is not None and "Note" in d.columns:
        d.loc[d["Report Date"] == seed_date, "Note"] = NOTE_SEED
    d["Cumulative PnL"] = seed_cum + d["Daily PnL Total"].fillna(0).cumsum()
    for b in BUCKETS:
        en = BUCKET_EN[b]
        if en in d.columns:
            d[f"{en} Cum."] = d[en].fillna(0).cumsum()
    d["NAV per Unit (USD)"] = (principal + d["Cumulative PnL"]) / principal
    d["AUM (Principal + Cum PnL)"] = principal + d["Cumulative PnL"]
    d["Daily Return (%)"] = (d["Daily PnL Total"] /
                             (principal + d["Cumulative PnL"].shift().fillna(seed_cum))
                             * 100).astype(float)
    if seed_date is not None:
        d.loc[d["Report Date"] <= seed_date, "Daily Return (%)"] = float("nan")
    # 기준일 → 누적손익 → 일일손익 순으로 맨 앞에 배치(열자마자 바로 읽히도록)
    order = (["Report Date", "Cumulative PnL", "Daily PnL Total",
              "NAV per Unit (USD)", "AUM (Principal + Cum PnL)", "Daily Return (%)"]
             + [BUCKET_EN[b] for b in BUCKETS]
             + [f"{BUCKET_EN[b]} Cum." for b in BUCKETS]
             + ["MS Account Market Value", "IPO Subscription Adjustment",
                "Total Market Value (adj.)", "External Cash Movement",
                "Prior Report Date", "Days Elapsed", "Computed", "Note"])
    return d[[c for c in order if c in d.columns] +
             [c for c in d.columns if c not in order]]


# ---------------------------------------------------------------------------
# 엑셀 출력
# ---------------------------------------------------------------------------

def write_workbook(path: str, sheets: List[Tuple[str, pd.DataFrame]],
                   meta: List[Tuple[str, str]]):
    os.makedirs(long_path(os.path.dirname(os.path.abspath(path))), exist_ok=True)
    with pd.ExcelWriter(path, engine="xlsxwriter",
                        datetime_format="yyyy-mm-dd", date_format="yyyy-mm-dd") as xw:
        wb = xw.book
        hdr = wb.add_format({"bold": True, "bg_color": "#1F3864", "font_color": "white",
                             "border": 1, "align": "center", "valign": "vcenter",
                             "text_wrap": True})
        title = wb.add_format({"bold": True, "font_size": 14})
        label = wb.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
        value = wb.add_format({"border": 1})
        money = wb.add_format({"num_format": "#,##0.00"})
        px = wb.add_format({"num_format": "0.00000000"})

        ws = wb.add_worksheet("00_Summary")
        xw.sheets["00_Summary"] = ws
        ws.set_column(0, 0, 38); ws.set_column(1, 1, 72)
        ws.write(0, 0, "Prelude Daily PnL Tracker", title)
        for r, (k, v) in enumerate(meta, start=2):
            ws.write(r, 0, k, label); ws.write(r, 1, v, value)

        for name, df in sheets:
            sn = name[:31]
            if df is None or len(df) == 0:
                w = wb.add_worksheet(sn); xw.sheets[sn] = w
                w.write(0, 0, "No data")
                continue
            df = sort_newest_first(df)      # 최신 날짜가 맨 위
            df.to_excel(xw, sheet_name=sn, index=False)
            w = xw.sheets[sn]
            for i, c in enumerate(df.columns):
                w.write(0, i, str(c), hdr)
                try:
                    width = max(len(str(c)) + 2,
                                int(df[c].astype(str).str.len().quantile(0.95)) + 2)
                except Exception:
                    width = len(str(c)) + 2
                width = min(max(width, 11), 40)
                name = str(c)
                numeric = df[c].dtype.kind in "if"
                if not numeric:
                    w.set_column(i, i, width)
                elif "NAV per Unit" in name:
                    w.set_column(i, i, width, px)
                elif any(k in name for k in ("PnL", "Market Value", "Amount", "Balance",
                                             "Cash", "AUM", "Cum", "Change", "Movement",
                                             "Settlement", "Principal", "Price", "Flow")):
                    w.set_column(i, i, width, money)
                else:
                    w.set_column(i, i, width)
            # 기준일 열까지는 고정해 두어 오른쪽으로 스크롤해도 날짜가 보이게 한다
            cols = list(df.columns)
            frozen = (cols.index("Report Date") + 1) if "Report Date" in cols else 1
            w.freeze_panes(1, min(frozen, 4))
            w.autofilter(0, 0, len(df), len(df.columns) - 1)

            if sn == DAILY_SHEET and len(df) > 2:
                cols = list(df.columns)
                if "Cumulative PnL" in cols and "Report Date" in cols:
                    ci, cc = cols.index("Report Date"), cols.index("Cumulative PnL")
                    ch = wb.add_chart({"type": "line"})
                    ch.add_series({"name": "Cumulative PnL (USD)",
                                   "categories": [sn, 1, ci, len(df), ci],
                                   "values": [sn, 1, cc, len(df), cc],
                                   "line": {"color": "#1F3864", "width": 2.0}})
                    ch.set_title({"name": "Cumulative PnL (EQSWAP seed + Prelude calculation)"})
                    # 시트는 최신순이지만 차트는 시간 순(과거→최근)으로 그린다
                    ch.set_x_axis({"reverse": True})
                    ch.set_y_axis({"num_format": "#,##0", "crossing": "max"})
                    ch.set_size({"width": 900, "height": 380})
                    ch.set_legend({"position": "bottom"})
                    ws.insert_chart(f"A{len(meta) + 5}", ch)


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------

def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Prelude 리포트 기반 일일손익/누적손익 누적 관리 엑셀 생성",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--src", required=True, help="Prelude 리포트 폴더 (하위 폴더 포함)")
    ap.add_argument("--out", default=None,
                    help=f"산출 엑셀 (기본: <src>/_output/{DEFAULT_OUTPUT})")
    ap.add_argument("--eqswap", default=None,
                    help=f"기존 스왑 관리 파일 (기본: <src>/{DEFAULT_EQSWAP})")
    ap.add_argument("--cutover", default=DEFAULT_CUTOVER.isoformat(),
                    help=f"Prelude 기반 계산 시작일 (기본 {DEFAULT_CUTOVER})")
    ap.add_argument("--seed-date", default=None,
                    help="EQSWAP 누적손익을 가져올 기준일 (기본: 컷오버 직전 영업일)")
    ap.add_argument("--seed-value", type=float, default=None,
                    help="시드 누적손익을 직접 지정 (EQSWAP 파일 대신)")
    ap.add_argument("--principal", type=float, default=DEFAULT_PRINCIPAL,
                    help=f"기준가 산출 원금 (기본 {DEFAULT_PRINCIPAL:,.0f})")
    ap.add_argument("--ipo-fee-rate", type=float, default=DEFAULT_IPO_FEE_RATE,
                    help=f"청약 수수료율. 청약대금 = 공모가 x 수량 x (1+수수료율) "
                         f"(기본 {DEFAULT_IPO_FEE_RATE:.1%})")
    ap.add_argument("--ipo-cost", action="append", default=[], metavar="종목=USD금액",
                    help="청약대금 Wire 가 아직 없는 배정주의 원가를 수동 지정 "
                         "(예: --ipo-cost \"NH SPECIAL=59000\")")
    ap.add_argument("--rebuild", action="store_true",
                    help="기존 산출 파일을 무시하고 전체 재계산")
    ap.add_argument("--max-gap-days", type=int, default=MAX_GAP_DAYS,
                    help=f"직전 리포트일과 간격이 이 일수를 넘으면 산출 제외 (기본 {MAX_GAP_DAYS})")
    args = ap.parse_args(argv)

    for s in (sys.stdout, sys.stderr):
        try:
            s.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    src = os.path.abspath(args.src)
    if not os.path.isdir(src):
        raise SystemExit(f"폴더를 찾을 수 없습니다: {src}")
    cutover = dt.date.fromisoformat(args.cutover)
    out_path = args.out or os.path.join(src, "_output", DEFAULT_OUTPUT)
    eqswap_path = args.eqswap or os.path.join(src, DEFAULT_EQSWAP)

    print(f"[1/7] 리포트 스캔: {src}")
    files = scan_files(src)
    by_code = index_by_code(files)
    print(f"      {len(files):,}개 파일")

    print("[2/7] 리포트 로드")
    positions = build_positions(load_code(by_code, "MAC001X"))
    activity = build_activity(load_code(by_code, "MAC002TDX"))
    raw_cash = load_code(by_code, "CASH005X")
    if positions.empty:
        raise SystemExit("MAC001X 파일이 없어 계산할 수 없습니다.")

    print("[3/7] EQSWAP.xlsx 시드 읽기")
    seed_cum, seed_date, ipo_master = 0.0, None, pd.DataFrame()
    if args.seed_value is not None:
        seed_cum = args.seed_value
        seed_date = dt.date.fromisoformat(args.seed_date) if args.seed_date else None
        print(f"      시드 직접 지정: {seed_cum:,.2f}")
    elif os.path.exists(long_path(eqswap_path)):
        summary = read_eqswap_summary(eqswap_path)
        ipo_master = read_eqswap_ipo(eqswap_path)
        if args.seed_date:
            want = dt.date.fromisoformat(args.seed_date)
            cand = summary[summary["기준일"] == want]
        else:
            cand = summary[summary["기준일"] < cutover]
        if cand.empty:
            raise SystemExit(f"EQSWAP.xlsx Summary 에서 시드 행을 찾지 못했습니다 ({args.seed_date or f'< {cutover}'})")
        row = cand.iloc[-1]
        seed_cum, seed_date = float(row["누적손익"]), row["기준일"]
        print(f"      시드: {seed_date} 누적손익 {seed_cum:,.2f} USD "
              f"(IPO 마스터 {len(ipo_master)}건)")
    else:
        print(f"      ! EQSWAP.xlsx 없음({eqswap_path}) - 시드 0 으로 진행")

    print("[4/7] IPO 배정·청약대금 매칭")
    allot = detect_allotments(activity)
    wires = detect_subscription_wires(activity)
    allot = match_allotment_costs(allot, wires, fee_rate=args.ipo_fee_rate)
    allot = apply_manual_costs(allot, args.ipo_cost)
    matched_wire_rows = allot.attrs.get("matched_wire_rows", {})

    ipo_wire_map: Dict[int, Tuple[str, str]] = {}
    external_idx: Set[int] = set()
    if len(wires):
        for j in wires.index:
            orig = activity.index[
                (activity["입력일"] == wires.at[j, "Wire입력일"])
                & (activity["정산금액_결제통화"] == wires.at[j, "대금_결제통화"])
                & activity["소분류"].isin(IPO_WIRE_CATEGORIES)]
            if j in matched_wire_rows:
                for k in orig:
                    ipo_wire_map[k] = matched_wire_rows[j]
            else:
                external_idx.update(orig)

    ok = allot[allot["원가확정"]] if len(allot) else pd.DataFrame()
    ng = allot[~allot["원가확정"]] if len(allot) else pd.DataFrame()
    print(f"      배정주 {len(allot)}건 중 원가 확정 {len(ok)}건 "
          f"(청약대금 Wire 매칭, 수수료율 {args.ipo_fee_rate:.1%})")
    for _, r in ok[ok["배정인식일"] >= cutover].iterrows() if len(ok) else []:
        print(f"      · {r['종목명'][:30]:30s} {r['배정수량']:>8,.0f}주 "
              f"공모가 {r['공모가']:>9,.0f} 원가 {r['청약대금_USD']:>11,.2f} USD "
              f"(배정 {r['배정인식일']} / 납입 {r['Wire인식일']})")
    for _, r in ng.iterrows() if len(ng) else []:
        print(f"      ! 원가 미확정: {r['종목명'][:30]} {r['배정수량']:,.0f}주 "
              f"(배정 {r['배정인식일']}) → 청약대금 Wire 미도착. "
              f"--ipo-cost \"종목=USD금액\" 으로 지정 가능")
    if external_idx:
        print(f"      외부 자금이동(청약 무관 Wire) {len(external_idx)}건")

    positions, activity = apply_ipo_buckets(positions, activity, allot, ipo_wire_map)
    all_dates = sorted(positions["기준일"].unique())
    synth_ipo, synth_detail = build_ipo_synthetic(allot, positions, all_dates)

    print("[5/7] 일일손익 계산")
    res = compute_daily(positions, activity, external_idx, cutover,
                        max_gap_days=args.max_gap_days,
                        synthetic_ipo=synth_ipo)
    new_daily = res["daily"][res["daily"]["Report Date"] >= cutover].copy()
    print(f"      계산 구간: {new_daily['Report Date'].min()} ~ {new_daily['Report Date'].max()} "
          f"({int(new_daily['Computed'].sum())}영업일)")

    print("[6/7] 기존 누적분 병합")
    merged = merge_history(new_daily, out_path, args.rebuild)
    daily = add_cumulative(merged, seed_cum, seed_date, args.principal)
    last = daily.iloc[-1]

    sec = compute_security_pnl(positions, activity, res["computable"], synth_detail,
                               bucket_pnl=res["pnl"])
    sec = sec[sec["Report Date"] >= cutover]
    detail = res["detail"][res["detail"]["Report Date"] >= cutover]
    recon = res["recon"][res["recon"]["Report Date"] >= cutover]
    cash_bal = to_english(build_cash_balance(positions, raw_cash))
    if len(cash_bal):
        cash_bal = cash_bal[cash_bal["Report Date"] >= cutover]

    ipo_rows = pd.DataFrame()
    if len(sec):
        ipo_rows = sec[sec["Asset Class"] == "IPO"].copy()
    ipo_sheet = pd.DataFrame()
    if len(ipo_master):
        ipo_sheet = ipo_master[[c for c in ["Trade Date", "Settle Date", "Stock Description",
                                            "ticker", "Allocation Shares", "IPO Price",
                                            "Payment Amount", "Settled Net Amount ($)",
                                            "PnL ($)"] if c in ipo_master.columns]].copy()

    trades = pd.DataFrame()
    if not activity.empty:
        tcols = ["기준일", "매매일", "결제일", "구분", "중분류", "소분류", "버킷", "포지션유형",
                 "상품유형", "종목명", "종목코드", "매매구분", "수량", "단가_USD",
                 "약정금액_USD", "정산금액_USD", "결제통화", "정산금액_결제통화", "현금원장"]
        trades = activity[[c for c in tcols if c in activity.columns]].copy()
        trades["자산군"] = trades["버킷"].map(BUCKET_EN).fillna(trades["버킷"])
        trades = (trades[trades["기준일"] >= cutover]
                  .sort_values(["기준일", "자산군", "종목명"])
                  .drop(columns=["버킷"]))
        trades = to_english(trades)
        trades = trades[["Report Date", "Asset Class"]
                        + [c for c in trades.columns if c not in ("Report Date", "Asset Class")]]

    allot_sheet = pd.DataFrame()
    if len(allot):
        allot_sheet = allot.rename(columns={
            "종목명": "Security", "종목코드": "Symbol", "배정수량": "Allotted Qty",
            "배정인식일": "Delivery Date", "배정입력일": "Delivery Entry Date",
            "Wire입력일": "Wire Entry Date", "Wire인식일": "Payment Date",
            "공모가": "Offer Price (KRW)", "청약대금_결제통화": "Subscription (KRW)",
            "청약대금_USD": "Subscription Cost (USD)", "원가확정": "Cost Matched"})
        cols = ["Security", "Symbol", "Allotted Qty", "Delivery Date", "Payment Date",
                "Offer Price (KRW)", "Subscription (KRW)", "Subscription Cost (USD)",
                "Cost Matched", "Delivery Entry Date", "Wire Entry Date"]
        allot_sheet = allot_sheet[[c for c in cols if c in allot_sheet.columns]]
        allot_sheet = allot_sheet.sort_values("Delivery Date")
    if len(synth_detail):
        synth_detail = synth_detail[synth_detail["Report Date"] >= cutover]

    monthly = daily[daily["Report Date"] >= cutover].copy()
    monthly["Year-Month"] = pd.to_datetime(monthly["Report Date"]).dt.strftime("%Y-%m")
    magg = (monthly.groupby("Year-Month", as_index=False)[[BUCKET_EN[b] for b in BUCKETS]
                                                          + ["Daily PnL Total"]].sum())
    mlast = monthly.groupby("Year-Month", as_index=False).last()[
        ["Year-Month", "Cumulative PnL", "NAV per Unit (USD)", "MS Account Market Value"]]
    monthly = (magg.merge(mlast, on="Year-Month", how="left")
               .rename(columns={"Daily PnL Total": "Monthly PnL Total"}))
    # 01_Daily_PnL 과 같은 배치: 기준(연월) → 누적손익 → 손익합계
    m_order = (["Year-Month", "Cumulative PnL", "Monthly PnL Total",
                "NAV per Unit (USD)", "MS Account Market Value"]
               + [BUCKET_EN[b] for b in BUCKETS])
    monthly = monthly[[c for c in m_order if c in monthly.columns]
                      + [c for c in monthly.columns if c not in m_order]]

    meta = [
        ("Generated at", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Source folder", src),
        ("Output file", out_path),
        ("Seed (EQSWAP.xlsx)",
         f"{seed_date}  cumulative PnL {seed_cum:,.2f} USD" if seed_date
         else f"{seed_cum:,.2f} USD"),
        ("Prelude calculation starts", str(cutover)),
        ("Latest report date", str(last["Report Date"])),
        ("Cumulative PnL (latest)", f"{last['Cumulative PnL']:,.2f} USD"),
        ("NAV per unit (USD)", f"{last['NAV per Unit (USD)']:.8f}"),
        ("AUM (principal + cum. PnL)", f"{last['AUM (Principal + Cum PnL)']:,.2f} USD"),
        ("PnL since cutover", f"{last['Cumulative PnL'] - seed_cum:,.2f} USD"),
        ("Cumulative by asset class (since cutover)",
         " | ".join(f"{BUCKET_EN[b]} {last.get(f'{BUCKET_EN[b]} Cum.', 0):,.0f}"
                    for b in BUCKETS)),
        ("IPO securities",
         ", ".join(sorted(ipo_rows["Security"].unique())) if len(ipo_rows) else "None"),
        ("IPO allotments with cost matched",
         f"{len(ok)} of {len(allot)}  (subscription = offer price x qty x "
         f"{1 + args.ipo_fee_rate:.2f})"),
        ("IPO allotments WITHOUT cost (check!)",
         " | ".join(f"{r['종목명']} {r['배정수량']:,.0f}sh @ {r['배정인식일']}"
                    for _, r in ng.iterrows()) if len(ng) else "None"),
        ("PnL formula",
         "Daily PnL = ΔMarket Value + attributed cash flow (Cash & Interest is the residual)"),
        ("IPO accounting",
         "Allotted shares are NOT free deliveries. Cost = offer price x qty x "
         f"{1 + args.ipo_fee_rate:.2f}, paid by wire a few days after delivery. "
         "An unpaid-subscription liability is carried from delivery until the wire "
         "clears, and unpriced allotments are held at cost, so day-1 PnL = "
         "market value - subscription cost."),
        ("Reconciliation",
         "Σ asset classes = Δ total market value (incl. IPO subscription adjustment) "
         "− external cash movement (see 04_Reconciliation)"),
    ]

    sheets = [
        (DAILY_SHEET, daily),
        ("02_Monthly_PnL", monthly),
        ("03_Asset_Class_Detail", detail),
        ("04_Reconciliation", recon),
        ("05_Security_PnL", sec),
        ("06_IPO_Detail", ipo_rows),
        ("07_IPO_Allotment_Cost", allot_sheet),
        ("08_IPO_Subscription_Adj", synth_detail),
        ("09_Cash_Balance", cash_bal),
        ("10_Transactions", trades),
        ("11_IPO_Master_EQSWAP", ipo_sheet),
    ]

    print(f"[7/7] 엑셀 작성: {out_path}")
    write_workbook(out_path, sheets, meta)
    print(f"      완료 · 최종 {last['Report Date']} 누적손익 {last['Cumulative PnL']:,.2f} USD "
          f"/ 기준가 {last['NAV per Unit (USD)']:.6f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
